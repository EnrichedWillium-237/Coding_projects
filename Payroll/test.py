### Code for reading in a payroll spreadsheet and organizing net payments ###
### This is a work in progress and does things in a brute force fashion. Will be improved later. ###

# Source files
import openpyxl
import array as arr
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
import numpy as np

# Payroll file location
workbook = load_workbook('payroll_input.xlsx')

# Payroll output
newbook1 = openpyxl.Workbook()
newsheet1 = newbook1.active

sheet = workbook.active

Nrow = sheet.max_row
Nrow = Nrow - 1 # offset from spreadsheet
Ncell = 1 # index for excel output

# hours per category
hrsUnarmed = 0
hrsArmed = 0
hrsAdmin = 0
hrsOT = 0
hrsTrain = 0
hrsSick = 0
hrsCOVID = 0
hrsTotal = 0

GrandUnarmed = 0
GrandArmed = 0
GrandAdmin = 0
GrandOT = 0
GrandTrain = 0
GrandSick = 0
GrandCovid = 0

Gross = 0 # total pay per employee
GrandHrs = 0 # total number of hours
GrandTot = 0 # total payroll
GrandReim = 0 # total reimbursement
GrandGross = 0 # total payout: total + reimbursement

unknownCat = False

# set up arrays
arrName  = np.empty(500, dtype = 'object')
arrCat   = np.empty(500, dtype = 'object')
arrHrs   = np.empty(500, dtype = 'f')
arrRate  = np.empty(500, dtype = 'f')
arrTot   = np.empty(500, dtype = 'f')
arrReim  = np.empty(500, dtype = 'f')
arrGross = np.empty(500, dtype = 'f')
arrNote  = np.empty(500, dtype = 'object')

# payrate for each category
rate        = np.empty(50, dtype = 'f')
ratehrs     = np.empty(50, dtype = 'f')
rateUnarmed = np.empty(50, dtype = 'f')
rateArmed   = np.empty(50, dtype = 'f')
rateAdmin   = np.empty(50, dtype = 'f')
rateOT      = np.empty(50, dtype = 'f')
rateTrain   = np.empty(50, dtype = 'f')
rateSick    = np.empty(50, dtype = 'f')
rateCovid   = np.empty(50, dtype = 'f')
# rateCovid   = np.empty(50, dtype = 'f') # (KEEP) this one works if the above double arrays do not

# entries per employee Name
arrNumRows = np.empty(500, dtype = 'int')

# fill arrays with Excel spreadsheet values
for i in range(2, Nrow):

    valName = sheet.cell(row = i, column = 1)
    valCat = sheet.cell(row = i, column = 2)
    valHrs = sheet.cell(row = i, column = 3)
    valRate = sheet.cell(row = i, column = 4)
    valTot = sheet.cell(row = i, column = 5)
    valReim = sheet.cell(row = i, column = 6)
    valGross = sheet.cell(row = i, column = 7)
    valNote = sheet.cell(row = i, column = 8)
    valName = valName.value
    valCat = valCat.value
    valHrs = valHrs.value
    valRate = valRate.value
    valTot = valTot.value
    valReim = valReim.value
    valGross = valGross.value
    valNote = valNote.value
    if valCat is None or 0:
        valCat = 0
    if valHrs is None:
        valHrs = 0
    if valRate is None:
        valRate = 0
    if valTot is None:
        valTot = 0
    if valReim is None:
        valReim = 0
    if valGross is None:
        valGross = 0
    if valNote is None or 0:
        valNote = 0
    arrName[i] = valName
    arrCat[i] = valCat
    arrHrs[i] = valHrs
    arrRate[i] = valRate
    arrTot[i] = valTot
    arrReim[i] = valReim
    arrGross[i] = valGross
    arrNote[i] = valNote
    if valName is None or 0:
        break

# determine total rows per name
nRowName = 1
for i in range(2, Nrow):
    if arrName[i] is None or 0:
        break
    if arrName[i+1] is arrName[i]:
        nRowName += 1
        arrNumRows[i] = 0
    if arrName[i+1] is not arrName[i]:
        # value gets stored to same array number as first enrty of that name
        arrNumRows[i-nRowName] = nRowName
        print(arrName[i], nRowName, arrNumRows[i-nRowName]) # (KEEP) test number of rows per name
        nRowName = 1
for i in range(2, Nrow):
    if arrName[i+1] is arrName[i]:
        arrNumRows[i] = arrNumRows[i-1]
    if arrName[i] is None or 0:
        break

# main event loop
for i in range(2, Nrow):
    #for i in range(2, 25):
    if arrName[i] is None or 0:
        break
    print(arrName[i],"  Cat",arrCat[i],"  Hrs",arrHrs[i],"  Rate",arrRate[i],"  Tot",arrTot[i],"  Reim",arrReim[i],"  Gross",arrGross[i],"  Note",arrNote[i]," rows",arrNumRows[i-1])
    ###print(arrName[i],arrNumRows[i-1])

    # calculate hours per rate
    #for j in range(1, arrNumRows[i-1]):
        #print(j,arrNumRows[i-1])

    # calculate total hours for each category
    GrandHrs += arrHrs[i]
    GrandTot += arrTot[i]
    GrandReim += arrReim[i]
    GrandGross += arrGross[i]
    if arrCat[i] is None:
        print("No category given!")
        unknownCat = True
    elif arrCat[i].__contains__("Un") and not arrCat[i].__contains__("OT"):
        hrsUnarmed += arrHrs[i]
        GrandUnarmed += arrHrs[i]
    elif arrCat[i].__contains__("Armed") and not arrCat[i].__contains__("OT"):
        hrsArmed += arrHrs[i]
        GrandArmed += arrHrs[i]
    elif arrCat[i].__contains__("Admin"):
        hrsAdmin += arrHrs[i]
        GrandAdmin += arrHrs[i]
    elif arrCat[i].__contains__("OT"):
        hrsOT += arrHrs[i]
        GrandOT += arrHrs[i]
    elif arrCat[i].__contains__("Train"):
        hrsTrain += arrHrs[i]
        GrandTrain += arrHrs[i]
    elif arrCat[i].__contains__("Sick"):
        hrsSick += arrHrs[i]
        GrandSick += arrHrs[i]
    elif arrCat[i].__contains__("COVID"):
        hrsCOVID += arrHrs[i]
        GrandCovid += arrHrs[i]
    else:
        print("Unknown category for hours!")
        unknownCat = True
    hrsTotal += arrHrs[i]



# end main loop

print()
print("Unarmed: %.2f  Armed: %.2f  Admin: %.2f  OT: %.2f  Train: %.2f  Sick: %.2f  COVID: %.2f" % (GrandUnarmed,GrandArmed,GrandAdmin,GrandOT,GrandTrain,GrandSick,GrandCovid))
print("Total hours: %.2f  Total pay: %.2f  Reimbursement: %.2f  Gross pay: %.2f" % (hrsTotal,GrandTot,GrandReim,GrandGross))
print()
if unknownCat is True:
    print("New or missing work category detected! Fix this!")
print()
print("Done")
print()
