# Code for calculating overtime values
# "sort.py" alphabetizes and sorts payroll input by employee name LAST then FIRST.
#
# Instructions:
#
# Download .csv file "Assigned Shift Details - General" with the following parameters:
# "Employee Number", "Position Name", "Date", "Start Time", "End Time", "Duration", "Employee First Name",
# "Employee Last Name", "Employee Pay Rate".
# Save .csv file as "input.xlsx" and put in same directory as this code.

# Source files
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import pandas as pd
import sys
import os
if os.path.isfile('input.xlsx') is False: print("\n\"File input.xlsx not found!\"\n")

# Input file
workbook = load_workbook('input.xlsx')

df = pd.read_excel('input.xlsx')
# df['Employee Name'] = df['Employee Last Name'] + ' ' + df['Employee First Name']
df_sorted = df.sort_values(
    by = ['Employee Last Name', 'Employee First Name', 'Date'],
    ascending = [True, True, True]
)
df_sorted['Employee Name'] = df_sorted['Employee First Name'] + ' ' + df_sorted['Employee Last Name']
df_sorted.to_excel('output_sorted.xlsx', header = True, index = False)

sheet = workbook.active
label_0 = sheet.cell(row = 1, column = 1).value
label_1 = sheet.cell(row = 1, column = 2).value
label_2 = sheet.cell(row = 1, column = 3).value
label_3 = sheet.cell(row = 1, column = 4).value
label_4 = sheet.cell(row = 1, column = 5).value
label_5 = sheet.cell(row = 1, column = 6).value
label_6 = sheet.cell(row = 1, column = 7).value
label_7 = sheet.cell(row = 1, column = 8).value
label_8 = sheet.cell(row = 1, column = 9).value
if ("Employee Number" not in label_0 or "Position Name" not in label_1 or "Date" not in label_2 or
    "Start Time" not in label_3 or "End Time" not in label_4 or "Duration" not in label_5 or
    "Employee First Name" not in label_6 or "Employee Last Name" not in label_7 or "Employee Pay Rate" not in label_8):
    print("\n\n")
    print("===================================================")
    print("---WARNING!!! INCORRECT SPREADSHEET HEADERS!!!---  ")
    print("                                                   ")
    print("            ---Fix then try again!---              ")
    print("===================================================")
    print("\n\n")
    sys.exit(0)

workbook = load_workbook('output_sorted.xlsx')
sheet = workbook.active
for c in sheet["J"]:
    new_cell = c.offset(column = -3)
    new_cell.value = c.value
for c in sheet["I"]:
    new_cell = c.offset(column = -1)
    new_cell.value = c.value
sheet.delete_cols(10)
sheet.delete_cols(9)
sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 40
sheet.column_dimensions["C"].width = 22
sheet.column_dimensions["D"].width = 12
sheet.column_dimensions["E"].width = 12
sheet.column_dimensions["F"].width = 10
sheet.column_dimensions["G"].width = 20
sheet.column_dimensions["H"].width = 20
output_name = "output_sorted.xlsx"
workbook.save(output_name)

print("\n\n\n")
print("=======================================")
print("      Step 1: Sorting input file       ")
print("=======================================")
print("\n\n")
print("Input file sorted by last name, first name.\n")
print("File output written to", output_name, "\n")
print("--Step 1 complete--\n")
