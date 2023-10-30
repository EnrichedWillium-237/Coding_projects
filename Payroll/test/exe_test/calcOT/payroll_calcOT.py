# Code for calculating overtime values
# "sort.py" alphabetizes and sorts payroll input by employee name LAST then FIRST.
#
# Instructions:
#
# Download .csv file "Assigned Shift Details - General" with the following parameters:
# "Employee Number", "Position Name", "Date", "Start Time", "End Time", "Duration", "Employee First Name",
# "Employee Last Name", "Employee Pay Rate".
# Save .csv file as "input.xlsx" and put in same directory as this code.

# Source files
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import pandas as pd
import sys
import os
if os.path.isfile('input.xlsx') is False: print("\n\"File input.xlsx not found!\"\n")

# Input file
workbook = load_workbook('input.xlsx')

# create output directory if one does not exist
newpath = "./output"
if not os.path.exists(newpath):
    os.makedirs(newpath)

df = pd.read_excel('input.xlsx')
# df['Employee Name'] = df['Employee Last Name'] + ' ' + df['Employee First Name']
df_sorted = df.sort_values(
    by = ['Employee Last Name', 'Employee First Name', 'Date'],
    ascending = [True, True, True]
)
df_sorted['Employee Name'] = df_sorted['Employee First Name'] + ' ' + df_sorted['Employee Last Name']
df_sorted.to_excel('output/input_sorted.xlsx', header = True, index = False)

sheet = workbook.active
label_0 = sheet.cell(row = 1, column = 1).value
label_1 = sheet.cell(row = 1, column = 2).value
label_2 = sheet.cell(row = 1, column = 3).value
label_3 = sheet.cell(row = 1, column = 4).value
label_4 = sheet.cell(row = 1, column = 5).value
label_5 = sheet.cell(row = 1, column = 6).value
label_6 = sheet.cell(row = 1, column = 7).value
label_7 = sheet.cell(row = 1, column = 8).value
label_8 = sheet.cell(row = 1, column = 9).value
if ("Employee Number" not in label_0 or "Position Name" not in label_1 or "Date" not in label_2 or
    "Start Time" not in label_3 or "End Time" not in label_4 or "Duration" not in label_5 or
    "Employee First Name" not in label_6 or "Employee Last Name" not in label_7 or "Employee Pay Rate" not in label_8):
    print("\n\n")
    print("===================================================")
    print("---WARNING!!! INCORRECT SPREADSHEET HEADERS!!!---  ")
    print("                                                   ")
    print("            ---Fix then try again!---              ")
    print("===================================================")
    print("\n\n")
    sys.exit(0)

workbook = load_workbook('output/input_sorted.xlsx')
sheet = workbook.active
for c in sheet["J"]:
    new_cell = c.offset(column = -3)
    new_cell.value = c.value
for c in sheet["I"]:
    new_cell = c.offset(column = -1)
    new_cell.value = c.value
sheet.delete_cols(10)
sheet.delete_cols(9)
sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 40
sheet.column_dimensions["C"].width = 22
sheet.column_dimensions["D"].width = 12
sheet.column_dimensions["E"].width = 12
sheet.column_dimensions["F"].width = 10
sheet.column_dimensions["G"].width = 20
sheet.column_dimensions["H"].width = 20
output_name = "output/input_sorted.xlsx"
workbook.save(output_name)

print("\n\n\n")
print("=======================================")
print("      Step 1: Sorting input file       ")
print("=======================================")
print("\n\n")
print("Input file sorted by last name, first name.\n")
print("--Step 1 complete--\n")


###----------------------------------------------------------------------------------------------###

# Code for calculating OT values
# Reads in file "input_sorted.xlsx"
# Calculates overtime hours

# Source files
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Alignment, Border, Side, Font
from datetime import datetime, date, timedelta

flagDebug1 = False
flagDebug2 = False

# Input file
workbook = load_workbook('output/input_sorted.xlsx')

# Calculate week ranges
sheet = workbook.active
Nrow = sheet.max_row
daterange = sheet['C']
date_list = [daterange[x].value for x in range(2, len(daterange) - 10)]
import datetime
week1start = min(date_list)
week1end = week1start + datetime.timedelta(days = 6)
week2start = week1end + datetime.timedelta(days = 1)
week2end = max(date_list)

# Column parameters (change if spreadsheet changes)
colENum    = 1
colPos     = 2
colDate    = 3
colTimeSrt = 4
colTimeEnd = 5
colHrs     = 6
colName    = 7
colRate    = 8
timeGap = 8 # gap allowed in hours for OT+12

# Calculation for crosschecks
GrandHrs = 0
GrandNames = 0
for i in range(2, Nrow + 1):
    valHrs = sheet.cell(row = i, column = colHrs)
    GrandHrs += valHrs.value
    valName = sheet.cell(row = i, column = colName).value
    valNameNxt = sheet.cell(row = i + 1, column = colName).value
    if valNameNxt is not None and valNameNxt not in valName: GrandNames += 1

print("\n\n\n")
print("=======================================")
print("        Step 2: Calculating OT         ")
print("=======================================")
print("\n")
print("Payroll date range:")
print("--- Week 1:", week1start.strftime("%Y-%m-%d"), "to", week1end.strftime("%Y-%m-%d"), "---")
print("--- Week 2:", week2start.strftime("%Y-%m-%d"), "to", week2end.strftime("%Y-%m-%d"), "---")
print("\n")
print("Global statistics:")
print("  Total number of employees:", GrandNames)
print("  Total number of shifts:   ", Nrow - 1)
print("  Total hours for all positions:  ", f'{GrandHrs:.9}')

# Setup output file and spreadsheet
newbook1 = openpyxl.Workbook()
newsheet1 = newbook1.active
printCnt = 8 # Needed for printing out hours
printCntInit = 8
rowCnt = 1
warnShift = "\nEMPLOYEE HAS WORKED TOO MANY SHIFTS IN ONE WEEK!!!  GIVE THEM SOME TIME OFF!!!\n"
flagMultShiftTot = False

# Main event loop
for i in range(2, Nrow + 1):
    valName = sheet.cell(row = i, column = colName).value
    valNameNxt = sheet.cell(row = i + 1, column = colName).value
    if valNameNxt is not None and valNameNxt.__contains__(valName):
        rowCnt += 1
    else:
        rowmin = i - rowCnt + 1
        rowmax = i
        emptyList = [None, None, 0, 0, 0, 0, 0, None, 0]
        listWeek1 = [emptyList, emptyList, emptyList, emptyList, emptyList,
                     emptyList, emptyList, emptyList, emptyList]
        listWeek2 = [emptyList, emptyList, emptyList, emptyList, emptyList,
                     emptyList, emptyList, emptyList, emptyList]

        # Begin OT calculation
        # Find middle cell between week 1 and week 2
        rowmid = 0
        rowmidcnt = 0
        week1Hrs = 0
        week2Hrs = 0
        for j in range(rowmin, rowmax + 1):
            valDate = sheet.cell(row = j, column = colDate).value
            valHrs = sheet.cell(row = j, column = colHrs).value
            if valHrs > 0 and valHrs < 0.01: valHrs = 0
            if valDate <= week1end:
                week1Hrs += valHrs
                rowmidcnt += 1
            else:
                week2Hrs += valHrs
            rowmid = rowmin + rowmidcnt

        # OT +12 for week one and week 2 without 40+
        if week1Hrs <= 40:
            regHrs = 0
            dayHrsCnt = 0
            dayHrsGap = 0
            flagMultShift1 = False
            valName = sheet.cell(row = rowmid-1, column = colName).value
            list1  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list2  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list3  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list4  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list5  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list6  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list7  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list8  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list9  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list10 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list11 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list12 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list13 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list14 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list15 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list16 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list17 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list18 = [None, None, 0, 0, 0, 0, 0, None, 0]
            for j in range(rowmid - 1, rowmin - 1, -1): # minus one offset in loop because we're counting backwards
                valPos     = sheet.cell(row = j, column = colPos).value
                valDate    = sheet.cell(row = j, column = colDate).value
                valDateNxt = sheet.cell(row = j - 1, column = colDate).value
                valRate    = sheet.cell(row = j, column = colRate).value
                valENum    = sheet.cell(row = j, column = colENum).value
                valHrs = sheet.cell(row = j, column = colHrs).value
                valHrsNxt = sheet.cell(row = j - 1, column = colHrs).value
                dayCnt = 1
                kmax = 9
                dayHrsTot = valHrs
                if j < 10: kmax = j
                if j > 2:
                    if valDateNxt.day is valDate.day:
                        dayCnt += 1
                        dayHrsTot += valHrsNxt
                        ###print("j ",j," name: ",valName," valNameNxt: ",valNameNxt," valDate: ",valDate.day," valDateNxt: ",valDateNxt.day," dayCnt: ",dayCnt," dayHrsTot: ",dayHrsTot)
                        # if valName.__contains__("Borleske"):
                        #     print(j, "  Date: ", valDate.day," DateNext: ", valDateNxt.day," valHrs: ", valHrs,"valHrsNxt: ", valHrsNxt," dayCnt: ",dayCnt,"\n")
                    if valDateNxt is not None and valName is not valNameNxt and valDateNxt.day is valDate.day:
                        flagMultShift1 = True # remove flags when confident with calculation
                        flagMultShiftTot = True
                        valTimeSrt = float(sheet.cell(row = j + 1, column = colTimeSrt).value[:2]) + float(sheet.cell(row = j + 1, column = colTimeSrt).value[3:4])/6.
                        valTimeEnd = float(sheet.cell(row = j, column = colTimeEnd).value[:2]) + float(sheet.cell(row = j, column = colTimeEnd).value[3:4])/6.
                        valTimeDif = valTimeSrt - valTimeEnd
                        if valTimeDif <= timeGap:
                            flagMultShift2 = True
                            flagMultShiftTot = True
                    if flagDebug1: print("---Multiple shifts in same day this week. Evaluate OT+12 by hand!---")
                if valHrs > 0 and valHrs < 0.01: valHrs = 0
                if dayCnt == 1: # Single shift in a day
                    if valHrs > 12:
                        otHrs = valHrs - 12
                        regHrs = valHrs - otHrs
                    else:
                        otHrs = 0
                        regHrs = valHrs
                if dayCnt > 1 and dayHrsTot < 12: # multiple shifts, no OT+12
                    if valHrs > 12:
                        otHrs = valHrs - 12
                        regHrs = valHrs - otHrs
                    else:
                        otHrs = 0
                        regHrs = valHrs
                if dayCnt > 1 and dayHrsTot > 12: # multiple shifts with OT+12
                    if valHrs > 12 and valTimeDif <= timeGap:
                        otHrs = valHrs - 12
                        regHrs = valHrs - otHrs
                    else:
                        otHrs = 0
                        regHrs = valHrs

                if regHrs > 0 and regHrs < 0.01: regHrs = 0
                if regHrs < 0: regHrs = 0
                if otHrs > 0 and otHrs < 0.01: otHrs = 0
                if j == rowmid - 1:   list1   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 2: list2   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 3: list3   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 4: list4   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 5: list5   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 6: list6   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 7: list7   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 8: list8   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 9: list9   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 10: list10 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 11: list11 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 12: list12 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 13: list13 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 14: list14 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 15: list15 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 16: list16 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 17: list17 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmid - 18: list18 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                if j <= rowmid - 19 and valName.__contains__("EXPRESS 1MGR") is not True: print(warnShift, "  ", valName)
                if flagDebug1: print(valName, "  ", valPos, "  Rate:", valRate, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", otHrs)
                # if valName.__contains__("Borleske"):
                #     print("  ", valPos, "  Rate:", valRate, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", otHrs)
            listWeek1 = [list1, list2, list3, list4, list5, list6, list7, list8, list9, list10, list11, list12, list13, list14, list15, list16, list17, list18]
            if flagDebug1: print(valName, "  Week 1 --- total: ", week1Hrs, " shift+40 total: ", 0, "\n")
        if week2Hrs <= 40:
            regHrs = 0
            flagMultShift2 = False
            valName = sheet.cell(row = rowmax, column = colName).value
            list1  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list2  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list3  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list4  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list5  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list6  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list7  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list8  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list9  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list10 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list11 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list12 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list13 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list14 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list15 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list16 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list17 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list18 = [None, None, 0, 0, 0, 0, 0, None, 0]
            for j in range(rowmax, rowmid - 1, -1):
                valPos = sheet.cell(row = j, column = colPos).value
                valDate = sheet.cell(row = j, column = colDate).value
                valDateNxt = sheet.cell(row = j + 1, column = colDate).value
                valRate = sheet.cell(row = j, column = colRate).value
                valENum = sheet.cell(row = j, column = colENum).value
                valTimeSrt = sheet.cell(row = j + 1, column = colTimeSrt).value
                valTimeEnd = sheet.cell(row = j, column = colTimeSrt).value
                valTimeDif = 0
                if j > 2 and valDateNxt is not None and valDateNxt.day is valDate.day:
                    flagMultShift1 = True # remove flags when confident with calculation
                    flagMultShiftTot = True
                    valTimeSrt = float(sheet.cell(row = j + 1, column = colTimeSrt).value[:2]) + float(sheet.cell(row = j + 1, column = colTimeSrt).value[3:4])/6.
                    valTimeEnd = float(sheet.cell(row = j, column = colTimeEnd).value[:2]) + float(sheet.cell(row = j, column = colTimeEnd).value[3:4])/6.
                    valTimeDif = valTimeSrt - valTimeEnd
                    if valTimeDif <= timeGap:
                        flagMultShift2 = True
                        flagMultShiftTot = True
                    if flagDebug1: print("---Multiple shifts in same day this week. Evaluate OT+12 by hand!---")
                valHrs = sheet.cell(row = j, column = colHrs).value
                if valHrs > 0 and valHrs < 0.01: valHrs = 0
                if valHrs > 12 and valTimeDif <= timeGap:
                    otHrs = valHrs - 12
                    regHrs = valHrs - otHrs
                else:
                    otHrs = 0
                    regHrs = valHrs
                if regHrs > 0 and regHrs < 0.01: regHrs = 0
                if otHrs > 0 and otHrs < 0.01: otHrs = 0
                if j == rowmax:       list1   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 1: list2   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 2: list3   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 3: list4   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 4: list5   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 5: list6   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 6: list7   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 7: list8   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 8: list9   = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 9: list10  = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 10: list11 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 11: list12 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 12: list13 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 13: list14 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 14: list15 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 15: list16 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 16: list17 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                elif j == rowmax - 17: list18 = [valName, valPos, valHrs, regHrs, otHrs, 0, valRate, valENum, 0]
                if j <= rowmid - 18 and valName.__contains__("EXPRESS 1MGR") is not True: print(warnShift, "  ", valName)
                if flagDebug1: print(valName, "  ", valPos, "  Rate:", valRate, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", otHrs)
            listWeek2 = [list1, list2, list3, list4, list5, list6, list7, list8, list9, list10, list11, list12, list13, list14, list15, list16, list17, list18]
            if flagDebug1: print(valName, "  Week 2 --- total: ", week2Hrs, " shift+40 total: ", 0, "\n")

        # OT +40 for week one
        if week1Hrs > 40:
            regHrs = 0
            OT40week1 = 0
            OT12 = 0
            OT40week1 = week1Hrs - 40
            OT = OT40week1
            OTn = OT
            OTcnt = 0
            flag1 = False
            flagMultShift1 = False
            valName = sheet.cell(row = rowmid-1, column = colName)
            valName = valName.value
            list1  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list2  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list3  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list4  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list5  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list6  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list7  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list8  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list9  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list10 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list11 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list12 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list13 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list14 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list15 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list16 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list17 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list18 = [None, None, 0, 0, 0, 0, 0, None, 0]
            for j in range(rowmid - 1, rowmin - 1, -1): # minus one offset in for loop because we're counting backwards
                valPos = sheet.cell(row = j, column = colPos).value
                valDate = sheet.cell(row = j, column = colDate).value
                valDateNxt = sheet.cell(row = j + 1, column = colDate).value
                valRate = sheet.cell(row = j, column = colRate).value
                valENum = sheet.cell(row = j, column = colENum).value
                if j > 2 and valDateNxt is not None and valDateNxt.day is valDate.day:
                    flagMultShift1 = True
                    flagMultShiftTot = True
                    if flagDebug1: print("---Multiple shifts in same day this week. Evaluate OT+12 by hand!---")
                valHrs = sheet.cell(row = j, column = colHrs).value
                if valHrs > 0 and valHrs < 0.01: valHrs = 0
                x = valHrs
                if flag1 is True:
                    OTn = 0
                    otHrs = 0
                if flag1 is False:
                    OTn = OTn - x
                    if OTn > 0:
                        otHrs = x
                        OTcnt += otHrs
                    if OTn <= 0:
                        otHrs = OT - OTcnt
                        flag1 = True
                if valHrs > 12: # OT +12 calculation
                    z = valHrs - 12
                    if otHrs > z: z = 0
                    else: z = z - otHrs
                else: z = 0
                regHrs = valHrs - otHrs - z
                if otHrs + z == 0: regHrs = valHrs
                if regHrs > 0 and regHrs < 0.01: regHrs = 0
                if z > 0 and z < 0.01: z = 0
                if otHrs > 0 and otHrs < 0.01: otHrs = 0
                if j == rowmid - 1:   list1   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 2: list2   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 3: list3   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 4: list4   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 5: list5   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 6: list6   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 7: list7   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 8: list8   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 9: list9   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 10: list10 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 11: list11 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 12: list12 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 13: list13 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 14: list14 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 15: list15 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 16: list16 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 17: list17 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmid - 18: list18 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                if j <= rowmid - 19 and valName.__contains__("EXPRESS 1MGR") is not True: print(warnShift, "  ", valName)
                if flagDebug1: print(valName, "  ", valPos, "  Rate:", valRate, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", otHrs)

            listWeek1 = [list1, list2, list3, list4, list5, list6, list7, list8, list9, list10, list11, list12, list13, list14, list15, list16, list17, list18]
            if flagDebug1: print(valName,"  Week 1 --- total: ", week1Hrs," shift+40 total: ",OT40week1,"\n")

        # OT +40 for week two
        if week2Hrs > 40:
            regHrs = 0
            OT40week2 = 0
            OT40week2 = week2Hrs - 40
            OT = OT40week2
            OTn = OT
            OTcnt = 0
            flag1 = False
            flagMultShift2 = False
            valName = sheet.cell(row = rowmax, column = colName)
            valName = valName.value
            list1  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list2  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list3  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list4  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list5  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list6  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list7  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list8  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list9  = [None, None, 0, 0, 0, 0, 0, None, 0]
            list10 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list11 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list12 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list13 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list14 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list15 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list16 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list17 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list18 = [None, None, 0, 0, 0, 0, 0, None, 0]
            for j in range(rowmax, rowmid - 1, -1):
                valPos = sheet.cell(row = j, column = colPos).value
                valDate = sheet.cell(row = j, column = colDate).value
                valDateNxt = sheet.cell(row = j + 1, column = colDate).value
                valRate = sheet.cell(row = j, column = colRate).value
                valENum = sheet.cell(row = j, column = colENum).value
                if j > 2 and valDateNxt is not None and valDateNxt.day is valDate.day:
                    flagMultShift2 = True
                    flagMultShiftTot = True
                    if flagDebug1: print("---Multiple shifts in same day this week. Evaluate OT+12 by hand!---")
                valHrs = sheet.cell(row = j, column = colHrs).value
                if valHrs > 0 and valHrs < 0.01: valHrs = 0
                x = valHrs
                if flag1 is True:
                    OTn = 0
                    otHrs = 0
                if flag1 is False:
                    OTn = OTn - x
                    if OTn > 0:
                        otHrs = x
                        OTcnt += otHrs
                    if OTn <= 0:
                        otHrs = OT - OTcnt
                        flag1 = True
                if valHrs > 12: # OT +12 calculation
                    z = valHrs - 12
                    if otHrs > z: z = 0
                    else: z = z - otHrs
                else: z = 0
                regHrs = valHrs - otHrs - z
                if otHrs + z == 0: regHrs = valHrs
                if regHrs > 0 and regHrs < 0.01: regHrs = 0
                if z > 0 and z < 0.01: z = 0
                if otHrs > 0 and otHrs < 0.01: otHrs = 0
                if j == rowmax:       list1   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 1: list2   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 2: list3   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 3: list4   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 4: list5   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 5: list6   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 6: list7   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 7: list8   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 8: list9   = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 9:  list10 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 10: list11 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 11: list12 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 12: list13 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 13: list14 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 14: list15 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 15: list16 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 16: list17 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                elif j == rowmax - 17: list18 = [valName, valPos, valHrs, regHrs, z, otHrs, valRate, valENum, 0]
                if j <= rowmid - 18 and valName.__contains__("EXPRESS 1MGR") is not True: print(warnShift, "  ", valName)
                if flagDebug1: print(valName, "  ", valPos, "  Rate:", valRate, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", otHrs)

            listWeek2 = [list1, list2, list3, list4, list5, list6, list7, list8, list9, list10, list11, list12, list13, list14, list15, list16, list17, list18]
            if flagDebug1: print(valName,"  Week 2 --- total: ", week2Hrs," shift+40 total: ",OT40week2,"\n")

        # Sort by position
        for j in range(0, 2):
            if j == 0:
                list1  = (listWeek1[0])
                list2  = (listWeek1[1])
                list3  = (listWeek1[2])
                list4  = (listWeek1[3])
                list5  = (listWeek1[4])
                list6  = (listWeek1[5])
                list7  = (listWeek1[6])
                list8  = (listWeek1[7])
                list9  = (listWeek1[8])
                list10 = (listWeek1[9])
                list11 = (listWeek1[10])
                list12 = (listWeek1[11])
                list13 = (listWeek1[12])
                list14 = (listWeek1[13])
                list15 = (listWeek1[14])
                list16 = (listWeek1[15])
                list17 = (listWeek1[16])
                list18 = (listWeek1[17])
            else:
                list1  = (listWeek2[0])
                list2  = (listWeek2[1])
                list3  = (listWeek2[2])
                list4  = (listWeek2[3])
                list5  = (listWeek2[4])
                list6  = (listWeek2[5])
                list7  = (listWeek2[6])
                list8  = (listWeek2[7])
                list9  = (listWeek2[8])
                list10 = (listWeek2[9])
                list11 = (listWeek2[10])
                list12 = (listWeek2[11])
                list13 = (listWeek2[12])
                list14 = (listWeek2[13])
                list15 = (listWeek2[14])
                list16 = (listWeek2[15])
                list17 = (listWeek2[16])
                list18 = (listWeek2[17])
            pos1  = list1[1]
            pos2  = list2[1]
            pos3  = list3[1]
            pos4  = list4[1]
            pos5  = list5[1]
            pos6  = list6[1]
            pos7  = list7[1]
            pos8  = list8[1]
            pos9  = list9[1]
            pos10 = list10[1]
            pos11 = list11[1]
            pos12 = list12[1]
            pos13 = list13[1]
            pos14 = list14[1]
            pos15 = list15[1]
            pos16 = list16[1]
            pos17 = list17[1]
            pos18 = list18[1]
            hrsPos1 = list1[2]
            RegHrs1 = list1[3]
            totOT12_1 = list1[4]
            totOT40_1 = list1[5]
            # Position 1
            if pos1 is not None:
                if pos2 is not None and pos2 in pos1:
                    hrsPos1 += list2[2]
                    RegHrs1 += list2[3]
                    totOT12_1 += list2[4]
                    totOT40_1 += list2[5]
                if pos3 is not None and pos3 in pos1:
                    hrsPos1 += list3[2]
                    RegHrs1 += list3[3]
                    totOT12_1 += list3[4]
                    totOT40_1 += list3[5]
                if pos4 is not None and pos4 in pos1:
                    hrsPos1 += list4[2]
                    RegHrs1 += list4[3]
                    totOT12_1 += list4[4]
                    totOT40_1 += list4[5]
                if pos5 is not None and pos5 in pos1:
                    hrsPos1 += list5[2]
                    RegHrs1 += list5[3]
                    totOT12_1 += list5[4]
                    totOT40_1 += list5[5]
                if pos6 is not None and pos6 in pos1:
                    hrsPos1 += list6[2]
                    RegHrs1 += list6[3]
                    totOT12_1 += list6[4]
                    totOT40_1 += list6[5]
                if pos7 is not None and pos7 in pos1:
                    hrsPos1 += list7[2]
                    RegHrs1 += list7[3]
                    totOT12_1 += list7[4]
                    totOT40_1 += list7[5]
                if pos8 is not None and pos8 in pos1:
                    hrsPos1 += list8[2]
                    RegHrs1 += list8[3]
                    totOT12_1 += list8[4]
                    totOT40_1 += list8[5]
                if pos9 is not None and pos9 in pos1:
                    hrsPos1 += list9[2]
                    RegHrs1 += list9[3]
                    totOT12_1 += list9[4]
                    totOT40_1 += list9[5]
                if pos10 is not None and pos10 in pos1:
                    hrsPos1 += list10[2]
                    RegHrs1 += list10[3]
                    totOT12_1 += list10[4]
                    totOT40_1 += list10[5]
                if pos11 is not None and pos11 in pos1:
                    hrsPos1 += list11[2]
                    RegHrs1 += list11[3]
                    totOT12_1 += list11[4]
                    totOT40_1 += list11[5]
                if pos12 is not None and pos12 in pos1:
                    hrsPos1 += list12[2]
                    RegHrs1 += list12[3]
                    totOT12_1 += list12[4]
                    totOT40_1 += list12[5]
                if pos13 is not None and pos13 in pos1:
                    hrsPos1 += list13[2]
                    RegHrs1 += list13[3]
                    totOT12_1 += list13[4]
                    totOT40_1 += list13[5]
                if pos14 is not None and pos14 in pos1:
                    hrsPos1 += list14[2]
                    RegHrs1 += list14[3]
                    totOT12_1 += list14[4]
                    totOT40_1 += list14[5]
                if pos15 is not None and pos15 in pos1:
                    hrsPos1 += list15[2]
                    RegHrs1 += list15[3]
                    totOT12_1 += list15[4]
                    totOT40_1 += list15[5]
                if pos16 is not None and pos16 in pos1:
                    hrsPos1 += list16[2]
                    RegHrs1 += list16[3]
                    totOT12_1 += list16[4]
                    totOT40_1 += list16[5]
                if pos17 is not None and pos17 in pos1:
                    hrsPos1 += list17[2]
                    RegHrs1 += list17[3]
                    totOT12_1 += list17[4]
                    totOT40_1 += list17[5]
                if pos18 is not None and pos18 in pos1:
                    hrsPos1 += list18[2]
                    RegHrs1 += list18[3]
                    totOT12_1 += list18[4]
                    totOT40_1 += list18[5]
            list1 = [valName, pos1, hrsPos1, RegHrs1, totOT12_1, totOT40_1, list1[6], list1[7], 0]
            # Position 2
            hrsPos2 = 0
            RegHrs2 = 0
            totOT12_2 = 0
            totOT40_2 = 0
            if pos2 is not None:
                if (    pos1 is not None and
                        pos2 not in pos1):
                    hrsPos2 = list2[2]
                    RegHrs2 = list2[3]
                    totOT12_2 = list2[4]
                    totOT40_2 = list2[5]
                    if pos3 is not None and pos3 in pos2:
                        hrsPos2 += list3[2]
                        RegHrs2 += list3[3]
                        totOT12_2 += list3[4]
                        totOT40_2 += list3[5]
                    if pos4 is not None and pos4 in pos2:
                        hrsPos2 += list4[2]
                        RegHrs2 += list4[3]
                        totOT12_2 += list4[4]
                        totOT40_2 += list4[5]
                    if pos5 is not None and pos5 in pos2:
                        hrsPos2 += list5[2]
                        RegHrs2 += list5[3]
                        totOT12_2 += list5[4]
                        totOT40_2 += list5[5]
                    if pos6 is not None and pos6 in pos2:
                        hrsPos2 += list6[2]
                        RegHrs2 += list6[3]
                        totOT12_2 += list6[4]
                        totOT40_2 += list6[5]
                    if pos7 is not None and pos7 in pos2:
                        hrsPos2 += list7[2]
                        RegHrs2 += list7[3]
                        totOT12_2 += list7[4]
                        totOT40_2 += list7[5]
                    if pos8 is not None and pos8 in pos2:
                        hrsPos2 += list8[2]
                        RegHrs2 += list8[3]
                        totOT12_2 += list8[4]
                        totOT40_2 += list8[5]
                    if pos9 is not None and pos9 in pos2:
                        hrsPos2 += list9[2]
                        RegHrs2 += list9[3]
                        totOT12_2 += list9[4]
                        totOT40_2 += list9[5]
                    if pos10 is not None and pos10 in pos2:
                        hrsPos2 += list10[2]
                        RegHrs2 += list10[3]
                        totOT12_2 += list10[4]
                        totOT40_2 += list10[5]
                    if pos11 is not None and pos11 in pos2:
                        hrsPos2 += list11[2]
                        RegHrs2 += list11[3]
                        totOT12_2 += list11[4]
                        totOT40_2 += list11[5]
                    if pos12 is not None and pos12 in pos2:
                        hrsPos2 += list12[2]
                        RegHrs2 += list12[3]
                        totOT12_2 += list12[4]
                        totOT40_2 += list12[5]
                    if pos13 is not None and pos13 in pos2:
                        hrsPos2 += list13[2]
                        RegHrs2 += list13[3]
                        totOT12_2 += list13[4]
                        totOT40_2 += list13[5]
                    if pos14 is not None and pos14 in pos2:
                        hrsPos2 += list14[2]
                        RegHrs2 += list14[3]
                        totOT12_2 += list14[4]
                        totOT40_2 += list14[5]
                    if pos15 is not None and pos15 in pos2:
                        hrsPos2 += list15[2]
                        RegHrs2 += list15[3]
                        totOT12_2 += list15[4]
                        totOT40_2 += list15[5]
                    if pos16 is not None and pos16 in pos2:
                        hrsPos2 += list16[2]
                        RegHrs2 += list16[3]
                        totOT12_2 += list16[4]
                        totOT40_2 += list16[5]
                    if pos17 is not None and pos17 in pos2:
                        hrsPos2 += list17[2]
                        RegHrs2 += list17[3]
                        totOT12_2 += list17[4]
                        totOT40_2 += list17[5]
                    if pos18 is not None and pos18 in pos2:
                        hrsPos2 += list18[2]
                        RegHrs2 += list18[3]
                        totOT12_2 += list18[4]
                        totOT40_2 += list18[5]
                list2 = [valName, pos2, hrsPos2, RegHrs2, totOT12_2, totOT40_2, list2[6], list2[7], 0]
            # Position 3
            hrsPos3 = 0
            RegHrs3 = 0
            totOT12_3 = 0
            totOT40_3 = 0
            if pos3 is not None:
                if (    pos2 is not None and pos1 is not None and
                        pos3 not in pos1 and pos3 not in pos2):
                    hrsPos3 += list3[2]
                    RegHrs3 += list3[3]
                    totOT12_3 += list3[4]
                    totOT40_3 += list3[5]
                    if pos4 is not None and pos4 in pos3:
                        hrsPos3 += list4[2]
                        RegHrs3 += list4[3]
                        totOT12_3 += list4[4]
                        totOT40_3 += list4[5]
                    if pos5 is not None and pos5 in pos3:
                        hrsPos3 += list5[2]
                        RegHrs3 += list5[3]
                        totOT12_3 += list5[4]
                        totOT40_3 += list5[5]
                    if pos6 is not None and pos6 in pos3:
                        hrsPos3 += list6[2]
                        RegHrs3 += list6[3]
                        totOT12_3 += list6[4]
                        totOT40_3 += list6[5]
                    if pos7 is not None and pos7 in pos3:
                        hrsPos3 += list7[2]
                        RegHrs3 += list7[3]
                        totOT12_3 += list7[4]
                        totOT40_3 += list7[5]
                    if pos8 is not None and pos8 in pos3:
                        hrsPos3 += list8[2]
                        RegHrs3 += list8[3]
                        totOT12_3 += list8[4]
                        totOT40_3 += list8[5]
                    if pos9 is not None and pos9 in pos3:
                        hrsPos3 += list9[2]
                        RegHrs3 += list9[3]
                        totOT12_3 += list9[4]
                        totOT40_3 += list9[5]
                    if pos10 is not None and pos10 in pos3:
                        hrsPos3 += list10[2]
                        RegHrs3 += list10[3]
                        totOT12_3 += list10[4]
                        totOT40_3 += list10[5]
                    if pos11 is not None and pos11 in pos3:
                        hrsPos3 += list11[2]
                        RegHrs3 += list11[3]
                        totOT12_3 += list11[4]
                        totOT40_3 += list11[5]
                    if pos12 is not None and pos12 in pos3:
                        hrsPos3 += list12[2]
                        RegHrs3 += list12[3]
                        totOT12_3 += list12[4]
                        totOT40_3 += list12[5]
                    if pos13 is not None and pos13 in pos3:
                        hrsPos3 += list13[2]
                        RegHrs3 += list13[3]
                        totOT12_3 += list13[4]
                        totOT40_3 += list13[5]
                    if pos14 is not None and pos14 in pos3:
                        hrsPos3 += list14[2]
                        RegHrs3 += list14[3]
                        totOT12_3 += list14[4]
                        totOT40_3 += list14[5]
                    if pos15 is not None and pos15 in pos3:
                        hrsPos3 += list15[2]
                        RegHrs3 += list15[3]
                        totOT12_3 += list15[4]
                        totOT40_3 += list15[5]
                    if pos16 is not None and pos16 in pos3:
                        hrsPos3 += list16[2]
                        RegHrs3 += list16[3]
                        totOT12_3 += list16[4]
                        totOT40_3 += list16[5]
                    if pos17 is not None and pos17 in pos3:
                        hrsPos3 += list17[2]
                        RegHrs3 += list17[3]
                        totOT12_3 += list17[4]
                        totOT40_3 += list17[5]
                    if pos18 is not None and pos18 in pos3:
                        hrsPos3 += list18[2]
                        RegHrs3 += list18[3]
                        totOT12_3 += list18[4]
                        totOT40_3 += list18[5]
                list3 = [valName, pos3, hrsPos3, RegHrs3, totOT12_3, totOT40_3, list3[6], list3[7], 0]
            # Position 4
            hrsPos4 = 0
            RegHrs4 = 0
            totOT12_4 = 0
            totOT40_4 = 0
            if pos4 is not None:
                if (    pos3 is not None and pos2 is not None and pos1 is not None and
                        pos4 not in pos1 and pos4 not in pos2 and pos4 not in pos3):
                    hrsPos4 += list4[2]
                    RegHrs4 += list4[3]
                    totOT12_4 += list4[4]
                    totOT40_4 += list4[5]
                    if pos5 is not None and pos5 in pos4:
                        hrsPos4 += list5[2]
                        RegHrs4 += list5[3]
                        totOT12_4 += list5[4]
                        totOT40_4 += list5[5]
                    if pos6 is not None and pos6 in pos4:
                        hrsPos4 += list6[2]
                        RegHrs4 += list6[3]
                        totOT12_4 += list6[4]
                        totOT40_4 += list6[5]
                    if pos7 is not None and pos7 in pos4:
                        hrsPos4 += list7[2]
                        RegHrs4 += list7[3]
                        totOT12_4 += list7[4]
                        totOT40_4 += list7[5]
                    if pos8 is not None and pos8 in pos4:
                        hrsPos4 += list8[2]
                        RegHrs4 += list8[3]
                        totOT12_4 += list8[4]
                        totOT40_4 += list8[5]
                    if pos9 is not None and pos9 in pos4:
                        hrsPos4 += list9[2]
                        RegHrs4 += list9[3]
                        totOT12_4 += list9[4]
                        totOT40_4 += list9[5]
                    if pos10 is not None and pos10 in pos4:
                        hrsPos4 += list10[2]
                        RegHrs4 += list10[3]
                        totOT12_4 += list10[4]
                        totOT40_4 += list10[5]
                    if pos11 is not None and pos11 in pos4:
                        hrsPos4 += list11[2]
                        RegHrs4 += list11[3]
                        totOT12_4 += list11[4]
                        totOT40_4 += list11[5]
                    if pos12 is not None and pos12 in pos4:
                        hrsPos4 += list12[2]
                        RegHrs4 += list12[3]
                        totOT12_4 += list12[4]
                        totOT40_4 += list12[5]
                    if pos13 is not None and pos13 in pos4:
                        hrsPos4 += list13[2]
                        RegHrs4 += list13[3]
                        totOT12_4 += list13[4]
                        totOT40_4 += list13[5]
                    if pos14 is not None and pos14 in pos4:
                        hrsPos4 += list14[2]
                        RegHrs4 += list14[3]
                        totOT12_4 += list14[4]
                        totOT40_4 += list14[5]
                    if pos15 is not None and pos15 in pos4:
                        hrsPos4 += list15[2]
                        RegHrs4 += list15[3]
                        totOT12_4 += list15[4]
                        totOT40_4 += list15[5]
                    if pos16 is not None and pos16 in pos4:
                        hrsPos4 += list16[2]
                        RegHrs4 += list16[3]
                        totOT12_4 += list16[4]
                        totOT40_4 += list16[5]
                    if pos17 is not None and pos17 in pos4:
                        hrsPos4 += list17[2]
                        RegHrs4 += list17[3]
                        totOT12_4 += list17[4]
                        totOT40_4 += list17[5]
                    if pos18 is not None and pos18 in pos4:
                        hrsPos4 += list18[2]
                        RegHrs4 += list18[3]
                        totOT12_4 += list18[4]
                        totOT40_4 += list18[5]
                list4 = [valName, pos4, hrsPos4, RegHrs4, totOT12_4, totOT40_4, list4[6], list4[7], 0]
            # Position 5
            hrsPos5 = 0
            RegHrs5 = 0
            totOT12_5 = 0
            totOT40_5 = 0
            if pos5 is not None:
                if (    pos4 is not None and pos3 is not None and pos2 is not None and
                        pos1 is not None and
                        pos5 not in pos1 and pos5 not in pos2 and pos5 not in pos3 and
                        pos5 not in pos4):
                    hrsPos5 += list5[2]
                    RegHrs5 += list5[3]
                    totOT12_5 += list5[4]
                    totOT40_5 += list5[5]
                    if pos6 is not None and pos6 in pos5:
                        hrsPos5 += list6[2]
                        RegHrs5 += list6[3]
                        totOT12_5 += list6[4]
                        totOT40_5 += list6[5]
                    if pos7 is not None and pos7 in pos5:
                        hrsPos5 += list7[2]
                        RegHrs5 += list7[3]
                        totOT12_5 += list7[4]
                        totOT40_5 += list7[5]
                    if pos8 is not None and pos8 in pos5:
                        hrsPos5 += list8[2]
                        RegHrs5 += list8[3]
                        totOT12_5 += list8[4]
                        totOT40_5 += list8[5]
                    if pos9 is not None and pos9 in pos5:
                        hrsPos5 += list9[2]
                        RegHrs5 += list9[3]
                        totOT12_5 += list9[4]
                        totOT40_5 += list9[5]
                    if pos10 is not None and pos10 in pos5:
                        hrsPos5 += list10[2]
                        RegHrs5 += list10[3]
                        totOT12_5 += list10[4]
                        totOT40_5 += list10[5]
                    if pos11 is not None and pos11 in pos5:
                        hrsPos5 += list11[2]
                        RegHrs5 += list11[3]
                        totOT12_5 += list11[4]
                        totOT40_5 += list11[5]
                    if pos12 is not None and pos12 in pos5:
                        hrsPos5 += list12[2]
                        RegHrs5 += list12[3]
                        totOT12_5 += list12[4]
                        totOT40_5 += list12[5]
                    if pos13 is not None and pos13 in pos5:
                        hrsPos5 += list13[2]
                        RegHrs5 += list13[3]
                        totOT12_5 += list13[4]
                        totOT40_5 += list13[5]
                    if pos14 is not None and pos14 in pos5:
                        hrsPos5 += list14[2]
                        RegHrs5 += list14[3]
                        totOT12_5 += list14[4]
                        totOT40_5 += list14[5]
                    if pos15 is not None and pos15 in pos5:
                        hrsPos5 += list15[2]
                        RegHrs5 += list15[3]
                        totOT12_5 += list15[4]
                        totOT40_5 += list15[5]
                    if pos16 is not None and pos16 in pos5:
                        hrsPos5 += list16[2]
                        RegHrs5 += list16[3]
                        totOT12_5 += list16[4]
                        totOT40_5 += list16[5]
                    if pos17 is not None and pos17 in pos5:
                        hrsPos5 += list17[2]
                        RegHrs5 += list17[3]
                        totOT12_5 += list17[4]
                        totOT40_5 += list17[5]
                    if pos18 is not None and pos18 in pos5:
                        hrsPos5 += list18[2]
                        RegHrs5 += list18[3]
                        totOT12_5 += list18[4]
                        totOT40_5 += list18[5]
                list5 = [valName, pos5, hrsPos5, RegHrs5, totOT12_5, totOT40_5, list5[6], list5[7], 0]
            # Position 6
            hrsPos6 = 0
            RegHrs6 = 0
            totOT12_6 = 0
            totOT40_6 = 0
            if pos6 is not None:
                if (    pos5 is not None and pos4 is not None and pos3 is not None and
                        pos2 is not None and pos1 is not None and
                        pos6 not in pos1 and pos6 not in pos2 and pos6 not in pos3 and
                        pos6 not in pos4 and pos6 not in pos5):
                    hrsPos6 += list6[2]
                    RegHrs6 += list6[3]
                    totOT12_6 += list6[4]
                    totOT40_6 += list6[5]
                    if pos7 is not None and pos7 in pos6:
                        hrsPos6 += list7[2]
                        RegHrs6 += list7[3]
                        totOT12_6 += list7[4]
                        totOT40_6 += list7[5]
                    if pos8 is not None and pos8 in pos6:
                        hrsPos6 += list8[2]
                        RegHrs6 += list8[3]
                        totOT12_6 += list8[4]
                        totOT40_6 += list8[5]
                    if pos9 is not None and pos9 in pos6:
                        hrsPos6 += list9[2]
                        RegHrs6 += list9[3]
                        totOT12_6 += list9[4]
                        totOT40_6 += list9[5]
                    if pos10 is not None and pos10 in pos6:
                        hrsPos6 += list10[2]
                        RegHrs6 += list10[3]
                        totOT12_6 += list10[4]
                        totOT40_6 += list10[5]
                    if pos11 is not None and pos11 in pos6:
                        hrsPos6 += list11[2]
                        RegHrs6 += list11[3]
                        totOT12_6 += list11[4]
                        totOT40_6 += list11[5]
                    if pos12 is not None and pos12 in pos6:
                        hrsPos6 += list12[2]
                        RegHrs6 += list12[3]
                        totOT12_6 += list12[4]
                        totOT40_6 += list12[5]
                    if pos13 is not None and pos13 in pos6:
                        hrsPos6 += list13[2]
                        RegHrs6 += list13[3]
                        totOT12_6 += list13[4]
                        totOT40_6 += list13[5]
                    if pos14 is not None and pos14 in pos6:
                        hrsPos6 += list14[2]
                        RegHrs6 += list14[3]
                        totOT12_6 += list14[4]
                        totOT40_6 += list14[5]
                    if pos15 is not None and pos15 in pos6:
                        hrsPos6 += list15[2]
                        RegHrs6 += list15[3]
                        totOT12_6 += list15[4]
                        totOT40_6 += list15[5]
                    if pos16 is not None and pos16 in pos6:
                        hrsPos6 += list16[2]
                        RegHrs6 += list16[3]
                        totOT12_6 += list16[4]
                        totOT40_6 += list16[5]
                    if pos17 is not None and pos17 in pos6:
                        hrsPos6 += list17[2]
                        RegHrs6 += list17[3]
                        totOT12_6 += list17[4]
                        totOT40_6 += list17[5]
                    if pos18 is not None and pos18 in pos6:
                        hrsPos6 += list18[2]
                        RegHrs6 += list18[3]
                        totOT12_6 += list18[4]
                        totOT40_6 += list18[5]
                list6 = [valName, pos6, hrsPos6, RegHrs6, totOT12_6, totOT40_6, list6[6], list6[7], 0]
            # Position 7
            hrsPos7 = 0
            RegHrs7 = 0
            totOT12_7 = 0
            totOT40_7 = 0
            if pos7 is not None:
                if (    pos6 is not None and pos5 is not None and pos4 is not None and
                        pos3 is not None and pos2 is not None and pos1 is not None and
                        pos7 not in pos1 and pos7 not in pos2 and pos7 not in pos3 and
                        pos7 not in pos4 and pos7 not in pos5 and pos7 not in pos6):
                    hrsPos7 += list7[2]
                    RegHrs7 += list7[3]
                    totOT12_7 += list7[4]
                    totOT40_7 += list7[5]
                    if pos8 is not None and pos8 in pos7:
                        hrsPos7 += list8[2]
                        RegHrs7 += list8[3]
                        totOT12_7 += list8[4]
                        totOT40_7 += list8[5]
                    if pos9 is not None and pos9 in pos7:
                        hrsPos7 += list9[2]
                        RegHrs7 += list9[3]
                        totOT12_7 += list9[4]
                        totOT40_7 += list9[5]
                    if pos10 is not None and pos10 in pos7:
                        hrsPos7 += list10[2]
                        RegHrs7 += list10[3]
                        totOT12_7 += list10[4]
                        totOT40_7 += list10[5]
                    if pos11 is not None and pos11 in pos7:
                        hrsPos7 += list11[2]
                        RegHrs7 += list11[3]
                        totOT12_7 += list11[4]
                        totOT40_7 += list11[5]
                    if pos12 is not None and pos12 in pos7:
                        hrsPos7 += list12[2]
                        RegHrs7 += list12[3]
                        totOT12_7 += list12[4]
                        totOT40_7 += list12[5]
                    if pos13 is not None and pos13 in pos7:
                        hrsPos7 += list13[2]
                        RegHrs7 += list13[3]
                        totOT12_7 += list13[4]
                        totOT40_7 += list13[5]
                    if pos14 is not None and pos14 in pos7:
                        hrsPos7 += list14[2]
                        RegHrs7 += list14[3]
                        totOT12_7 += list14[4]
                        totOT40_7 += list14[5]
                    if pos15 is not None and pos15 in pos7:
                        hrsPos7 += list15[2]
                        RegHrs7 += list15[3]
                        totOT12_7 += list15[4]
                        totOT40_7 += list15[5]
                    if pos16 is not None and pos16 in pos7:
                        hrsPos7 += list16[2]
                        RegHrs7 += list16[3]
                        totOT12_7 += list16[4]
                        totOT40_7 += list16[5]
                    if pos17 is not None and pos17 in pos7:
                        hrsPos7 += list17[2]
                        RegHrs7 += list17[3]
                        totOT12_7 += list17[4]
                        totOT40_7 += list17[5]
                    if pos18 is not None and pos18 in pos7:
                        hrsPos7 += list18[2]
                        RegHrs7 += list18[3]
                        totOT12_7 += list18[4]
                        totOT40_7 += list18[5]
                list7 = [valName, pos7, hrsPos7, RegHrs7, totOT12_7, totOT40_7, list7[6], list7[7], 0]
            # Position 8
            hrsPos8 = 0
            RegHrs8 = 0
            totOT12_8 = 0
            totOT40_8 = 0
            if pos8 is not None:
                if (    pos7 is not None and pos6 is not None and pos5 is not None and
                        pos4 is not None and pos3 is not None and pos2 is not None and
                        pos1 is not None and
                        pos8 not in pos1 and pos8 not in pos2 and pos8 not in pos3 and
                        pos8 not in pos4 and pos8 not in pos5 and pos8 not in pos6 and
                        pos8 not in pos7):
                    hrsPos8 += list8[2]
                    RegHrs8 += list8[3]
                    totOT12_8 += list8[4]
                    totOT40_8 += list8[5]
                    if pos9 is not None and pos9 in pos8:
                        hrsPos8 += list9[2]
                        RegHrs8 += list9[3]
                        totOT12_8 += list9[4]
                        totOT40_8 += list9[5]
                    if pos10 is not None and pos10 in pos8:
                        hrsPos8 += list10[2]
                        RegHrs8 += list10[3]
                        totOT12_8 += list10[4]
                        totOT40_8 += list10[5]
                    if pos11 is not None and pos11 in pos8:
                        hrsPos8 += list11[2]
                        RegHrs8 += list11[3]
                        totOT12_8 += list11[4]
                        totOT40_8 += list11[5]
                    if pos12 is not None and pos12 in pos8:
                        hrsPos8 += list12[2]
                        RegHrs8 += list12[3]
                        totOT12_8 += list12[4]
                        totOT40_8 += list12[5]
                    if pos13 is not None and pos13 in pos8:
                        hrsPos8 += list13[2]
                        RegHrs8 += list13[3]
                        totOT12_8 += list13[4]
                        totOT40_8 += list13[5]
                    if pos14 is not None and pos14 in pos8:
                        hrsPos8 += list14[2]
                        RegHrs8 += list14[3]
                        totOT12_8 += list14[4]
                        totOT40_8 += list14[5]
                    if pos15 is not None and pos15 in pos8:
                        hrsPos8 += list15[2]
                        RegHrs8 += list15[3]
                        totOT12_8 += list15[4]
                        totOT40_8 += list15[5]
                    if pos16 is not None and pos16 in pos8:
                        hrsPos8 += list16[2]
                        RegHrs8 += list16[3]
                        totOT12_8 += list16[4]
                        totOT40_8 += list16[5]
                    if pos17 is not None and pos17 in pos8:
                        hrsPos8 += list17[2]
                        RegHrs8 += list17[3]
                        totOT12_8 += list17[4]
                        totOT40_8 += list17[5]
                    if pos18 is not None and pos18 in pos8:
                        hrsPos8 += list18[2]
                        RegHrs8 += list18[3]
                        totOT12_8 += list18[4]
                        totOT40_8 += list18[5]
                list8 = [valName, pos8, hrsPos8, RegHrs8, totOT12_8, totOT40_8, list8[6], list8[7], 0]
            # Position 9
            hrsPos9 = 0
            RegHrs9 = 0
            totOT12_9 = 0
            totOT40_9 = 0
            if pos9 is not None:
                if (    pos8 is not None and pos7 is not None and pos6 is not None and
                        pos5 is not None and pos4 is not None and pos3 is not None and
                        pos2 is not None and pos1 is not None and
                        pos9 not in pos1 and pos9 not in pos2 and pos9 not in pos3 and
                        pos9 not in pos4 and pos9 not in pos5 and pos9 not in pos6 and
                        pos9 not in pos7 and pos9 not in pos8):
                    hrsPos9 += list9[2]
                    RegHrs9 += list9[3]
                    totOT12_9 += list9[4]
                    totOT40_9 += list9[5]
                    if pos10 is not None and pos10 in pos9:
                        hrsPos9 += list10[2]
                        RegHrs9 += list10[3]
                        totOT12_9 += list10[4]
                        totOT40_9 += list10[5]
                    if pos11 is not None and pos11 in pos9:
                        hrsPos9 += list11[2]
                        RegHrs9 += list11[3]
                        totOT12_9 += list11[4]
                        totOT40_9 += list11[5]
                    if pos12 is not None and pos12 in pos9:
                        hrsPos9 += list12[2]
                        RegHrs9 += list12[3]
                        totOT12_9 += list12[4]
                        totOT40_9 += list12[5]
                    if pos13 is not None and pos13 in pos9:
                        hrsPos9 += list13[2]
                        RegHrs9 += list13[3]
                        totOT12_9 += list13[4]
                        totOT40_9 += list13[5]
                    if pos14 is not None and pos14 in pos9:
                        hrsPos9 += list14[2]
                        RegHrs9 += list14[3]
                        totOT12_9 += list14[4]
                        totOT40_9 += list14[5]
                    if pos15 is not None and pos15 in pos9:
                        hrsPos9 += list15[2]
                        RegHrs9 += list15[3]
                        totOT12_9 += list15[4]
                        totOT40_9 += list15[5]
                    if pos16 is not None and pos16 in pos9:
                        hrsPos9 += list16[2]
                        RegHrs9 += list16[3]
                        totOT12_9 += list16[4]
                        totOT40_9 += list16[5]
                    if pos17 is not None and pos17 in pos9:
                        hrsPos9 += list17[2]
                        RegHrs9 += list17[3]
                        totOT12_9 += list17[4]
                        totOT40_9 += list17[5]
                    if pos18 is not None and pos18 in pos9:
                        hrsPos9 += list18[2]
                        RegHrs9 += list18[3]
                        totOT12_9 += list18[4]
                        totOT40_9 += list18[5]
                list9 = [valName, pos9, hrsPos9, RegHrs9, totOT12_9, totOT40_9, list9[6], list9[7], 0]
            # Position 10
            hrsPos10 = 0
            RegHrs10 = 0
            totOT12_10 = 0
            totOT40_10 = 0
            if pos10 is not None:
                if (    pos9 is not None and pos8 is not None and pos7 is not None and
                        pos6 is not None and pos5 is not None and pos4 is not None and
                        pos3 is not None and pos2 is not None and pos1 is not None and
                        pos10 not in pos1 and pos10 not in pos2 and pos10 not in pos3 and
                        pos10 not in pos4 and pos10 not in pos5 and pos10 not in pos6 and
                        pos10 not in pos7 and pos10 not in pos8 and pos10 not in pos9):
                    hrsPos10 += list10[2]
                    RegHrs10 += list10[3]
                    totOT12_10 += list10[4]
                    totOT40_10 += list10[5]
                    if pos11 is not None and pos11 in pos10:
                        hrsPos10 += list11[2]
                        RegHrs10 += list11[3]
                        totOT12_10 += list11[4]
                        totOT40_10 += list11[5]
                    if pos12 is not None and pos12 in pos10:
                        hrsPos10 += list12[2]
                        RegHrs10 += list12[3]
                        totOT12_10 += list12[4]
                        totOT40_10 += list12[5]
                    if pos13 is not None and pos13 in pos10:
                        hrsPos10 += list13[2]
                        RegHrs10 += list13[3]
                        totOT12_10 += list13[4]
                        totOT40_10 += list13[5]
                    if pos14 is not None and pos14 in pos10:
                        hrsPos10 += list14[2]
                        RegHrs10 += list14[3]
                        totOT12_10 += list14[4]
                        totOT40_10 += list14[5]
                    if pos15 is not None and pos15 in pos10:
                        hrsPos10 += list15[2]
                        RegHrs10 += list15[3]
                        totOT12_10 += list15[4]
                        totOT40_10 += list15[5]
                    if pos16 is not None and pos16 in pos10:
                        hrsPos10 += list16[2]
                        RegHrs10 += list16[3]
                        totOT12_10 += list16[4]
                        totOT40_10 += list16[5]
                    if pos17 is not None and pos17 in pos10:
                        hrsPos10 += list17[2]
                        RegHrs10 += list17[3]
                        totOT12_10 += list17[4]
                        totOT40_10 += list17[5]
                    if pos18 is not None and pos18 in pos10:
                        hrsPos10 += list18[2]
                        RegHrs10 += list18[3]
                        totOT12_10 += list18[4]
                        totOT40_10 += list18[5]
                list10 = [valName, pos10, hrsPos10, RegHrs10, totOT12_10, totOT40_10, list10[6], list10[7], 0]
            # Position 11
            hrsPos11 = 0
            RegHrs11 = 0
            totOT12_11 = 0
            totOT40_11 = 0
            if pos11 is not None:
                if (    pos10 is not None and pos9 is not None and pos8 is not None and
                        pos7 is not None and pos6 is not None and pos5 is not None and
                        pos4 is not None and pos3 is not None and pos2 is not None and
                        pos1 is not None and
                        pos11 not in pos1 and pos11 not in pos2 and pos11 not in pos3 and
                        pos11 not in pos4 and pos11 not in pos5 and pos11 not in pos6 and
                        pos11 not in pos7 and pos11 not in pos8 and pos11 not in pos9 and
                        pos11 not in pos10):
                    hrsPos11 += list11[2]
                    RegHrs11 += list11[3]
                    totOT12_11 += list11[4]
                    totOT40_11 += list11[5]
                    if pos12 is not None and pos12 in pos11:
                        hrsPos11 += list12[2]
                        RegHrs11 += list12[3]
                        totOT12_11 += list12[4]
                        totOT40_11 += list12[5]
                    if pos13 is not None and pos13 in pos11:
                        hrsPos11 += list13[2]
                        RegHrs11 += list13[3]
                        totOT12_11 += list13[4]
                        totOT40_11 += list13[5]
                    if pos14 is not None and pos14 in pos11:
                        hrsPos11 += list14[2]
                        RegHrs11 += list14[3]
                        totOT12_11 += list14[4]
                        totOT40_11 += list14[5]
                    if pos15 is not None and pos15 in pos11:
                        hrsPos11 += list15[2]
                        RegHrs11 += list15[3]
                        totOT12_11 += list15[4]
                        totOT40_11 += list15[5]
                    if pos16 is not None and pos16 in pos11:
                        hrsPos11 += list16[2]
                        RegHrs11 += list16[3]
                        totOT12_11 += list16[4]
                        totOT40_11 += list16[5]
                    if pos17 is not None and pos17 in pos11:
                        hrsPos11 += list17[2]
                        RegHrs11 += list17[3]
                        totOT12_11 += list17[4]
                        totOT40_11 += list17[5]
                    if pos18 is not None and pos18 in pos11:
                        hrsPos11 += list18[2]
                        RegHrs11 += list18[3]
                        totOT12_11 += list18[4]
                        totOT40_11 += list18[5]
                list11 = [valName, pos11, hrsPos11, RegHrs11, totOT12_11, totOT40_11, list11[6], list11[7], 0]
            # Position 12
            hrsPos12 = 0
            RegHrs12 = 0
            totOT12_12 = 0
            totOT40_12 = 0
            if pos12 is not None:
                if (    pos11 is not None and pos10 is not None and pos9 is not None and
                        pos8 is not None and pos7 is not None and pos6 is not None and
                        pos5 is not None and pos4 is not None and pos3 is not None and
                        pos2 is not None and pos1 is not None and
                        pos12 not in pos1 and pos12 not in pos2 and pos12 not in pos3 and
                        pos12 not in pos4 and pos12 not in pos5 and pos12 not in pos6 and
                        pos12 not in pos7 and pos12 not in pos8 and pos12 not in pos9 and
                        pos12 not in pos10 and pos12 not in pos11):
                    hrsPos12 += list12[2]
                    RegHrs12 += list12[3]
                    totOT12_12 += list12[4]
                    totOT40_12 += list12[5]
                    if pos13 is not None and pos13 in pos12:
                        hrsPos12 += list13[2]
                        RegHrs12 += list13[3]
                        totOT12_12 += list13[4]
                        totOT40_12 += list13[5]
                    if pos14 is not None and pos14 in pos12:
                        hrsPos12 += list14[2]
                        RegHrs12 += list14[3]
                        totOT12_12 += list14[4]
                        totOT40_12 += list14[5]
                    if pos15 is not None and pos15 in pos12:
                        hrsPos12 += list15[2]
                        RegHrs12 += list15[3]
                        totOT12_12 += list15[4]
                        totOT40_12 += list15[5]
                    if pos16 is not None and pos16 in pos12:
                        hrsPos12 += list16[2]
                        RegHrs12 += list16[3]
                        totOT12_12 += list16[4]
                        totOT40_12 += list16[5]
                    if pos17 is not None and pos17 in pos12:
                        hrsPos12 += list17[2]
                        RegHrs12 += list17[3]
                        totOT12_12 += list17[4]
                        totOT40_12 += list17[5]
                    if pos18 is not None and pos18 in pos12:
                        hrsPos12 += list18[2]
                        RegHrs12 += list18[3]
                        totOT12_12 += list18[4]
                        totOT40_12 += list18[5]
                list12 = [valName, pos12, hrsPos12, RegHrs12, totOT12_12, totOT40_12, list12[6], list12[7], 0]
            # Position 13
            hrsPos13 = 0
            RegHrs13 = 0
            totOT12_13 = 0
            totOT40_13 = 0
            if pos13 is not None:
                if (    pos12 is not None and pos11 is not None and pos10 is not None and
                        pos9 is not None and pos8 is not None and pos7 is not None and
                        pos6 is not None and pos5 is not None and pos4 is not None and
                        pos3 is not None and pos2 is not None and pos1 is not None and
                        pos13 not in pos1 and pos13 not in pos2 and pos13 not in pos3 and
                        pos13 not in pos4 and pos13 not in pos5 and pos13 not in pos6 and
                        pos13 not in pos7 and pos13 not in pos8 and pos13 not in pos9 and
                        pos13 not in pos10 and pos13 not in pos11 and pos13 not in pos12):
                    hrsPos13 += list13[2]
                    RegHrs13 += list13[3]
                    totOT12_13 += list13[4]
                    totOT40_13 += list13[5]
                    if pos14 is not None and pos14 in pos14:
                        hrsPos13 += list14[2]
                        RegHrs13 += list14[3]
                        totOT12_13 += list14[4]
                        totOT40_13 += list14[5]
                    if pos15 is not None and pos15 in pos13:
                        hrsPos13 += list15[2]
                        RegHrs13 += list15[3]
                        totOT12_13 += list15[4]
                        totOT40_13 += list15[5]
                    if pos16 is not None and pos16 in pos13:
                        hrsPos13 += list16[2]
                        RegHrs13 += list16[3]
                        totOT12_13 += list16[4]
                        totOT40_13 += list16[5]
                    if pos17 is not None and pos17 in pos13:
                        hrsPos13 += list17[2]
                        RegHrs13 += list17[3]
                        totOT12_13 += list17[4]
                        totOT40_13 += list17[5]
                    if pos18 is not None and pos18 in pos13:
                        hrsPos13 += list18[2]
                        RegHrs13 += list18[3]
                        totOT12_13 += list18[4]
                        totOT40_13 += list18[5]
                list13 = [valName, pos13, hrsPos13, RegHrs13, totOT12_13, totOT40_13, list13[6], list13[7], 0]
            # Position 14
            hrsPos14 = 0
            RegHrs14 = 0
            totOT12_14 = 0
            totOT40_14 = 0
            if pos14 is not None:
                if (    pos13 is not None and pos12 is not None and pos11 is not None and
                        pos10 is not None and pos9 is not None and pos8 is not None and
                        pos7 is not None and pos6 is not None and pos5 is not None and
                        pos4 is not None and pos3 is not None and pos2 is not None and
                        pos1 is not None and
                        pos14 not in pos1 and pos14 not in pos2 and pos14 not in pos3 and
                        pos14 not in pos4 and pos14 not in pos5 and pos14 not in pos6 and
                        pos14 not in pos7 and pos14 not in pos8 and pos14 not in pos9 and
                        pos14 not in pos10 and pos14 not in pos11 and pos14 not in pos12 and
                        pos14 not in pos13):
                    hrsPos14 += list14[2]
                    RegHrs14 += list14[3]
                    totOT12_14 += list14[4]
                    totOT40_14 += list14[5]
                    if pos15 is not None and pos15 in pos14:
                        hrsPos14 += list15[2]
                        RegHrs14 += list15[3]
                        totOT12_14 += list15[4]
                        totOT40_14 += list15[5]
                    if pos16 is not None and pos16 in pos14:
                        hrsPos14 += list16[2]
                        RegHrs14 += list16[3]
                        totOT12_14 += list16[4]
                        totOT40_14 += list16[5]
                    if pos17 is not None and pos17 in pos14:
                        hrsPos14 += list17[2]
                        RegHrs14 += list17[3]
                        totOT12_14 += list17[4]
                        totOT40_14 += list17[5]
                    if pos18 is not None and pos18 in pos14:
                        hrsPos14 += list18[2]
                        RegHrs14 += list18[3]
                        totOT12_14 += list18[4]
                        totOT40_14 += list18[5]
                list14 = [valName, pos14, hrsPos14, RegHrs14, totOT12_14, totOT40_14, list14[6], list14[7], 0]
            # Position 15
            hrsPos15 = 0
            RegHrs15 = 0
            totOT12_15 = 0
            totOT40_15 = 0
            if pos15 is not None:
                if (    pos14 is not None and pos13 is not None and pos12 is not None and
                        pos11 is not None and pos10 is not None and pos9 is not None and
                        pos8 is not None and pos7 is not None and pos6 is not None and
                        pos5 is not None and pos4 is not None and pos3 is not None and
                        pos2 is not None and pos1 is not None and
                        pos15 not in pos1 and pos15 not in pos2 and pos15 not in pos3 and
                        pos15 not in pos4 and pos15 not in pos5 and pos15 not in pos6 and
                        pos15 not in pos7 and pos15 not in pos8 and pos15 not in pos9 and
                        pos15 not in pos10 and pos15 not in pos11 and pos15 not in pos12 and
                        pos15 not in pos13 and pos15 not in pos14):
                    hrsPos15 += list15[2]
                    RegHrs15 += list15[3]
                    totOT12_15 += list15[4]
                    totOT40_15 += list15[5]
                    if pos16 is not None and pos16 in pos15:
                        hrsPos15 += list16[2]
                        RegHrs15 += list16[3]
                        totOT12_15 += list16[4]
                        totOT40_15 += list16[5]
                    if pos17 is not None and pos17 in pos15:
                        hrsPos15 += list17[2]
                        RegHrs15 += list17[3]
                        totOT12_15 += list17[4]
                        totOT40_15 += list17[5]
                    if pos18 is not None and pos18 in pos15:
                        hrsPos15 += list18[2]
                        RegHrs15 += list18[3]
                        totOT12_15 += list18[4]
                        totOT40_15 += list18[5]
                list15 = [valName, pos15, hrsPos15, RegHrs15, totOT12_15, totOT40_15, list15[6], list15[7], 0]
            # Position 16
            hrsPos16 = 0
            RegHrs16 = 0
            totOT12_16 = 0
            totOT40_16 = 0
            if pos16 is not None:
                if (    pos15 is not None and pos14 is not None and pos13 is not None and
                        pos12 is not None and pos11 is not None and pos10 is not None and
                        pos9 is not None and pos8 is not None and pos7 is not None and
                        pos6 is not None and pos5 is not None and pos4 is not None and
                        pos3 is not None and pos2 is not None and pos1 is not None and
                        pos16 not in pos1 and pos16 not in pos2 and pos16 not in pos3 and
                        pos16 not in pos4 and pos16 not in pos5 and pos16 not in pos6 and
                        pos16 not in pos7 and pos16 not in pos8 and pos16 not in pos9 and
                        pos16 not in pos10 and pos16 not in pos11 and pos16 not in pos12 and
                        pos16 not in pos13 and pos16 not in pos14 and pos16 not in pos15):
                    hrsPos16 += list16[2]
                    RegHrs16 += list16[3]
                    totOT12_16 += list16[4]
                    totOT40_16 += list16[5]
                    if pos17 is not None and pos17 in pos16:
                        hrsPos16 += list17[2]
                        RegHrs16 += list17[3]
                        totOT12_16 += list17[4]
                        totOT40_16 += list17[5]
                    if pos18 is not None and pos18 in pos16:
                        hrsPos16 += list18[2]
                        RegHrs16 += list18[3]
                        totOT12_16 += list18[4]
                        totOT40_16 += list18[5]
                list16 = [valName, pos16, hrsPos16, RegHrs16, totOT12_16, totOT40_16, list16[6], list16[7], 0]
            # Position 17
            hrsPos17 = 0
            RegHrs17 = 0
            totOT12_17 = 0
            totOT40_17 = 0
            if pos17 is not None:
                if (    pos16 is not None and pos15 is not None and pos14 is not None and
                        pos13 is not None and pos12 is not None and pos11 is not None and
                        pos10 is not None and pos9 is not None and pos8 is not None and
                        pos7 is not None and pos6 is not None and pos5 is not None and
                        pos4 is not None and pos3 is not None and pos2 is not None and
                        pos1 is not None and
                        pos17 not in pos1 and pos17 not in pos2 and pos17 not in pos3 and
                        pos17 not in pos4 and pos17 not in pos5 and pos17 not in pos6 and
                        pos17 not in pos7 and pos17 not in pos8 and pos17 not in pos9 and
                        pos17 not in pos10 and pos17 not in pos11 and pos17 not in pos12 and
                        pos17 not in pos13 and pos17 not in pos14 and pos17 not in pos15 and
                        pos17 not in pos16):
                    hrsPos17 += list17[2]
                    RegHrs17 += list17[3]
                    totOT12_17 += list17[4]
                    totOT40_17 += list17[5]
                    if pos18 is not None and pos18 in pos17:
                        hrsPos17 += list18[2]
                        RegHrs17 += list18[3]
                        totOT12_17 += list18[4]
                        totOT40_17 += list18[5]
                list17 = [valName, pos17, hrsPos17, RegHrs17, totOT12_17, totOT40_17, list17[6], list17[7], 0]
            # Position 18
            hrsPos18 = 0
            RegHrs18 = 0
            totOT12_18 = 0
            totOT40_18 = 0
            if pos18 is not None:
                if (    pos17 is not None and pos16 is not None and pos15 is not None and
                        pos14 is not None and pos13 is not None and pos12 is not None and
                        pos11 is not None and pos10 is not None and pos9 is not None and
                        pos8 is not None and pos7 is not None and pos6 is not None and
                        pos5 is not None and pos4 is not None and pos3 is not None and
                        pos2 is not None and pos1 is not None and
                        pos18 not in pos1 and pos18 not in pos2 and pos18 not in pos3 and
                        pos18 not in pos4 and pos18 not in pos5 and pos18 not in pos6 and
                        pos18 not in pos7 and pos18 not in pos8 and pos18 not in pos9 and
                        pos18 not in pos10 and pos18 not in pos11 and pos18 not in pos12 and
                        pos18 not in pos13 and pos18 not in pos14 and pos18 not in pos15 and
                        pos18 not in pos16 and pos18 not in pos17):
                    hrsPos18 += list18[2]
                    RegHrs18 += list18[3]
                    totOT12_18 += list18[4]
                    totOT40_18 += list18[5]
                list18 = [valName, pos18, hrsPos18, RegHrs18, totOT12_18, totOT40_18, list18[6], list18[7], 0]


            if (    list2[1] is not None and (list2[1] is list1[1])):
                list2 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list3[1] is not None and (list3[1] is list1[1] or list3[1] is list2[1])):
                list3 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list4[1] is not None and (list4[1] is list1[1] or list4[1] is list2[1] or list4[1] is list3[1])):
                list4 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list5[1] is not None and (list5[1] is list1[1] or list5[1] is list2[1] or list5[1] is list3[1] or
                    list5[1] is list4[1])):
                list5 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list6[1] is not None and (list6[1] is list1[1] or list6[1] is list2[1] or list6[1] is list3[1] or
                    list6[1] is list4[1] or list6[1] is list5[1])):
                list6 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list7[1] is not None and (list7[1] is list1[1] or list7[1] is list2[1] or list7[1] is list3[1] or
                    list7[1] is list4[1] or list7[1] is list5[1] or list7[1] is list6[1])):
                list7 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list8[1] is not None and (list8[1] is list1[1] or list8[1] is list2[1] or list8[1] is list3[1] or
                    list8[1] is list4[1] or list8[1] is list5[1] or list8[1] is list6[1] or list8[1] is list7[1])):
                list8 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list9[1] is not None and (list9[1] is list1[1] or list9[1] is list2[1] or list9[1] is list3[1] or
                    list9[1] is list4[1] or list9[1] is list5[1] or list9[1] is list6[1] or list9[1] is list7[1] or
                    list9[1] is list8[1])):
                list9 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list10[1] is not None and (list10[1] is list1[1] or list10[1] is list2[1] or list10[1] is list3[1] or
                    list10[1] is list4[1] or list10[1] is list5[1] or list10[1] is list6[1] or list10[1] is list7[1] or
                    list10[1] is list8[1] or list10[1] is list9[1])):
                list10 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list11[1] is not None and (list11[1] is list1[1] or list11[1] is list2[1] or list11[1] is list3[1] or
                    list11[1] is list4[1] or list11[1] is list5[1] or list11[1] is list6[1] or list11[1] is list7[1] or
                    list11[1] is list8[1] or list11[1] is list9[1] or list11[1] is list10[1])):
                list11 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list12[1] is not None and (list12[1] is list1[1] or list12[1] is list2[1] or list12[1] is list3[1] or
                    list12[1] is list4[1] or list12[1] is list5[1] or list12[1] is list6[1] or list12[1] is list7[1] or
                    list12[1] is list8[1] or list12[1] is list9[1] or list12[1] is list10[1] or list12[1] is list11[1])):
                list12 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list13[1] is not None and (list13[1] is list1[1] or list13[1] is list2[1] or list13[1] is list3[1] or
                    list13[1] is list4[1] or list13[1] is list5[1] or list13[1] is list6[1] or list13[1] is list7[1] or
                    list13[1] is list8[1] or list13[1] is list9[1] or list13[1] is list10[1] or list13[1] is list11[1] or
                    list13[1] is list12[1])):
                list13 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list14[1] is not None and (list14[1] is list1[1] or list14[1] is list2[1] or list14[1] is list3[1] or
                    list14[1] is list4[1] or list14[1] is list5[1] or list14[1] is list6[1] or list14[1] is list7[1] or
                    list14[1] is list8[1] or list14[1] is list9[1] or list14[1] is list10[1] or list14[1] is list11[1] or
                    list14[1] is list12[1] or list14[1] is list13[1])):
                list14 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list15[1] is not None and (list15[1] is list1[1] or list15[1] is list2[1] or list15[1] is list3[1] or
                    list15[1] is list4[1] or list15[1] is list5[1] or list15[1] is list6[1] or list15[1] is list7[1] or
                    list15[1] is list8[1] or list15[1] is list9[1] or list15[1] is list10[1] or list15[1] is list11[1] or
                    list15[1] is list12[1] or list15[1] is list13[1] or list15[1] is list14[1])):
                list15 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list16[1] is not None and (list16[1] is list1[1] or list16[1] is list2[1] or list16[1] is list3[1] or
                    list16[1] is list4[1] or list16[1] is list5[1] or list16[1] is list6[1] or list16[1] is list7[1] or
                    list16[1] is list8[1] or list16[1] is list9[1] or list16[1] is list10[1] or list16[1] is list11[1] or
                    list16[1] is list12[1] or list16[1] is list13[1] or list16[1] is list14[1])):
                list16 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list17[1] is not None and (list17[1] is list1[1] or list17[1] is list2[1] or list17[1] is list3[1] or
                    list17[1] is list4[1] or list17[1] is list5[1] or list17[1] is list6[1] or list17[1] is list7[1] or
                    list17[1] is list8[1] or list17[1] is list9[1] or list17[1] is list10[1] or list17[1] is list11[1] or
                    list17[1] is list12[1] or list17[1] is list13[1] or list17[1] is list14[1])):
                list17 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list18[1] is not None and (list18[1] is list1[1] or list18[1] is list2[1] or list18[1] is list3[1] or
                    list18[1] is list4[1] or list18[1] is list5[1] or list18[1] is list6[1] or list18[1] is list7[1] or
                    list18[1] is list8[1] or list18[1] is list9[1] or list18[1] is list10[1] or list18[1] is list11[1] or
                    list18[1] is list12[1] or list18[1] is list13[1] or list18[1] is list14[1])):
                list18 = [None, None, 0, 0, 0, 0, 0, None, 0]

            if list17[1] is None and list18[1] is not None:
                list17 = list18
                list18 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list16[1] is None and list17[1] is not None:
                list16 = list17
                list17 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list15[1] is None and list16[1] is not None:
                list15 = list16
                list16 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list14[1] is None and list15[1] is not None:
                list14 = list15
                list15 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list13[1] is None and list14[1] is not None:
                list13 = list14
                list14 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list12[1] is None and list13[1] is not None:
                list12 = list13
                list13 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list11[1] is None and list12[1] is not None:
                list11 = list12
                list12 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list10[1] is None and list11[1] is not None:
                list10 = list11
                list11 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list9[1] is None and list10[1] is not None:
                list9 = list10
                list10 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list8[1] is None and list9[1] is not None:
                list8 = list9
                list9 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list7[1] is None and list8[1] is not None:
                list7 = list8
                list8 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list6[1] is None and list7[1] is not None:
                list6 = list7
                list7 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list5[1] is None and list6[1] is not None:
                list5 = list6
                list6 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list4[1] is None and list5[1] is not None:
                list4 = list5
                list5 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list3[1] is None and list4[1] is not None:
                list3 = list4
                list4 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list2[1] is None and list3[1] is not None:
                list2 = list3
                list3 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if j == 0: listWeek1 = [list1, list2, list3, list4, list5, list6, list7, list8, list9, list10, list11, list12, list13, list14, list15, list16, list17, list18]
            else:      listWeek2 = [list1, list2, list3, list4, list5, list6, list7, list8, list9, list10, list11, list12, list13, list14, list15, list16, list17, list18]

        if flagDebug2:
            print("\n", valName)
            print("--Week 1 totals--")
            print("    Position:\t\t  total hrs:   reg hrs:   OT12:   OT40:   Rate:")
            for j in range(0, 17):
                if listWeek1[j][0] is not None:
                    if (listWeek1[j][2] != 0. or listWeek1[j][5] != 0.):
                        print(listWeek1[j][1], " ", listWeek1[j][2], " ", listWeek1[j][3], " ", listWeek1[j][4], " ", listWeek1[j][5], " ", listWeek1[j][6])
            print("--Week 2 totals--")
            print("    Position:\t\t  total hrs:   reg hrs:   OT12:   OT40:   Rate:")
            for j in range(0, 17):
                if listWeek2[j][0] is not None:
                    if (listWeek2[j][2] != 0. or listWeek2[j][5] != 0.):
                        print(listWeek2[j][1], " ", listWeek2[j][2], " ", listWeek2[j][3], " ", listWeek2[j][4], " ", listWeek2[j][5], " ", listWeek2[j][6])


        # Print values to .xlsx file
        line = Side(border_style="thin", color="000000")
        c0 = newsheet1.cell(row = printCnt, column = 1)
        c0.value = listWeek1[0][0]
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt, column = 2)
        c0.border = Border(bottom = line)
        c0.value = "Rate"
        c0 = newsheet1.cell(row = printCnt, column = 3)
        c0.value = "Regular"
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt, column = 4)
        c0.value = "OT+12"
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt, column = 5)
        c0.value = "OT+40"
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt, column = 6)
        c0.value = "OT Total"
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt, column = 7)
        c0.value = "Emp # "
        c0.border = Border(bottom = line)
        c0.alignment = Alignment(horizontal = 'right')
        c0 = newsheet1.cell(row = printCnt, column = 8)
        if listWeek1[0][7] is not None:
            c0.value = listWeek1[0][7]
        else: c0.value = listWeek2[0][7]
        c0.border = Border(bottom = line)
        c0.alignment = Alignment(horizontal = 'left')
        c0 = newsheet1.cell(row = printCnt + 1, column = 1)
        c0.value = "---Week 1---"
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0 = newsheet1.cell(row = printCnt + 1, column = 2)
        c0.border = Border(right = line)
        c0 = newsheet1.cell(row = printCnt + 2, column = 1)
        if listWeek1[0][1] is None: c0.value = "No shifts"
        else: c0.value = listWeek1[0][1]
        c0.border = Border(right = line)
        c0 = newsheet1.cell(row = printCnt + 2, column = 2)
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0.value = listWeek1[0][6]
        c0 = newsheet1.cell(row = printCnt + 2, column = 3)
        c0.value = listWeek1[0][3]
        c0 = newsheet1.cell(row = printCnt + 2, column = 4)
        c0.value = listWeek1[0][4]
        c0 = newsheet1.cell(row = printCnt + 2, column = 5)
        c0.value = listWeek1[0][5]
        c0 = newsheet1.cell(row = printCnt + 2, column = 6)
        c0.value = listWeek1[0][4] + listWeek1[0][5]
        c0 = newsheet1.cell(row = printCnt + 2, column = 9)
        if flagMultShift1 is True:
            c0 = newsheet1.cell(row = printCnt + 2, column = 10)
            c0.value = "Multiple shifts in same day for WEEK 1. Check for OT+12 by hand."
        if listWeek1[1][0] is None: printCnt += 4
        if listWeek1[1][0] is not None:
            c0 = newsheet1.cell(row = printCnt + 3, column = 1)
            c0.value = listWeek1[1][1]
            c0.border = Border(right = line)
            c0 = newsheet1.cell(row = printCnt + 3, column = 2)
            c0.border = Border(right = line)
            c0.font = Font(italic = 'single')
            c0.value = listWeek1[1][6]
            c0 = newsheet1.cell(row = printCnt + 3, column = 3)
            c0.value = listWeek1[1][3]
            c0 = newsheet1.cell(row = printCnt + 3, column = 4)
            c0.value = listWeek1[1][4]
            c0 = newsheet1.cell(row = printCnt + 3, column = 5)
            c0.value = listWeek1[1][5]
            c0 = newsheet1.cell(row = printCnt + 3, column = 6)
            c0.value = listWeek1[1][4] + listWeek1[1][5]
            c0 = newsheet1.cell(row = printCnt + 3, column = 9)
            printCnt += 5
            for j in range(2, 15):
                if listWeek1[j][0] is not None:
                    if (listWeek1[j][2] != 0. or listWeek1[j][5] != 0.):
                        c0 = newsheet1.cell(row = printCnt - 1, column = 1)
                        c0.value = listWeek1[j][1]
                        c0.border = Border(right = line)
                        c0 = newsheet1.cell(row = printCnt - 1, column = 2)
                        c0.border = Border(right = line)
                        c0.font = Font(italic = 'single')
                        c0.value = listWeek1[j][6]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 3)
                        c0.value = listWeek1[j][3]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 4)
                        c0.value = listWeek1[j][4]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 5)
                        c0.value = listWeek1[j][5]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 6)
                        c0.value = listWeek1[j][4] + listWeek1[j][5]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 9)
                        printCnt += 1

        c0 = newsheet1.cell(row = printCnt - 1, column = 1)
        c0.value = "---Week 2---"
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0 = newsheet1.cell(row = printCnt - 1, column = 2)
        c0.border = Border(right = line)
        c0 = newsheet1.cell(row = printCnt, column = 1)
        if listWeek2[0][1] is None: c0.value = "No shifts"
        else: c0.value = listWeek2[0][1]
        c0.border = Border(right = line)
        c0 = newsheet1.cell(row = printCnt, column = 2)
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0.value = listWeek2[0][6]
        c0 = newsheet1.cell(row = printCnt, column = 3)
        c0.value = listWeek2[0][3]
        c0 = newsheet1.cell(row = printCnt, column = 4)
        c0.value = listWeek2[0][4]
        c0 = newsheet1.cell(row = printCnt, column = 5)
        c0.value = listWeek2[0][5]
        c0 = newsheet1.cell(row = printCnt, column = 6)
        c0.value = listWeek2[0][4] + listWeek2[0][5]
        c0 = newsheet1.cell(row = printCnt, column = 9)
        if flagMultShift2 is True:
            c0 = newsheet1.cell(row = printCnt, column = 10)
            c0.value = "Multiple shifts in same day for WEEK 2. Check for OT+12 by hand."
        if listWeek2[1][0] is None: printCnt += 2
        if listWeek2[1][0] is not None:
            c0 = newsheet1.cell(row = printCnt + 1, column = 1)
            c0.value = listWeek2[1][1]
            c0.border = Border(right = line)
            c0 = newsheet1.cell(row = printCnt + 1, column = 2)
            c0.border = Border(right = line)
            c0.font = Font(italic = 'single')
            c0.value = listWeek2[1][6]
            c0 = newsheet1.cell(row = printCnt + 1, column = 3)
            c0.value = listWeek2[1][3]
            c0 = newsheet1.cell(row = printCnt + 1, column = 4)
            c0.value = listWeek2[1][4]
            c0 = newsheet1.cell(row = printCnt + 1, column = 5)
            c0.value = listWeek2[1][5]
            c0 = newsheet1.cell(row = printCnt + 1, column = 6)
            c0.value = listWeek2[1][4] + listWeek2[1][5]
            c0 = newsheet1.cell(row = printCnt + 1, column = 9)
            printCnt += 3
            for j in range(2, 15):
                if listWeek2[j][0] is not None:
                    if (listWeek2[j][2] != 0. or listWeek2[j][5] != 0.):
                        c0 = newsheet1.cell(row = printCnt - 1, column = 1)
                        c0.value = listWeek2[j][1]
                        c0.border = Border(right = line)
                        c0 = newsheet1.cell(row = printCnt - 1, column = 2)
                        c0.border = Border(right = line)
                        c0.font = Font(italic = 'single')
                        c0.value = listWeek2[j][6]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 3)
                        c0.value = listWeek2[j][3]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 4)
                        c0.value = listWeek2[j][4]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 5)
                        c0.value = listWeek2[j][5]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 6)
                        c0.value = listWeek2[j][4] + listWeek2[j][5]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 9)
                        printCnt += 1

        # Add up both weeks
        xreg = 0
        xOT = 0
        c0 = newsheet1.cell(row = printCnt - 1, column = 1)
        c0.value = "---Total (reg, OT, shifts)---"
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0 = newsheet1.cell(row = printCnt - 1, column = 2)
        c0.border = Border(right = line, bottom = line)
        c0 = newsheet1.cell(row = printCnt - 1, column = 3)
        c0.value = "tot reg"
        c0.font = Font(italic = 'single')
        c0.alignment = Alignment(horizontal = 'right')
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt - 1, column = 4)
        c0.value = "tot OT"
        c0.font = Font(italic = 'single')
        c0.alignment = Alignment(horizontal = 'right')
        c0.border = Border(bottom = line)
        for j in range(0, 15):
            if listWeek1[j][1] is not None:
                if (listWeek1[j][2] != 0. or listWeek1[j][5] != 0.):
                    xreg = listWeek1[j][3]
                    xOT  = listWeek1[j][4] + listWeek1[j][5]
                    if listWeek2[0][1] is not None and listWeek2[0][1] in listWeek1[j][1]:
                        xreg += listWeek2[0][3]
                        xOT  += listWeek2[0][4] + listWeek2[0][5]
                    if listWeek2[1][1] is not None and listWeek2[1][1] in listWeek1[j][1]:
                        xreg += listWeek2[1][3]
                        xOT  += listWeek2[1][4] + listWeek2[1][5]
                    if listWeek2[2][1] is not None and listWeek2[2][1] in listWeek1[j][1]:
                        xreg += listWeek2[2][3]
                        xOT  += listWeek2[2][4] + listWeek2[2][5]
                    if listWeek2[3][1] is not None and listWeek2[3][1] in listWeek1[j][1]:
                        xreg += listWeek2[3][3]
                        xOT  += listWeek2[3][4] + listWeek2[3][5]
                    if listWeek2[4][1] is not None and listWeek2[4][1] in listWeek1[j][1]:
                        xreg += listWeek2[4][3]
                        xOT  += listWeek2[4][4] + listWeek2[4][5]
                    if listWeek2[5][1] is not None and listWeek2[5][1] in listWeek1[j][1]:
                        xreg += listWeek2[5][3]
                        xOT  += listWeek2[5][4] + listWeek2[5][5]
                    if listWeek2[6][1] is not None and listWeek2[6][1] in listWeek1[j][1]:
                        xreg += listWeek2[6][3]
                        xOT  += listWeek2[6][4] + listWeek2[6][5]
                    if listWeek2[7][1] is not None and listWeek2[7][1] in listWeek1[j][1]:
                        xreg += listWeek2[7][3]
                        xOT  += listWeek2[7][4] + listWeek2[7][5]
                    if listWeek2[8][1] is not None and listWeek2[8][1] in listWeek1[j][1]:
                        xreg += listWeek2[8][3]
                        xOT  += listWeek2[8][4] + listWeek2[8][5]
                    if listWeek2[9][1] is not None and listWeek2[9][1] in listWeek1[j][1]:
                        xreg += listWeek2[9][3]
                        xOT  += listWeek2[9][4] + listWeek2[9][5]
                    if listWeek2[10][1] is not None and listWeek2[10][1] in listWeek1[j][1]:
                        xreg += listWeek2[10][3]
                        xOT  += listWeek2[10][4] + listWeek2[10][5]
                    if listWeek2[11][1] is not None and listWeek2[11][1] in listWeek1[j][1]:
                        xreg += listWeek2[11][3]
                        xOT  += listWeek2[11][4] + listWeek2[11][5]
                    if listWeek2[12][1] is not None and listWeek2[12][1] in listWeek1[j][1]:
                        xreg += listWeek2[12][3]
                        xOT  += listWeek2[12][4] + listWeek2[12][5]
                    if listWeek2[13][1] is not None and listWeek2[13][1] in listWeek1[j][1]:
                        xreg += listWeek2[13][3]
                        xOT  += listWeek2[13][4] + listWeek2[13][5]
                    if listWeek2[14][1] is not None and listWeek2[14][1] in listWeek1[j][1]:
                        xreg += listWeek2[14][3]
                        xOT  += listWeek2[14][4] + listWeek2[14][5]
                    if listWeek2[15][1] is not None and listWeek2[15][1] in listWeek1[j][1]:
                        xreg += listWeek2[15][3]
                        xOT  += listWeek2[15][4] + listWeek2[15][5]
                    if listWeek2[16][1] is not None and listWeek2[16][1] in listWeek1[j][1]:
                        xreg += listWeek2[16][3]
                        xOT  += listWeek2[16][4] + listWeek2[16][5]
                    if listWeek2[17][1] is not None and listWeek2[17][1] in listWeek1[j][1]:
                        xreg += listWeek2[17][3]
                        xOT  += listWeek2[17][4] + listWeek2[17][5]
                    c0 = newsheet1.cell(row = printCnt, column = 1)
                    c0.value = listWeek1[j][1]
                    c0.border = Border(right = line)
                    c0 = newsheet1.cell(row = printCnt, column = 2)
                    c0.border = Border(right = line)
                    c0.font = Font(italic = 'single')
                    c0.value = listWeek1[j][6]
                    c0 = newsheet1.cell(row = printCnt, column = 3)
                    c0.font = Font(bold = 'single')
                    c0.value = xreg
                    c0 = newsheet1.cell(row = printCnt, column = 4)
                    c0.font = Font(bold = 'single')
                    c0.value = xOT
                    printCnt += 1
        xreg = 0
        xOT = 0
        for j in range(0, 15):
            if listWeek2[j][1] is not None:
                if (listWeek2[j][2] != 0. or listWeek2[j][5] != 0.):
                    if   (listWeek1[0][1]  is not None and listWeek2[j][1].__contains__(listWeek1[0][1]) ):  continue
                    elif (listWeek1[1][1]  is not None and listWeek2[j][1].__contains__(listWeek1[1][1]) ):  continue
                    elif (listWeek1[2][1]  is not None and listWeek2[j][1].__contains__(listWeek1[2][1]) ):  continue
                    elif (listWeek1[3][1]  is not None and listWeek2[j][1].__contains__(listWeek1[3][1]) ):  continue
                    elif (listWeek1[4][1]  is not None and listWeek2[j][1].__contains__(listWeek1[4][1]) ):  continue
                    elif (listWeek1[5][1]  is not None and listWeek2[j][1].__contains__(listWeek1[5][1]) ):  continue
                    elif (listWeek1[6][1]  is not None and listWeek2[j][1].__contains__(listWeek1[6][1]) ):  continue
                    elif (listWeek1[7][1]  is not None and listWeek2[j][1].__contains__(listWeek1[7][1]) ):  continue
                    elif (listWeek1[8][1]  is not None and listWeek2[j][1].__contains__(listWeek1[8][1]) ):  continue
                    elif (listWeek1[9][1]  is not None and listWeek2[j][1].__contains__(listWeek1[9][1]) ):  continue
                    elif (listWeek1[10][1] is not None and listWeek2[j][1].__contains__(listWeek1[10][1]) ): continue
                    elif (listWeek1[11][1] is not None and listWeek2[j][1].__contains__(listWeek1[11][1]) ): continue
                    elif (listWeek1[12][1] is not None and listWeek2[j][1].__contains__(listWeek1[12][1]) ): continue
                    elif (listWeek1[13][1] is not None and listWeek2[j][1].__contains__(listWeek1[13][1]) ): continue
                    elif (listWeek1[14][1] is not None and listWeek2[j][1].__contains__(listWeek1[14][1]) ): continue
                    elif (listWeek1[15][1] is not None and listWeek2[j][1].__contains__(listWeek1[15][1]) ): continue
                    elif (listWeek1[16][1] is not None and listWeek2[j][1].__contains__(listWeek1[16][1]) ): continue
                    elif (listWeek1[17][1] is not None and listWeek2[j][1].__contains__(listWeek1[17][1]) ): continue
                    else:
                        xreg = listWeek2[j][3]
                        xOT  = listWeek2[j][4] + listWeek2[j][5]
                        c0 = newsheet1.cell(row = printCnt, column = 1)
                        c0.value = listWeek2[j][1]
                        c0.border = Border(right = line)
                        c0 = newsheet1.cell(row = printCnt, column = 2)
                        c0.border = Border(right = line)
                        c0.font = Font(italic = 'single')
                        c0.value = listWeek2[j][6]
                        c0 = newsheet1.cell(row = printCnt, column = 3)
                        c0.font = Font(bold = 'single')
                        c0.value = xreg
                        c0 = newsheet1.cell(row = printCnt, column = 4)
                        c0.font = Font(bold = 'single')
                        c0.value = xOT
                        printCnt += 1
        printCnt += 2
        rowCnt = 1 # Reset count parameter for next employee

# Check for empty rows
for i in range(1, Nrow):
    c0 = newsheet1.cell(row = i, column = 1).value
    c1 = newsheet1.cell(row = i, column = 1).value
    c2 = newsheet1.cell(row = i, column = 3).value
    c3 = newsheet1.cell(row = i, column = 4).value
    c4 = newsheet1.cell(row = i, column = 5).value
    c5 = newsheet1.cell(row = i, column = 6).value
    if c0 is c1 and c2 == 0 and c3 == 0 and c4 == 0 and c5 == 0:
        newsheet1.delete_rows(i, 1)


# More headers and design options for output
c0 = newsheet1.cell(row = 1, column = 1)
c0.value = "Week 1 dates:  "
c0.font = Font(bold = 'single')
c0 = newsheet1.cell(row = 1, column = 2)
c0.value = week1start.strftime("%Y-%m-%d") + "  to  " + week1end.strftime("%Y-%m-%d")
c0 = newsheet1.cell(row = 2, column = 1)
c0.value = "Week 2 dates:  "
c0.font = Font(bold = 'single')
c0 = newsheet1.cell(row = 2, column = 2)
c0.value = week2start.strftime("%Y-%m-%d") + "  to  " + week2end.strftime("%Y-%m-%d")
c0 = newsheet1.cell(row = 3, column = 1)
c0.value = "Total number of employees:  "
c0.font = Font(bold = 'single')
c0 = newsheet1.cell(row = 3, column = 2)
c0.value = GrandNames
c0 = newsheet1.cell(row = 4, column = 1)
c0.value = "Total number of shifts:  "
c0.font = Font(bold = 'single')
c0 = newsheet1.cell(row = 4, column = 2)
c0.value = Nrow - 1
c0 = newsheet1.cell(row = 5, column = 1)
c0.value = "Total number of hours:  "
c0.font = Font(bold = 'single')
c0 = newsheet1.cell(row = 5, column = 2)
c0.value = GrandHrs
c0 = newsheet1.cell(row = 8, column = 10)
c0.value = "Notes:"

c0 = newsheet1.cell(row = 1, column = 1)
c0.alignment = Alignment(horizontal='right')
c0 = newsheet1.cell(row = 2, column = 1)
c0.alignment = Alignment(horizontal='right')
c0 = newsheet1.cell(row = 3, column = 1)
c0.alignment = Alignment(horizontal='right')
c0 = newsheet1.cell(row = 4, column = 1)
c0.alignment = Alignment(horizontal='right')
c0 = newsheet1.cell(row = 5, column = 1)
c0.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["A"].width = 48
newsheet1.column_dimensions["B"].width = 8
newsheet1.column_dimensions["C"].width = 8
newsheet1.column_dimensions["D"].width = 8

output_name = "output/OT_calculation_details.xlsx"
newbook1.save(output_name)
print("\n")
print("====================================================")
print("---WARNING! Sum up EXPRESS 1MGR hours by hand!---")
print("====================================================\n")
if flagMultShiftTot is True: print("====================================================")
if flagMultShiftTot is True: print("---WARNING! Hand-check days with multiple shifts!---")
if flagMultShiftTot is True: print("====================================================\n")
print("\nOT calculation completed. Check OT+12 for days with multiple shifts.\n")
print("File output written to", output_name,"\n")
print("--Step 2 complete--\n")


###----------------------------------------------------------------------------------------------###


# Code for organizing payroll overtime calculation.
# Reads in calc_details.xlsx

# Source files
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Alignment, Border, Side, Font
from datetime import datetime, date, timedelta

flagDebug = False

# Input file
workbook = load_workbook('output/OT_calculation_details.xlsx')
sheet = workbook.active
for i in range(1, 10000):
    val0 = sheet.cell(row = i, column = 1).value
    val1 = sheet.cell(row = i + 1, column = 1).value
    val2 = sheet.cell(row = i + 2, column = 1).value
    if val0 is None and val1 is None and val2 is None:
        continue
    Nrow = i

# Setup output file and spreadsheet
newbook1 = openpyxl.Workbook()
newsheet1 = newbook1.active

print("\n\n\n")
print("=======================================")
print("   Step 3: Merging final spreadsheet   ")
print("=======================================")
print("\n")
print("\nCreating final spreadsheet...\n")

for i in range(8, Nrow + 1):
    val0 = sheet.cell(row = i - 1, column = 1).value
    val1 = sheet.cell(row = i, column = 1).value
    if val1 is not None:
        if val0 is None:
            valName = val1
            valENum = sheet.cell(row = i, column = 8).value
            valNote = sheet.cell(row = i, column = 10).value
        if "---Total" in val1:
            cnt = 0
            for j in range(0, 12):
                pos  = sheet.cell(row = i + 1 + j, column = 1).value
                reg  = sheet.cell(row = i + 1 + j, column = 3).value
                ot   = sheet.cell(row = i + 1 + j, column = 4).value
                rate = sheet.cell(row = i + 1 + j, column = 2).value
                if pos is None: break
                else:
                    c0 = newsheet1.cell(row = i + 1 + cnt, column = 2)
                    c0.value = valName
                    c0 = newsheet1.cell(row = i + 1 + cnt, column = 3)
                    c0.value = pos
                    if valENum is not None:
                        c4 = newsheet1.cell(row = i + 1 + cnt, column = 1)
                        c4.value = valENum
                    if valNote is not None and "Check OT" in valNote:
                        c5 = newsheet1.cell(row = i + 1 + cnt, column = 11)
                        c5.value = "Check OT+12 by hand"
                        # remove redundant flag for OT+12 calculation
                        for k in range(0, 12):
                            c6 = newsheet1.cell(row = i + cnt - k, column = 11)
                            c7 = newsheet1.cell(row = i + cnt - k, column = 2)
                            if c6.value is not None and c7.value is valName:
                                c5.value = " "
                    cat = "Unarmed"
                    if pos.__contains__("Training"): cat = "Training"
                    elif pos.__contains__("Admin Work"): cat = "Admin"
                    elif pos.__contains__("ARMED") or pos.__contains__("Armed"): cat = "Armed"
                    elif pos.__contains__("Sick"): cat = "Sick"
                    elif pos.__contains__("Covid") or pos.__contains__("COVID"): cat = "Covid"
                    if reg != 0 and ot == 0:
                        c0 = newsheet1.cell(row = i + 1 + cnt, column = 6)
                        c0.value = reg
                        c1 = newsheet1.cell(row = i + 1 + cnt, column = 4)
                        c1.value = cat
                        c2 = newsheet1.cell(row = i + 1 + cnt, column = 7)
                        c2.value = rate
                        c3 = newsheet1.cell(row = i + 1 + cnt, column = 8)
                        c3.value = c0.value*c2.value
                        cnt += 1
                    if reg == 0 and ot != 0:
                        c0 = newsheet1.cell(row = i + 1 + cnt, column = 6)
                        c0.value = ot
                        if pos.__contains__("Training"): cat = "Training"
                        if cat.__contains__("Unarmed"): cat = "OT Unarmed"
                        if pos.__contains__("Admin"): cat = "OT Admin"
                        if pos.__contains__("ARMED") or pos.__contains__("Armed"): cat = "OT Armed"
                        c1 = newsheet1.cell(row = i + 1 + cnt, column = 4)
                        c1.value = cat
                        c2 = newsheet1.cell(row = i + 1 + cnt, column = 7)
                        c2.value = rate * 1.5
                        c3 = newsheet1.cell(row = i + 1 + cnt, column = 8)
                        c3.value = c0.value * c2.value
                        cnt += 1
                    if reg != 0 and ot != 0:
                        c0 = newsheet1.cell(row = i + 1 + cnt, column = 6)
                        c0.value = reg
                        c1 = newsheet1.cell(row = i + 1 + cnt, column = 4)
                        c1.value = cat
                        c2 = newsheet1.cell(row = i + 1 + cnt, column = 7)
                        c2.value = rate
                        c3 = newsheet1.cell(row = i + 1 + cnt, column = 8)
                        c3.value = c0.value*c2.value
                        c0 = newsheet1.cell(row = i + 2 + cnt, column = 2)
                        c0.value = valName
                        c0 = newsheet1.cell(row = i + 2 + cnt, column = 3)
                        c0.value = pos
                        c0 = newsheet1.cell(row = i + 2 + cnt, column = 6)
                        c0.value = ot
                        if valENum is not None:
                            c4 = newsheet1.cell(row = i + 2 + cnt, column = 1)
                            c4.value = valENum
                        if pos.__contains__("Training"): cat = "Training"
                        if cat.__contains__("Unarmed"): cat = "OT Unarmed"
                        if pos.__contains__("Admin"): cat = "OT Admin"
                        if pos.__contains__("ARMED") or pos.__contains__("Armed"): cat = "OT Armed"
                        c1 = newsheet1.cell(row = i + 2 + cnt, column = 4)
                        c1.value = cat
                        c2 = newsheet1.cell(row = i + 2 + cnt, column = 7)
                        c2.value = rate * 1.5
                        c3 = newsheet1.cell(row = i + 2 + cnt, column = 8)
                        c3.value = c0.value*c2.value
                        cnt += 2


c0 = newsheet1.cell(row = 1, column = 1)
c0.value = "Employee Number"
c0 = newsheet1.cell(row = 1, column = 2)
c0.value = "Employee Name"
c0 = newsheet1.cell(row = 1, column = 3)
c0.value = "Position"
c0 = newsheet1.cell(row = 1, column = 4)
c0.value = "Category"
c0 = newsheet1.cell(row = 1, column = 5)
c0.value = "Number of Shifts"
c0 = newsheet1.cell(row = 1, column = 6)
c0.value = "Number of Hours"
c0 = newsheet1.cell(row = 1, column = 7)
c0.value = "Pay Rate"
c0 = newsheet1.cell(row = 1, column = 8)
c0.value = "Total Pay"
c0 = newsheet1.cell(row = 1, column = 9)
c0.value = "Reimbursement/Deductions"
c0 = newsheet1.cell(row = 1, column = 10)
c0.value = "Grand Total"
c0 = newsheet1.cell(row = 1, column = 11)
c0.value = "Notes"

newsheet1.column_dimensions["A"].width = 20
newsheet1.column_dimensions["B"].width = 25
newsheet1.column_dimensions["C"].width = 48
newsheet1.column_dimensions["D"].width = 12
newsheet1.column_dimensions["E"].width = 12
newsheet1.column_dimensions["F"].width = 14
newsheet1.column_dimensions["G"].width = 12
newsheet1.column_dimensions["H"].width = 12
newsheet1.column_dimensions["I"].width = 12
newsheet1.column_dimensions["J"].width = 12
newsheet1.column_dimensions["K"].width = 12

# Delete empty rows
indx = []
for i in range(len(tuple(newsheet1.rows))):
    flag = False
    for cell in tuple(newsheet1.rows)[i]:
        if cell.value != None:
            flag = True
            break
    if flag == False:
        indx.append(i)
indx.sort()
for i in range(len(indx)):
    newsheet1.delete_rows(idx = indx[i] + 1 - i)

# Clean-up extra notes
for i in range(8, Nrow + 1):
    c0 = newsheet1.cell(row = i, column = 1)
    if c0 is None:
        newsheet.delete_rows(i)

output_name = "output/OT_calculation_final.xlsx"
newbook1.save(output_name)
print("\nSpreadsheet informarion merged.\n")
print("File output written to", output_name, "\n")
print("--Step 3 complete. OT calculation finished.--\n\n")
