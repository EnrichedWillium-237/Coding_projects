### Code for reading in a payroll spreadsheet and producing final output for Paylocity submission ###

# Source files
import openpyxl
import array as arr
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
import sys
import os

# Payroll file location
workbook = load_workbook('payroll_with_OT.xlsx')

# Payroll output
newbook1 = openpyxl.Workbook()
newsheet1 = newbook1.active
sheet = workbook.active

Nrow = sheet.max_row
Ncell = 1 # index for excel output

c0 = newsheet1.cell(row = 1, column = 1)
c0.value = 0
c0 = newsheet1.cell(row = 1, column = 2)
c0.value = "E"
c0 = newsheet1.cell(row = 1, column = 3)
c0.value = "UNARM"
c0 = newsheet1.cell(row = 1, column = 4)
c0.value = 0
c0 = newsheet1.cell(row = 1, column = 5)
c0.value = 0
c0 = newsheet1.cell(row = 1, column = 6)
c0.value = 0
n = 0
for i in range (2, Nrow + 1):
    valID = sheet.cell(row = i, column = 1)
    valName = sheet.cell(row = i, column = 2)
    valCat = sheet.cell(row = i, column = 3)
    valHours = sheet.cell(row = i, column = 4)
    valRate = sheet.cell(row = i, column = 5)
    valTot = sheet.cell(row = i, column = 6)
    valReim = sheet.cell(row = i, column = 7)
    valGross = sheet.cell(row = i, column = 8)
    valNote = sheet.cell(row = i, column = 9)

    c0 = newsheet1.cell(row = i + n, column = 1)
    c0.value = valID.value
    c0 = newsheet1.cell(row = i + n, column = 2)
    c0.value = "E"
    if (valCat.value == "Unarmed"):
        c0 = newsheet1.cell(row = i + n, column = 3)
        c0.value = "UNARM"
    if (valCat.value == "Armed"):
        c0 = newsheet1.cell(row = i + n, column = 3)
        c0.value = "ARMED"
    if (valCat.value == "Training"):
        c0 = newsheet1.cell(row = i + n, column = 3)
        c0.value = "REG"
    if (valCat.value == "Admin"):
        c0 = newsheet1.cell(row = i + n, column = 3)
        c0.value = "REG"
    if (valCat.value == "OT Admin"):
        c0 = newsheet1.cell(row = i + n, column = 3)
        c0.value = "REG"
    if (valCat.value == "Sick"):
        c0 = newsheet1.cell(row = i + n, column = 3)
        c0.value = "SICK"
    if (valCat.value == "OT Unarmed"):
        c0 = newsheet1.cell(row = i + n, column = 3)
        c0.value = "OTUNA"
    if (valCat.value == "OT Armed"):
        c0 = newsheet1.cell(row = i + n, column = 3)
        c0.value = "OTARM"
    if (valCat.value == "Holiday"):
        c0 = newsheet1.cell(row = i + n, column = 3)
        c0.value = "HOL"

    c0 = newsheet1.cell(row = i + n, column = 4)
    c0.value = valHours.value
    c0 = newsheet1.cell(row = i + n, column = 5)
    c0.value = valTot.value
    c0 = newsheet1.cell(row = i + n, column = 6)
    c0.value = valRate.value

    if (valReim.value is not None):
        if (valReim.value is not None and valReim != 0):
            n += 1
            c0 = newsheet1.cell(row = i + n, column = 1)
            c0.value = valID.value
            c0 = newsheet1.cell(row = i + n, column = 2)
            c0.value = "D"
            c0 = newsheet1.cell(row = i + n, column = 4)
            c0.value = 0
            c0 = newsheet1.cell(row = i + n, column = 6)
            c0.value = 0
            if (valReim.value > 0):
                c0 = newsheet1.cell(row = i + n, column = 3)
                c0.value = "REIMB"
                c0 = newsheet1.cell(row = i + n, column = 5)
                c0.value = valReim.value
            if (valReim.value < 0):
                c0 = newsheet1.cell(row = i + n, column = 3)
                c0.value = "ADVNC"
                c0 = newsheet1.cell(row = i + n, column = 5)
                c0.value = -1.*valReim.value

output_name = "output/final_payroll_for_paylocity.csv"
newbook1.save(output_name)
print("\nFinal spreadsheet ready for Paylocity submission.\n")
print("File output written to", output_name, "\n")
