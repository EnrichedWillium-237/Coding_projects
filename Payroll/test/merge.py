# Code for organizing payroll overtime calculation.
# Reads in calc_details.xlsx

# Source files
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Alignment, Border, Side, Font
from datetime import datetime, date, timedelta

flagDebug = False

# Input file
workbook = load_workbook('output/OT_calculation_details.xlsx')
sheet = workbook.active
for i in range(1, 3000):
    val0 = sheet.cell(row = i, column = 1).value
    val1 = sheet.cell(row = i + 1, column = 1).value
    val2 = sheet.cell(row = i + 2, column = 1).value
    if val0 is None and val1 is None and val2 is None:
        continue
    Nrow = i

# Setup output file and spreadsheet
newbook1 = openpyxl.Workbook()
newsheet1 = newbook1.active

print("\n\n\n")
print("=======================================")
print("   Step 3: Merging final spreadsheet   ")
print("=======================================")
print("\n")
print("\nCreating final spreadsheet...\n")

for i in range(8, Nrow + 1):
    val0 = sheet.cell(row = i - 1, column = 1).value
    val1 = sheet.cell(row = i, column = 1).value
    if val1 is not None:
        if val0 is None:
            valName = val1
            valENum = sheet.cell(row = i, column = 8).value
            valNote = sheet.cell(row = i, column = 10).value
        if "---Total" in val1:
            cnt = 0
            for j in range(0, 12):
                pos  = sheet.cell(row = i + 1 + j, column = 1).value
                reg  = sheet.cell(row = i + 1 + j, column = 3).value
                ot   = sheet.cell(row = i + 1 + j, column = 4).value
                rate = sheet.cell(row = i + 1 + j, column = 2).value
                if pos is None: break
                else:
                    c0 = newsheet1.cell(row = i + 1 + cnt, column = 2)
                    c0.value = valName
                    c0 = newsheet1.cell(row = i + 1 + cnt, column = 3)
                    c0.value = pos
                    if valENum is not None:
                        c4 = newsheet1.cell(row = i + 1 + cnt, column = 1)
                        c4.value = valENum
                    if valNote is not None and "Check OT" in valNote:
                        c5 = newsheet1.cell(row = i + 1 + cnt, column = 11)
                        c5.value = "Check OT+12 by hand"
                        # remove redundant flag for OT+12 calculation
                        for k in range(0, 12):
                            c6 = newsheet1.cell(row = i + cnt - k, column = 11)
                            c7 = newsheet1.cell(row = i + cnt - k, column = 2)
                            if c6.value is not None and c7.value is valName:
                                c5.value = " "
                    cat = "Unarmed"
                    if pos.__contains__("Training"): cat = "Training"
                    elif pos.__contains__("Admin Work"): cat = "Admin"
                    elif pos.__contains__("ARMED") or pos.__contains__("Armed"): cat = "Armed"
                    elif pos.__contains__("Sick"): cat = "Sick"
                    if reg != 0 and ot == 0:
                        c0 = newsheet1.cell(row = i + 1 + cnt, column = 6)
                        c0.value = reg
                        c1 = newsheet1.cell(row = i + 1 + cnt, column = 4)
                        c1.value = cat
                        c2 = newsheet1.cell(row = i + 1 + cnt, column = 7)
                        c2.value = rate
                        c3 = newsheet1.cell(row = i + 1 + cnt, column = 8)
                        c3.value = c0.value*c2.value
                        cnt += 1
                    if reg == 0 and ot != 0:
                        c0 = newsheet1.cell(row = i + 1 + cnt, column = 6)
                        c0.value = ot
                        if pos.__contains__("Training"): cat = "Training"
                        if cat.__contains__("Unarmed"): cat = "OT Unarmed"
                        if pos.__contains__("Admin"): cat = "OT Admin"
                        if pos.__contains__("ARMED") or pos.__contains__("Armed"): cat = "OT Armed"
                        c1 = newsheet1.cell(row = i + 1 + cnt, column = 4)
                        c1.value = cat
                        c2 = newsheet1.cell(row = i + 1 + cnt, column = 7)
                        c2.value = rate * 1.5
                        c3 = newsheet1.cell(row = i + 1 + cnt, column = 8)
                        c3.value = c0.value * c2.value
                        cnt += 1
                    if reg != 0 and ot != 0:
                        c0 = newsheet1.cell(row = i + 1 + cnt, column = 6)
                        c0.value = reg
                        c1 = newsheet1.cell(row = i + 1 + cnt, column = 4)
                        c1.value = cat
                        c2 = newsheet1.cell(row = i + 1 + cnt, column = 7)
                        c2.value = rate
                        c3 = newsheet1.cell(row = i + 1 + cnt, column = 8)
                        c3.value = c0.value*c2.value
                        c0 = newsheet1.cell(row = i + 2 + cnt, column = 2)
                        c0.value = valName
                        c0 = newsheet1.cell(row = i + 2 + cnt, column = 3)
                        c0.value = pos
                        c0 = newsheet1.cell(row = i + 2 + cnt, column = 6)
                        c0.value = ot
                        if valENum is not None:
                            c4 = newsheet1.cell(row = i + 2 + cnt, column = 1)
                            c4.value = valENum
                        if pos.__contains__("Training"): cat = "Training"
                        if cat.__contains__("Unarmed"): cat = "OT Unarmed"
                        if pos.__contains__("Admin"): cat = "OT Admin"
                        if pos.__contains__("ARMED") or pos.__contains__("Armed"): cat = "OT Armed"
                        c1 = newsheet1.cell(row = i + 2 + cnt, column = 4)
                        c1.value = cat
                        c2 = newsheet1.cell(row = i + 2 + cnt, column = 7)
                        c2.value = rate * 1.5
                        c3 = newsheet1.cell(row = i + 2 + cnt, column = 8)
                        c3.value = c0.value*c2.value
                        cnt += 2


c0 = newsheet1.cell(row = 1, column = 1)
c0.value = "Employee Number"
c0 = newsheet1.cell(row = 1, column = 2)
c0.value = "Employee Name"
c0 = newsheet1.cell(row = 1, column = 3)
c0.value = "Position"
c0 = newsheet1.cell(row = 1, column = 4)
c0.value = "Category"
c0 = newsheet1.cell(row = 1, column = 5)
c0.value = "Number of Shifts"
c0 = newsheet1.cell(row = 1, column = 6)
c0.value = "Number of Hours"
c0 = newsheet1.cell(row = 1, column = 7)
c0.value = "Pay Rate"
c0 = newsheet1.cell(row = 1, column = 8)
c0.value = "Total Pay"
c0 = newsheet1.cell(row = 1, column = 9)
c0.value = "Reimbursement/Deductions"
c0 = newsheet1.cell(row = 1, column = 10)
c0.value = "Grand Total"
c0 = newsheet1.cell(row = 1, column = 11)
c0.value = "Notes"

newsheet1.column_dimensions["A"].width = 20
newsheet1.column_dimensions["B"].width = 25
newsheet1.column_dimensions["C"].width = 48
newsheet1.column_dimensions["D"].width = 12
newsheet1.column_dimensions["E"].width = 12
newsheet1.column_dimensions["F"].width = 14
newsheet1.column_dimensions["G"].width = 12
newsheet1.column_dimensions["H"].width = 12
newsheet1.column_dimensions["I"].width = 12
newsheet1.column_dimensions["J"].width = 12
newsheet1.column_dimensions["K"].width = 12

# Delete empty rows
indx = []
for i in range(len(tuple(newsheet1.rows))):
    flag = False
    for cell in tuple(newsheet1.rows)[i]:
        if cell.value != None:
            flag = True
            break
    if flag == False:
        indx.append(i)
indx.sort()
for i in range(len(indx)):
    newsheet1.delete_rows(idx = indx[i] + 1 - i)

# Clean-up extra notes
for i in range(8, Nrow + 1):
    c0 = newsheet1.cell(row = i, column = 1)
    if c0 is None:
        newsheet.delete_rows(i)

output_name = "output/payroll_with_OT.xlsx"
newbook1.save(output_name)
print("\nSpreadsheet informarion merged.\n")
print("File output written to", output_name, "\n")
# Sum column totals for comparison
totHrs = 0
totPay = 0
for i in range(2, Nrow + 1):
    valHrs = newsheet1.cell(row = i, column = 6).value
    valPay = newsheet1.cell(row = i, column = 8).value
    if valHrs is None:
        continue
    totHrs = totHrs + valHrs
    totPay = totPay + valPay
print("\n\nFinal totals: ")
print("  Hours worked:  ", totHrs)
print("  Gross pay:  ", '${:,.2f}'.format(totPay), "\n\n")
print("--Step 3 complete. OT calculation finished.--\n\n\n")
