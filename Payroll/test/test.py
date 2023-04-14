# Code for calculating OT values
# Download .csv file "Assigned Shift Details - General" with the following parameters:
# "Position Name", "Date", "Start Time", "End Time", "Duration", "Employee Name", "Employee Pay Rate".
# Save .csv file as "data.xlms" and put in same directory as code.

# Source files
import openpyxl
import array as arr
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from datetime import datetime, date, timedelta

flagDebug = False

# Input file location
workbook = load_workbook('data.xlsx')

# Output file
newbook1 = openpyxl.Workbook()
newsheet1 = newbook1.active

sheet = workbook.active
label_1 = sheet.cell(row=1, column=1)
label_2 = sheet.cell(row=1, column=2)
label_3 = sheet.cell(row=1, column=3)
label_4 = sheet.cell(row=1, column=4)
label_5 = sheet.cell(row=1, column=5)
label_6 = sheet.cell(row=1, column=6)
label_7 = sheet.cell(row=1, column=7)
Nrow = sheet.max_row
daterange = sheet['B']
date_list = [daterange[x].value for x in range(2,len(daterange))]
import datetime
week1start = min(date_list)
week1end = week1start + datetime.timedelta(days = 6)
week2start = week1end + datetime.timedelta(days = 1)
week2end = max(date_list)

# Calculation for crosschecks
GrandHrs = 0
for i in range(2, Nrow-1):
    valHrs = sheet.cell(row = i, column = 5)
    GrandHrs += valHrs.value
# if flagDebug:
print("\n")
print("--- Week 1:", week1start.strftime("%Y-%m-%d"), "to", week1end.strftime("%Y-%m-%d"),
      "--- Week 2:", week2start.strftime("%Y-%m-%d"), "to", week2end.strftime("%Y-%m-%d"),"---")
print(" Overall stats")
print("Total hours for all names and positions:  ", f'{GrandHrs:.9}')
print("\n")

rowCnt = 1
warnShift = "\nEMPLOYEE HAS WORKED TOO MANY SHIFTS IN ONE WEEK!!!  GIVE THEM SOME TIME OFF!!!\n"
# Main event loop
for i in range(2, Nrow+1):
# for i in range(2, 24):
    valName = sheet.cell(row = i, column = 6)
    valName = valName.value
    valNameNxt = sheet.cell(row = i+1, column = 6)
    valNameNxt = valNameNxt.value
    if valNameNxt is valName:
        rowCnt += 1
    else:
        rowmin = i - rowCnt + 1
        rowmax = i
        emptyList = [None, None, 0, 0, 0, 0]
        listWeek1 = [emptyList, emptyList, emptyList, emptyList, emptyList,
                     emptyList, emptyList, emptyList, emptyList]
        listWeek2 = [emptyList, emptyList, emptyList, emptyList, emptyList,
                     emptyList, emptyList, emptyList, emptyList]

        # Begin OT calculation
        # Find middle cell between week 1 and week 2
        rowmid = 0
        rowmidcnt = 0
        week1Hrs = 0
        week2Hrs = 0
        for j in range(rowmin, rowmax + 1):
            valDate = sheet.cell(row = j, column = 2)
            valHrs = sheet.cell(row = j, column = 5)
            valHrs = valHrs.value
            valDate = valDate.value
            if valDate <= week1end:
                week1Hrs += valHrs
                rowmidcnt += 1
            else:
                week2Hrs += valHrs
            rowmid = rowmin + rowmidcnt

        # OT +12 for week one and week 2 without 40+
        if week1Hrs <= 40:
            regHrs = 0
            valName = sheet.cell(row = rowmid-1, column = 6)
            valName = valName.value
            list1 = [None, None, 0, 0, 0, 0]
            list2 = [None, None, 0, 0, 0, 0]
            list3 = [None, None, 0, 0, 0, 0]
            list4 = [None, None, 0, 0, 0, 0]
            list5 = [None, None, 0, 0, 0, 0]
            list6 = [None, None, 0, 0, 0, 0]
            list7 = [None, None, 0, 0, 0, 0]
            list8 = [None, None, 0, 0, 0, 0]
            list9 = [None, None, 0, 0, 0, 0]
            for i in range(rowmid-1, rowmin-1, -1): # minus one offset in loop because we're counting backwards
                valPos = sheet.cell(row = i, column = 1)
                valHrs = sheet.cell(row = i, column = 5)
                valPos = valPos.value
                valHrs = valHrs.value
                if valHrs > 12:
                    z = valHrs - 12
                    regHrs = valHrs - z
                else:
                    z = 0
                    regHrs = valHrs
                if i == rowmid - 1:   list1 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmid - 2: list2 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmid - 3: list3 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmid - 4: list4 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmid - 5: list5 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmid - 6: list6 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmid - 7: list7 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmid - 8: list8 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmid - 9: list9 = [valName, valPos, valHrs, regHrs, z, 0]
                if i <= rowmid - 10: print(warnShift)
                if flagDebug: print(valName, "  ", valPos, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", z)
            listWeek1 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            if flagDebug: print(valName, "  Week 1 --- total: ", week1Hrs, " shift+40 total: ", 0, "\n")
        if week2Hrs <= 40:
            regHrs = 0
            valName = sheet.cell(row = rowmax, column = 6)
            valName = valName.value
            list1 = [None, None, 0, 0, 0, 0]
            list2 = [None, None, 0, 0, 0, 0]
            list3 = [None, None, 0, 0, 0, 0]
            list4 = [None, None, 0, 0, 0, 0]
            list5 = [None, None, 0, 0, 0, 0]
            list6 = [None, None, 0, 0, 0, 0]
            list7 = [None, None, 0, 0, 0, 0]
            list8 = [None, None, 0, 0, 0, 0]
            list9 = [None, None, 0, 0, 0, 0]
            for i in range(rowmax, rowmid-1, -1):
                valPos = sheet.cell(row = i, column = 1)
                valHrs = sheet.cell(row = i, column = 5)
                valPos = valPos.value
                valHrs = valHrs.value
                if valHrs > 12:
                    z = valHrs - 12
                    regHrs = valHrs - z
                else:
                    z = 0
                    regHrs = valHrs
                if i == rowmax:   list1 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmax - 1: list2 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmax - 2: list3 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmax - 3: list4 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmax - 4: list5 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmax - 5: list6 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmax - 6: list7 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmax - 7: list8 = [valName, valPos, valHrs, regHrs, z, 0]
                elif i == rowmax - 8: list9 = [valName, valPos, valHrs, regHrs, z, 0]
                if i <= rowmid - 9: print(warnShift)
                if flagDebug: print(valName, "  ", valPos, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", z)
            listWeek2 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            if flagDebug: print(valName, "  Week 2 --- total: ", week2Hrs, " shift+40 total: ", 0, "\n")

        # OT +40 for week one
        if week1Hrs > 40:
            regHrs = 0
            OT40week1 = 0
            OT12 = 0
            OT40week1 = week1Hrs - 40
            OT = OT40week1
            OTn = OT
            OTcnt = 0
            flag1 = False
            valName = sheet.cell(row = rowmid-1, column = 6)
            valName = valName.value
            list1 = [None, None, 0, 0, 0, 0]
            list2 = [None, None, 0, 0, 0, 0]
            list3 = [None, None, 0, 0, 0, 0]
            list4 = [None, None, 0, 0, 0, 0]
            list5 = [None, None, 0, 0, 0, 0]
            list6 = [None, None, 0, 0, 0, 0]
            list7 = [None, None, 0, 0, 0, 0]
            list8 = [None, None, 0, 0, 0, 0]
            list9 = [None, None, 0, 0, 0, 0]
            for i in range(rowmid-1, rowmin-1, -1): # minus one offset in for loop because we're counting backwards
                valPos = sheet.cell(row = i, column = 1)
                valHrs = sheet.cell(row = i, column = 5)
                valPos = valPos.value
                valHrs = valHrs.value
                x = valHrs
                if flag1 is True:
                    OTn = 0
                    y = 0
                if flag1 is False:
                    OTn = OTn - x
                    if OTn > 0:
                        y = x
                        OTcnt += y
                    if OTn < 0:
                        y = OT - OTcnt
                        flag1 = True
                if valHrs > 12: # OT +12 calculation
                    z = valHrs - 12
                    if y > z: z = 0
                    else: z = z - y
                else: z = 0
                regHrs = valHrs - y - z
                if y + z == 0: regHrs = valHrs
                if i == rowmid - 1:   list1 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmid - 2: list2 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmid - 3: list3 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmid - 4: list4 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmid - 5: list5 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmid - 6: list6 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmid - 7: list7 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmid - 8: list8 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmid - 9: list9 = [valName, valPos, valHrs, regHrs, z, y]
                if i <= rowmid - 10: print(warnShift)
                if flagDebug: print(valName, "  ", valPos, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", OT12, "  OT+40: ", y)

            listWeek1 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            if flagDebug: print(valName,"  Week 1 --- total: ", week1Hrs," shift+40 total: ",OT40week1,"\n")

        # OT +40 for week two
        if week2Hrs > 40:
            regHrs = 0
            OT40week2 = 0
            OT40week2 = week2Hrs - 40
            OT = OT40week2
            OTn = OT
            OTcnt = 0
            flag1 = False
            valName = sheet.cell(row = rowmax, column = 6)
            valName = valName.value
            list1 = [None, None, 0, 0, 0, 0]
            list2 = [None, None, 0, 0, 0, 0]
            list3 = [None, None, 0, 0, 0, 0]
            list4 = [None, None, 0, 0, 0, 0]
            list5 = [None, None, 0, 0, 0, 0]
            list6 = [None, None, 0, 0, 0, 0]
            list7 = [None, None, 0, 0, 0, 0]
            list8 = [None, None, 0, 0, 0, 0]
            list9 = [None, None, 0, 0, 0, 0]
            for i in range(rowmax, rowmid-1, -1):
                valPos = sheet.cell(row = i, column = 1)
                valHrs = sheet.cell(row = i, column = 5)
                valPos = valPos.value
                valHrs = valHrs.value
                x = valHrs
                if flag1 is True:
                    OTn = 0
                    y = 0
                if flag1 is False:
                    OTn = OTn - x
                    if OTn > 0:
                        y = x
                        OTcnt += y
                    if OTn < 0:
                        y = OT - OTcnt
                        flag1 = True
                if valHrs > 12: # OT +12 calculation
                    z = valHrs - 12
                    if y > z: z = 0
                    else: z = z - y
                else: z = 0
                regHrs = valHrs - y - z
                if y + z == 0: regHrs = valHrs
                if i == rowmax:   list1 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmax - 1: list2 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmax - 2: list3 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmax - 3: list4 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmax - 4: list5 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmax - 5: list6 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmax - 6: list7 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmax - 7: list8 = [valName, valPos, valHrs, regHrs, z, y]
                elif i == rowmax - 8: list9 = [valName, valPos, valHrs, regHrs, z, y]
                if i <= rowmax - 9: print(warnShift)
                if flagDebug: print(valName, "  ", valPos, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", z, "  OT+40: ", y)

            listWeek2 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            if flagDebug: print(valName,"  Week 2 --- total: ", week2Hrs," shift+40 total: ",OT40week2,"\n")

        # Sort by position
        for j in range(0,2):
            if j == 0:
                list1 = (listWeek1[0])
                list2 = (listWeek1[1])
                list3 = (listWeek1[2])
                list4 = (listWeek1[3])
                list5 = (listWeek1[4])
                list6 = (listWeek1[5])
                list7 = (listWeek1[6])
                list8 = (listWeek1[7])
                list9 = (listWeek1[8])
            else:
                list1 = (listWeek2[0])
                list2 = (listWeek2[1])
                list3 = (listWeek2[2])
                list4 = (listWeek2[3])
                list5 = (listWeek2[4])
                list6 = (listWeek2[5])
                list7 = (listWeek2[6])
                list8 = (listWeek2[7])
                list9 = (listWeek2[8])
            pos1 = list1[1]
            pos2 = list2[1]
            pos3 = list3[1]
            pos4 = list4[1]
            pos5 = list5[1]
            pos6 = list6[1]
            pos7 = list7[1]
            pos8 = list8[1]
            pos9 = list9[1]
            hrsPos1 = list1[2]
            RegHrs1 = list1[3]
            totOT12_1 = list1[4]
            totOT40_1 = list1[5]
            # Position 1
            if pos2 is not None and pos2 in pos1:
                hrsPos1 += list2[2]
                RegHrs1 += list2[3]
                totOT12_1 += list2[4]
                totOT40_1 += list2[5]
            if pos3 is not None and pos3 in pos1:
                hrsPos1 += list3[2]
                RegHrs1 += list3[3]
                totOT12_1 += list3[4]
                totOT40_1 += list3[5]
            if pos4 is not None and pos4 in pos1:
                hrsPos1 += list4[2]
                RegHrs1 += list4[3]
                totOT12_1 += list4[4]
                totOT40_1 += list4[5]
            if pos5 is not None and pos5 in pos1:
                hrsPos1 += list5[2]
                RegHrs1 += list5[3]
                totOT12_1 += list5[4]
                totOT40_1 += list5[5]
            if pos6 is not None and pos6 in pos1:
                hrsPos1 += list6[2]
                RegHrs1 += list6[3]
                totOT12_1 += list6[4]
                totOT40_1 += list6[5]
            if pos7 is not None and pos7 in pos1:
                hrsPos1 += list7[2]
                RegHrs1 += list7[3]
                totOT12_1 += list7[4]
                totOT40_1 += list7[5]
            if pos8 is not None and pos8 in pos1:
                hrsPos1 += list8[2]
                RegHrs1 += list8[3]
                totOT12_1 += list8[4]
                totOT40_1 += list8[5]
            if pos9 is not None and pos9 in pos1:
                hrsPos1 += list9[2]
                RegHrs1 += list9[3]
                totOT12_1 += list9[4]
                totOT40_1 += list9[5]
            list1 = [valName, pos1, hrsPos1, RegHrs1, totOT12_1, totOT40_1]
            # Position 2
            hrsPos2 = 0
            RegHrs2 = 0
            totOT12_2 = 0
            totOT40_2 = 0
            if (    pos2 is not None and pos2 not in pos1):
                hrsPos2 = list2[2]
                RegHrs2 = list2[3]
                totOT12_2 = list2[4]
                totOT40_2 = list2[5]
                if pos3 is not None and pos3 in pos2:
                    hrsPos2 += list3[2]
                    RegHrs2 += list3[3]
                    totOT12_2 += list3[4]
                    totOT40_2 += list3[5]
                if pos4 is not None and pos4 in pos2:
                    hrsPos2 += list4[2]
                    RegHrs2 += list4[3]
                    totOT12_2 += list4[4]
                    totOT40_2 += list4[5]
                if pos5 is not None and pos5 in pos2:
                    hrsPos2 += list5[2]
                    RegHrs2 += list5[3]
                    totOT12_2 += list5[4]
                    totOT40_2 += list5[5]
                if pos6 is not None and pos6 in pos2:
                    hrsPos2 += list6[2]
                    RegHrs2 += list6[3]
                    totOT12_2 += list6[4]
                    totOT40_2 += list6[5]
                if pos7 is not None and pos7 in pos2:
                    hrsPos2 += list7[2]
                    RegHrs2 += list7[3]
                    totOT12_2 += list7[4]
                    totOT40_2 += list7[5]
                if pos8 is not None and pos8 in pos2:
                    hrsPos2 += list8[2]
                    RegHrs2 += list8[3]
                    totOT12_2 += list8[4]
                    totOT40_2 += list8[5]
                if pos9 is not None and pos9 in pos2:
                    hrsPos2 += list9[2]
                    RegHrs2 += list9[3]
                    totOT12_2 += list9[4]
                    totOT40_2 += list9[5]
                list2 = [valName, pos2, hrsPos2, RegHrs2, totOT12_2, totOT40_2]
            # Position 3
            hrsPos3 = 0
            RegHrs3 = 0
            totOT12_3 = 0
            totOT40_3 = 0
            if (    pos3 is not None and pos3 not in pos1 and pos3 not in pos2):
                hrsPos3 += list3[2]
                RegHrs3 += list3[3]
                totOT12_3 += list3[4]
                totOT40_3 += list3[5]
                if pos4 is not None and pos4 in pos3:
                    hrsPos3 += list4[2]
                    RegHrs3 += list4[3]
                    totOT12_3 += list4[4]
                    totOT40_3 += list4[5]
                if pos5 is not None and pos5 in pos3:
                    hrsPos3 += list5[2]
                    RegHrs3 += list5[3]
                    totOT12_3 += list5[4]
                    totOT40_3 += list5[5]
                if pos6 is not None and pos6 in pos3:
                    hrsPos3 += list6[2]
                    RegHrs3 += list6[3]
                    totOT12_3 += list6[4]
                    totOT40_3 += list6[5]
                if pos7 is not None and pos7 in pos3:
                    hrsPos3 += list7[2]
                    RegHrs3 += list7[3]
                    totOT12_3 += list7[4]
                    totOT40_3 += list7[5]
                if pos8 is not None and pos8 in pos3:
                    hrsPos3 += list8[2]
                    RegHrs3 += list8[3]
                    totOT12_3 += list8[4]
                    totOT40_3 += list8[5]
                if pos9 is not None and pos9 in pos3:
                    hrsPos3 += list9[2]
                    RegHrs3 += list9[3]
                    totOT12_3 += list9[4]
                    totOT40_3 += list9[5]
                list3 = [valName, pos3, hrsPos3, RegHrs3, totOT12_3, totOT40_3]
            # Position 4
            hrsPos4 = 0
            RegHrs4 = 0
            totOT12_4 = 0
            totOT40_4 = 0
            if (    pos4 is not None and pos4 not in pos1 and pos4 not in pos2 and pos4 not in pos3):
                hrsPos4 += list4[2]
                RegHrs4 += list4[3]
                totOT12_4 += list4[4]
                totOT40_4 += list4[5]
                if pos5 is not None and pos5 in pos4:
                    hrsPos4 += list5[2]
                    RegHrs4 += list5[3]
                    totOT12_4 += list5[4]
                    totOT40_4 += list5[5]
                if pos6 is not None and pos6 in pos4:
                    hrsPos4 += list6[2]
                    RegHrs4 += list6[3]
                    totOT12_4 += list6[4]
                    totOT40_4 += list6[5]
                if pos7 is not None and pos7 in pos4:
                    hrsPos4 += list7[2]
                    RegHrs4 += list7[3]
                    totOT12_4 += list7[4]
                    totOT40_4 += list7[5]
                if pos8 is not None and pos8 in pos4:
                    hrsPos4 += list8[2]
                    RegHrs4 += list8[3]
                    totOT12_4 += list8[4]
                    totOT40_4 += list8[5]
                if pos9 is not None and pos9 in pos4:
                    hrsPos4 += list9[2]
                    RegHrs4 += list9[3]
                    totOT12_4 += list9[4]
                    totOT40_4 += list9[5]
                list4 = [valName, pos4, hrsPos4, RegHrs4, totOT12_4, totOT40_4]
            # Position 5
            hrsPos5 = 0
            RegHrs5 = 0
            totOT12_5 = 0
            totOT40_5 = 0
            if (    pos5 is not None and pos5 not in pos1 and pos5 not in pos2 and pos5 not in pos3 and
                    pos5 not in pos4):
                hrsPos5 += list5[2]
                RegHrs5 += list5[3]
                totOT12_5 += list5[4]
                totOT40_5 += list5[5]
                if pos6 is not None and pos6 in pos5:
                    hrsPos5 += list6[2]
                    RegHrs5 += list6[3]
                    totOT12_5 += list6[4]
                    totOT40_5 += list6[5]
                if pos7 is not None and pos7 in pos5:
                    hrsPos5 += list7[2]
                    RegHrs5 += list7[3]
                    totOT12_5 += list7[4]
                    totOT40_5 += list7[5]
                if pos8 is not None and pos8 in pos5:
                    hrsPos5 += list8[2]
                    RegHrs5 += list8[3]
                    totOT12_5 += list8[4]
                    totOT40_5 += list8[5]
                if pos9 is not None and pos9 in pos5:
                    hrsPos5 += list9[2]
                    RegHrs5 += list9[3]
                    totOT12_5 += list9[4]
                    totOT40_5 += list9[5]
                list5 = [valName, pos5, hrsPos5, RegHrs5, totOT12_5, totOT40_5]
            # Position 6
            hrsPos6 = 0
            RegHrs6 = 0
            totOT12_6 = 0
            totOT40_6 = 0
            if (    pos6 is not None and pos6 not in pos1 and pos6 not in pos2 and pos6 not in pos3 and
                    pos6 not in pos4 and pos6 not in pos5):
                hrsPos6 += list6[2]
                RegHrs6 += list6[3]
                totOT12_6 += list6[4]
                totOT40_6 += list6[5]
                if pos7 is not None and pos7 in pos6:
                    hrsPos6 += list7[2]
                    RegHrs6 += list7[3]
                    totOT12_6 += list7[4]
                    totOT40_6 += list7[5]
                if pos8 is not None and pos8 in pos6:
                    hrsPos6 += list8[2]
                    RegHrs6 += list8[3]
                    totOT12_6 += list8[4]
                    totOT40_6 += list8[5]
                if pos9 is not None and pos9 in pos6:
                    hrsPos6 += list9[2]
                    RegHrs6 += list9[3]
                    totOT12_6 += list9[4]
                    totOT40_6 += list9[5]
                list6 = [valName, pos6, hrsPos6, RegHrs6, totOT12_6, totOT40_6]
            # Position 7
            hrsPos7 = 0
            RegHrs7 = 0
            totOT12_7 = 0
            totOT40_7 = 0
            if (    pos7 is not None and pos7 not in pos1 and pos7 not in pos2 and pos7 not in pos3 and
                    pos7 not in pos4 and pos7 not in pos5 and pos7 not in pos6):
                hrsPos7 += list7[2]
                RegHrs7 += list7[3]
                totOT12_7 += list7[4]
                totOT40_7 += list7[5]
                if pos8 is not None and pos8 in pos7:
                    hrsPos7 += list8[2]
                    RegHrs7 += list8[3]
                    totOT12_7 += list8[4]
                    totOT40_7 += list8[5]
                if pos9 is not None and pos9 in pos7:
                    hrsPos7 += list9[2]
                    RegHrs7 += list9[3]
                    totOT12_7 += list9[4]
                    totOT40_7 += list9[5]
                list7 = [valName, pos7, hrsPos7, RegHrs7, totOT12_7, totOT40_7]
            # Position 8
            hrsPos8 = 0
            RegHrs8 = 0
            totOT12_8 = 0
            totOT40_8 = 0
            if (    pos8 is not None and pos8 not in pos1 and pos8 not in pos2 and pos8 not in pos3 and
                    pos8 not in pos4 and pos8 not in pos5 and pos8 not in pos6 and pos8 not in pos7):
                hrsPos8 += list8[2]
                RegHrs8 += list8[3]
                totOT12_8 += list8[4]
                totOT40_8 += list8[5]
                if pos9 is not None and pos9 in pos8:
                    hrsPos8 += list9[2]
                    RegHrs8 += list9[3]
                    totOT12_8 += list9[4]
                    totOT40_8 += list9[5]
                list8 = [valName, pos8, hrsPos8, RegHrs8, totOT12_8, totOT40_8]
            # Position 9
            hrsPos9 = 0
            RegHrs9 = 0
            totOT12_9 = 0
            totOT40_9 = 0
            if (    pos9 is not None and pos9 not in pos1 and pos9 not in pos2 and pos9 not in pos3 and
                    pos9 not in pos4 and pos9 not in pos5 and pos9 not in pos6 and pos9 not in pos7 and
                    pos9 not in pos8):
                hrsPos9 += list9[2]
                RegHrs9 += list9[3]
                totOT12_9 += list9[4]
                totOT40_9 += list9[5]
                list9 = [valName, pos9, hrsPos9, RegHrs9, totOT12_9, totOT40_9]

            if (    list2[1] is not None and (list2[1] is list1[1])):
                list2 = [None, None, 0, 0, 0, 0]
            if (    list3[1] is not None and (list3[1] is list1[1] or list3[1] is list2[1])):
                list3 = [None, None, 0, 0, 0, 0]
            if (    list4[1] is not None and (list4[1] is list1[1] or list4[1] is list2[1] or list4[1] is list3[1])):
                list4 = [None, None, 0, 0, 0, 0]
            if (    list5[1] is not None and (list5[1] is list1[1] or list5[1] is list2[1] or list5[1] is list3[1] or
                    list5[1] is list4[1])):
                list5 = [None, None, 0, 0, 0, 0]
            if (    list6[1] is not None and (list6[1] is list1[1] or list6[1] is list2[1] or list6[1] is list3[1] or
                    list6[1] is list4[1] or list6[1] is list5[1])):
                list6 = [None, None, 0, 0, 0, 0]
            if (    list7[1] is not None and (list7[1] is list1[1] or list7[1] is list2[1] or list7[1] is list3[1] or
                    list7[1] is list4[1] or list7[1] is list5[1] or list7[1] is list6[1])):
                list7 = [None, None, 0, 0, 0, 0]
            if (    list8[1] is not None and (list8[1] is list1[1] or list8[1] is list2[1] or list8[1] is list3[1] or
                    list8[1] is list4[1] or list8[1] is list5[1] or list8[1] is list6[1] or list8[1] is list7[1])):
                list8 = [None, None, 0, 0, 0, 0]
            if (    list9[1] is not None and (list9[1] is list1[1] or list9[1] is list2[1] or list9[1] is list3[1] or
                    list9[1] is list4[1] or list9[1] is list5[1] or list9[1] is list6[1] or list9[1] is list7[1] or
                    list9[1] is list8[1])):
                list9 = [None, None, 0, 0, 0, 0]

            if list8[1] is None and list9[1] is not None:
                list8 = list9
                list9 = [None, None, 0, 0, 0, 0]
            if list7[1] is None and list8[1] is not None:
                list7 = list8
                list8 = [None, None, 0, 0, 0, 0]
            if list6[1] is None and list7[1] is not None:
                list6 = list7
                list7 = [None, None, 0, 0, 0, 0]
            if list5[1] is None and list6[1] is not None:
                list5 = list6
                list6 = [None, None, 0, 0, 0, 0]
            if list4[1] is None and list5[1] is not None:
                list4 = list5
                list5 = [None, None, 0, 0, 0, 0]
            if list3[1] is None and list4[1] is not None:
                list3 = list4
                list4 = [None, None, 0, 0, 0, 0]
            if list2[1] is None and list3[1] is not None:
                list2 = list3
                list3 = [None, None, 0, 0, 0, 0]
            if j == 0: listWeek1 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            else:      listWeek2 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]

        print("\n", valName)
        print("--Week 1 totals--")
        print("    Position:\t\t  total hrs:   reg hrs:   OT12:   OT40:")
        if listWeek1[0][0] is not None: print(listWeek1[0])
        if listWeek1[1][0] is not None: print(listWeek1[1])
        if listWeek1[2][0] is not None: print(listWeek1[2])
        if listWeek1[3][0] is not None: print(listWeek1[3])
        if listWeek1[4][0] is not None: print(listWeek1[4])
        if listWeek1[5][0] is not None: print(listWeek1[5])
        if listWeek1[6][0] is not None: print(listWeek1[6])
        if listWeek1[7][0] is not None: print(listWeek1[7])
        if listWeek1[8][0] is not None: print(listWeek1[8])
        print("--Week 2 totals--")
        print("    Position:\t\t  total hrs:   reg hrs:   OT12:   OT40:")
        if listWeek2[0][0] is not None: print(listWeek2[0])
        if listWeek2[1][0] is not None: print(listWeek2[1])
        if listWeek2[2][0] is not None: print(listWeek2[2])
        if listWeek2[3][0] is not None: print(listWeek2[3])
        if listWeek2[4][0] is not None: print(listWeek2[4])
        if listWeek2[5][0] is not None: print(listWeek2[5])
        if listWeek2[6][0] is not None: print(listWeek2[6])
        if listWeek2[7][0] is not None: print(listWeek2[7])
        if listWeek2[8][0] is not None: print(listWeek2[8])
        # Reset count parameter for next employee
        rowCnt = 1
