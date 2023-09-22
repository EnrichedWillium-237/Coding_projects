### Code for reading in a payroll spreadsheet and organizing gross payments ###
### This is a work in progress and does things in a brute force fashion. Will be improved later. ###

# Source files
import openpyxl
import array as arr
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
import sys

# Input payroll file
workbook = load_workbook('payroll_input.xlsx')

# Payroll output
newbook1 = openpyxl.Workbook()
newsheet1 = newbook1.active

sheet = workbook.active
label_0 = sheet.cell(row = 1, column = 1).value
label_1 = sheet.cell(row = 1, column = 2).value
label_2 = sheet.cell(row = 1, column = 3).value
label_3 = sheet.cell(row = 1, column = 4).value
label_4 = sheet.cell(row = 1, column = 5).value
label_5 = sheet.cell(row = 1, column = 6).value
label_6 = sheet.cell(row = 1, column = 7).value
label_7 = sheet.cell(row = 1, column = 8).value
if ("Employee Number" not in label_0 or "Employee Name" not in label_1 or "Category" not in label_2 or
    "Number of Hours" not in label_3 or "Pay Rate" not in label_4 or "Total Pay" not in label_5 or "Reimbursement/Deductions" not in label_6):
    print("\n\n")
    print("===================================================")
    print("---WARNING!!! INCORRECT SPREADSHEET HEADERS!!!---  ")
    print("                                                   ")
    print("            ---Fix then try again!---              ")
    print("===================================================")
    print("\n\n")
    sys.exit(0)
print(label_0, "\t", label_1, "\t", label_2, "\t", label_3, "\t", label_4, "\t", label_5, "\t", label_6)

Nrow = sheet.max_row
Nrow = Nrow - 1 # offset from spreadsheet
Ncell = 1 # index for excel output

# hours per category
hrsUnarmed = 0
hrsArmed = 0
hrsReg = 0
hrsOT = 0
hrsSick = 0
hrsTotal = 0

GrandUnarmed = 0
GrandArmed = 0
GrandReg = 0
GrandOT = 0
GrandSick = 0

Gross = 0 # total pay per employee
GrandHrs = 0 # total number of hours
GrandTot = 0 # total payroll
GrandReim = 0 # total reimbursement
GrandGross = 0 # total payout: total + reimbursement

c0 = newsheet1.cell(row = 1, column = 1)
c0.value = "Employee name"
c0 = newsheet1.cell(row = 1, column = 2)
c0.font = Font(bold = 'single')
c0.value = "ID"
c0 = newsheet1.cell(row = 1, column = 3)
c0.value = "Pay rate"
c0 = newsheet1.cell(row = 1, column = 4)
c0.value = "Category"
c0 = newsheet1.cell(row = 1, column = 21)
c0.value = "Reimbursement"

nrowsEmp = 0 # number of rows per employee
rate1  = 0
rate2  = 0
rate3  = 0
rate4  = 0
rate5  = 0
rate6  = 0
rate7  = 0
rate8  = 0
rate9  = 0
rate10 = 0
rate11 = 0
rate12 = 0
rate13 = 0
rate14 = 0
rate15 = 0
rate1hrs  = 0
rate2hrs  = 0
rate3hrs  = 0
rate4hrs  = 0
rate5hrs  = 0
rate6hrs  = 0
rate7hrs  = 0
rate8hrs  = 0
rate9hrs  = 0
rate10hrs = 0
rate11hrs = 0
rate12hrs = 0
rate13hrs = 0
rate14hrs = 0
rate15hrs = 0
rate1unarmed  = 0
rate2unarmed  = 0
rate3unarmed  = 0
rate4unarmed  = 0
rate5unarmed  = 0
rate6unarmed  = 0
rate7unarmed  = 0
rate8unarmed  = 0
rate9unarmed  = 0
rate10unarmed = 0
rate11unarmed = 0
rate12unarmed = 0
rate13unarmed = 0
rate14unarmed = 0
rate15unarmed = 0
rate1armed  = 0
rate2armed  = 0
rate3armed  = 0
rate4armed  = 0
rate5armed  = 0
rate6armed  = 0
rate7armed  = 0
rate8armed  = 0
rate9armed  = 0
rate10armed = 0
rate11armed = 0
rate12armed = 0
rate13armed = 0
rate14armed = 0
rate15armed = 0
rate1reg  = 0
rate2reg  = 0
rate3reg  = 0
rate4reg  = 0
rate5reg  = 0
rate6reg  = 0
rate7reg  = 0
rate8reg  = 0
rate9reg  = 0
rate10reg = 0
rate11reg = 0
rate12reg = 0
rate13reg = 0
rate14reg = 0
rate15reg = 0
rate1OT  = 0
rate2OT  = 0
rate3OT  = 0
rate4OT  = 0
rate5OT  = 0
rate6OT  = 0
rate7OT  = 0
rate8OT  = 0
rate9OT  = 0
rate10OT = 0
rate11OT = 0
rate12OT = 0
rate13OT = 0
rate14OT = 0
rate15OT = 0
rate1sick  = 0
rate2sick  = 0
rate3sick  = 0
rate4sick  = 0
rate5sick  = 0
rate6sick  = 0
rate7sick  = 0
rate8sick  = 0
rate9sick  = 0
rate10sick = 0
rate11sick = 0
rate12sick = 0
rate13sick = 0
rate14sick = 0
rate15sick = 0

for i in range (2, Nrow):
#for i in range(2, 25):

    valID = sheet.cell(row = i, column = 1)
    valName = sheet.cell(row = i, column = 2)
    valCat = sheet.cell(row = i, column = 3)
    valHours = sheet.cell(row = i, column = 4)
    valRate = sheet.cell(row = i, column = 5)
    valTot = sheet.cell(row = i, column = 6)
    valReim = sheet.cell(row = i, column = 7)
    valGross = sheet.cell(row = i, column = 8)
    valNote = sheet.cell(row = i, column = 9)

    valID = valID.value
    Name = valName.value
    Cat = valCat.value
    Hours = valHours.value
    if Hours is None:
        Hours = 0
    Rate = valRate.value
    Tot = valTot.value
    Reim = valReim.value
    if Reim is None:
        Reim = 0
#    Gross = Tot + Reim # calculate reimbursements
    # valGross = valGross.value
    # Note = valNote.value
    # Gross += valGross
    # Tot_nxt = Tot
    #
    # GrandHrs += Hours
    # GrandTot += Tot
    # GrandReim += Reim
    # GrandGross += valGross
    #
    # # Sort by category
    # if Cat is None:
    #     print("No category given")
    # elif Cat.__contains__("Un") and not Cat.__contains__("OT"):
    #     hrsUnarmed += Hours
    #     GrandUnarmed += Hours
    # elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
    #     hrsArmed += Hours
    #     GrandArmed += Hours
    # elif Cat.__contains__("Admin"):
    #     hrsReg += Hours
    #     GrandReg += Hours
    # elif Cat.__contains__("Train"):
    #     hrsReg += Hours
    #     GrandReg += Hours
    # elif Cat.__contains__("OT"):
    #     hrsOT += Hours
    #     GrandOT += Hours
    # elif Cat.__contains__("Sick"):
    #     hrsSick += Hours
    #     GrandSick += Hours
    # else:
    #     print("Unknown category for hours!")
    # hrsTotal += Hours
    #
    # Name_nxt = sheet.cell(row = i + 1, column = 2)
    # Tot_nxt = sheet.cell(row = i + 1, column = 6)
    # Reim_nxt = sheet.cell(row = i + 1, column = 7)
    # Note_nxt = sheet.cell(row = i +1, column = 8)
    # Name_nxt = Name_nxt.value
    # Tot_nxt = Tot_nxt.value
    # Reim_nxt = Reim_nxt.value
    # Note_nxt = Note_nxt.value
    #
    # nrowsEmp += 1
