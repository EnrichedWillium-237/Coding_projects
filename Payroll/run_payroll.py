### Code for reading in a payroll spreadsheet and organizing gross payments ###
### This is a work in progress and does things in a brute force fashion. Will be improved later. ###

# Source files
import openpyxl
import array as arr
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side

# Payroll file location
workbook = load_workbook('payroll_input.xlsx')

# Payroll output
newbook1 = openpyxl.Workbook()
newsheet1 = newbook1.active

sheet = workbook.active
label_0 = sheet.cell(row=1,column=1)
label_1 = sheet.cell(row=1,column=2)
label_2 = sheet.cell(row=1,column=3)
label_3 = sheet.cell(row=1,column=4)
label_4 = sheet.cell(row=1,column=5)
label_5 = sheet.cell(row=1,column=6)
label_6 = sheet.cell(row=1,column=7)
print(label_0.value, "\t", label_1.value, "\t", label_2.value, "\t", label_3.value, "\t", label_4.value, "\t", label_5.value, "\t", label_6.value)

Nrow = sheet.max_row
Nrow = Nrow - 1 # offset from spreadsheet
Ncell = 1 # index for excel output

# hours per category
hrsUnarmed = 0
hrsArmed = 0
hrsAdmin = 0
hrsOT = 0
hrsTrain = 0
hrsSick = 0
hrsCOVID = 0
hrsTotal = 0

GrandUnarmed = 0
GrandArmed = 0
GrandAdmin = 0
GrandOT = 0
GrandTrain = 0
GrandSick = 0
GrandCovid = 0

Gross = 0 # total pay per employee
GrandHrs = 0 # total number of hours
GrandTot = 0 # total payroll
GrandReim = 0 # total reimbursement
GrandGross = 0 # total payout: total + reimbursement

c0 = newsheet1.cell(row = 1, column = 1)
c0.value = "Employee name"
c0 = newsheet1.cell(row = 1, column = 2)
c0.value = "Pay rate"
c0 = newsheet1.cell(row = 1, column = 21)
c0.value = "Reimbursement"

nrowsEmp = 0 # number of rows per employee
rate1 = 0
rate2 = 0
rate3 = 0
rate4 = 0
rate5 = 0
rate6 = 0
rate7 = 0
rate8 = 0
rate9 = 0
rate10 = 0
rate11 = 0
rate12 = 0
rate1hrs = 0
rate2hrs = 0
rate3hrs = 0
rate4hrs = 0
rate5hrs = 0
rate6hrs = 0
rate7hrs = 0
rate8hrs = 0
rate9hrs = 0
rate10hrs = 0
rate11hrs = 0
rate12hrs = 0
rate1unarmed = 0
rate2unarmed = 0
rate3unarmed = 0
rate4unarmed = 0
rate5unarmed = 0
rate6unarmed = 0
rate7unarmed = 0
rate8unarmed = 0
rate9unarmed = 0
rate10unarmed = 0
rate11unarmed = 0
rate12unarmed = 0
rate1armed = 0
rate2armed = 0
rate3armed = 0
rate4armed = 0
rate5armed = 0
rate6armed = 0
rate7armed = 0
rate8armed = 0
rate9armed = 0
rate10armed = 0
rate11armed = 0
rate12armed = 0
rate1admin = 0
rate2admin = 0
rate3admin = 0
rate4admin = 0
rate5admin = 0
rate6admin = 0
rate7admin = 0
rate8admin = 0
rate9admin = 0
rate10admin = 0
rate11admin = 0
rate12admin = 0
rate1OT = 0
rate2OT = 0
rate3OT = 0
rate4OT = 0
rate5OT = 0
rate6OT = 0
rate7OT = 0
rate8OT = 0
rate9OT = 0
rate10OT = 0
rate11OT = 0
rate12OT = 0
rate1train = 0
rate2train = 0
rate3train = 0
rate4train = 0
rate5train = 0
rate6train = 0
rate7train = 0
rate8train = 0
rate9train = 0
rate10train = 0
rate11train = 0
rate12train = 0
rate1sick = 0
rate2sick = 0
rate3sick = 0
rate4sick = 0
rate5sick = 0
rate6sick = 0
rate7sick = 0
rate8sick = 0
rate9sick = 0
rate10sick = 0
rate11sick = 0
rate12sick = 0
rate1covid = 0
rate2covid = 0
rate3covid = 0
rate4covid = 0
rate5covid = 0
rate6covid = 0
rate7covid = 0
rate8covid = 0
rate9covid = 0
rate10covid = 0
rate11covid = 0
rate12covid = 0

for i in range (2, Nrow):
#for i in range(2, 25):

    valName = sheet.cell(row = i, column = 1)
    valCat = sheet.cell(row = i, column = 2)
    valHours = sheet.cell(row = i, column = 3)
    valRate = sheet.cell(row = i, column = 4)
    valTot = sheet.cell(row = i, column = 5)
    valReim = sheet.cell(row = i, column = 6)
    valGross = sheet.cell(row = i, column = 7)
    valNote = sheet.cell(row = i, column = 8)

    Name = valName.value
    Cat = valCat.value
    Hours = valHours.value
    Rate = valRate.value
    Tot = valTot.value
    Reim = valReim.value
    if Reim is None:
        Reim = 0
#    Gross = Tot + Reim # calculate reimbursements
    valGross = valGross.value
    Note = valNote.value
    Gross += valGross
    Tot_nxt = Tot

    GrandHrs += Hours
    GrandTot += Tot
    GrandReim += Reim
    GrandGross += valGross

    # Sort by category
    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
        hrsUnarmed += Hours
        GrandUnarmed += Hours
    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
        hrsArmed += Hours
        GrandArmed += Hours
    elif Cat.__contains__("Admin"):
        hrsAdmin += Hours
        GrandAdmin += Hours
    elif Cat.__contains__("OT"):
        hrsOT += Hours
        GrandOT += Hours
    elif Cat.__contains__("Train"):
        hrsTrain += Hours
        GrandTrain += Hours
    elif Cat.__contains__("Sick"):
        hrsSick += Hours
        GrandSick += Hours
    elif Cat.__contains__("COVID"):
        hrsCOVID += Hours
        GrandCovid += Hours
    else:
        print("Unknown category for hours!")
    hrsTotal += Hours

    Name_nxt = sheet.cell(row = i + 1, column = 1)
    Tot_nxt = sheet.cell(row = i + 1, column = 5)
    Reim_nxt = sheet.cell(row = i + 1, column = 6)
    Note_nxt = sheet.cell(row = i +1, column = 7)
    Name_nxt = Name_nxt.value
    Tot_nxt = Tot_nxt.value
    Reim_nxt = Reim_nxt.value
    Note_nxt = Note_nxt.value

    nrowsEmp += 1

    # Sort by pay rate
    # The following is messy as heck. Need to fix at a later time.
    # Note to self: learn how to use arrays and goto commands in python
    if Name_nxt is not Name:
        k = nrowsEmp
        l = 0
        for j in range(i-k+1, i+1):
            valRate = sheet.cell(row=j,column=4)
            Rate = valRate.value
            valHours_tmp = sheet.cell(row=j,column=3)
            Hours_tmp = valHours_tmp.value
            valCat_tmp = sheet.cell(row=j,column=2)
            Cat = valCat_tmp.value
            l += 1
            if l == 1:
                rate1 = Rate
                rate1hrs += Hours_tmp
                if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                    rate1unarmed += Hours_tmp
                elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                    rate1armed += Hours_tmp
                elif Cat.__contains__("Admin"):
                    rate1admin += Hours_tmp
                elif Cat.__contains__("OT"):
                    rate1OT += Hours_tmp
                elif Cat.__contains__("Train"):
                    rate1train += Hours_tmp
                elif Cat.__contains__("Sick"):
                    rate1sick += Hours_tmp
                elif Cat.__contains__("COVID"):
                    rate1covid += Hours_tmp
                else:
                    print("Unknown category for hours!")
            if l == 2:
                rate2 = Rate
                if rate2 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 3:
                rate3 = Rate
                if rate3 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate3 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 4:
                rate4 = Rate
                if rate4 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate4 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate4 == rate3:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate4hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate4unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate4armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate4admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate4OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate4train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate4sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate4covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 5:
                rate5 = Rate
                if rate5 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate5 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate5 == rate3:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate5 == rate4:
                    rate4hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate4unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate4armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate4admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate4OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate4train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate4sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate4covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate5hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate5unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate5armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate5admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate5OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate5train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate5sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate5covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 6:
                rate6 = Rate
                if rate6 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate6 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate6 == rate3:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate6 == rate4:
                    rate4hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate4unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate4armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate4admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate4OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate4train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate4sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate4covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate6 == rate5:
                    rate5hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate5unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate5armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate5admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate5OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate5train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate5sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate5covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate6hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate6unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate6armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate6admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate6OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate6train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate6sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate6covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 7:
                rate7 = Rate
                if rate7 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate7 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate7 == rate3:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate7 == rate4:
                    rate4hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate4unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate4armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate4admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate4OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate4train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate4sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate4covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate7 == rate5:
                    rate5hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate5unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate5armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate5admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate5OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate5train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate5sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate5covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate7 == rate6:
                    rate6hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate6unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate6armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate6admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate6OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate6train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate6sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate6covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate7hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate7unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate7armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate7admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate7OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate7train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate7sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate7covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 8:
                rate8 = Rate
                if rate8 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate8 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate8 == rate3:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate8 == rate4:
                    rate4hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate4unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate4armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate4admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate4OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate4train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate4sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate4covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate8 == rate5:
                    rate5hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate5unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate5armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate5admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate5OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate5train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate5sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate5covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate8 == rate6:
                    rate6hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate6unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate6armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate6admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate6OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate6train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate6sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate6covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate8 == rate7:
                    rate7hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate7unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate7armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate7admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate7OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate7train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate7sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate7covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate8hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate8unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate8armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate8admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate8OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate8train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate8sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate8covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 9:
                rate9 = Rate
                if rate9 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate9 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate9 == rate3:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate9 == rate4:
                    rate4hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate4unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate4armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate4admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate4OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate4train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate4sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate4covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate9 == rate5:
                    rate5hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate5unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate5armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate5admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate5OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate5train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate5sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate5covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate9 == rate6:
                    rate6hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate6unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate6armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate6admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate6OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate6train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate6sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate6covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate9 == rate7:
                    rate7hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate7unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate7armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate7admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate7OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate7train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate7sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate7covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate9 == rate8:
                    rate8hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate8unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate8armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate8admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate8OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate8train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate8sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate8covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate9hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate9unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate9armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate9admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate9OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate9train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate9sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate9covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 10:
                rate10 = Rate
                if rate10 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate10 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate10 == rate3:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate10 == rate4:
                    rate4hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate4unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate4armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate4admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate4OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate4train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate4sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate4covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate10 == rate5:
                    rate5hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate5unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate5armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate5admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate5OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate5train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate5sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate5covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate10 == rate6:
                    rate6hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate6unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate6armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate6admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate6OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate6train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate6sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate6covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate10 == rate7:
                    rate7hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate7unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate7armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate7admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate7OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate7train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate7sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate7covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate10 == rate8:
                    rate8hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate8unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate8armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate8admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate8OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate8train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate8sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate8covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate10 == rate9:
                    rate9hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate9unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate9armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate9admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate9OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate9train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate9sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate9covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate10hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate10unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate10armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate10admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate10OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate10train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate10sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate10covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 11:
                rate11 = Rate
                if rate11 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate11 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate11 == rate3:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate11 == rate4:
                    rate4hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate4unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate4armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate4admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate4OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate4train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate4sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate4covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate11 == rate5:
                    rate5hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate5unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate5armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate5admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate5OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate5train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate5sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate5covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate11 == rate6:
                    rate6hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate6unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate6armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate6admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate6OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate6train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate6sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate6covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate11 == rate7:
                    rate7hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate7unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate7armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate7admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate7OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate7train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate7sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate7covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate11 == rate8:
                    rate8hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate8unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate8armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate8admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate8OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate8train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate8sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate8covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate11 == rate9:
                    rate9hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate9unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate9armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate9admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate9OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate9train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate9sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate9covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate11 == rate10:
                    rate10hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate10unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate10armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate10admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate10OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate10train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate10sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate10covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate11hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate11unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate11armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate11admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate11OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate11train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate11sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate11covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            if l == 12:
                rate12 = Rate
                if rate12 == rate1:
                    rate1hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate1unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate1armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate1admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate1OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate1train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate1sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate1covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate2:
                    rate2hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate2unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate2armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate2admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate2OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate2train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate2sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate2covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate3:
                    rate3hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate3unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate3armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate3admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate3OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate3train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate3sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate3covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate4:
                    rate4hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate4unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate4armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate4admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate4OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate4train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate4sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate4covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate5:
                    rate5hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate5unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate5armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate5admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate5OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate5train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate5sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate5covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate6:
                    rate6hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate6unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate6armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate6admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate6OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate6train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate6sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate6covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate7:
                    rate7hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate7unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate7armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate7admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate7OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate7train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate7sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate7covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate8:
                    rate8hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate8unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate8armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate8admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate8OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate8train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate8sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate8covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate9:
                    rate9hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate9unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate9armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate9admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate9OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate9train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate9sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate9covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate10:
                    rate10hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate10unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate10armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate10admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate10OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate10train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate10sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate10covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                elif rate12 == rate11:
                    rate11hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate11unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate11armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate11admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate11OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate11train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate11sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate11covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
                else:
                    rate12hrs += Hours_tmp
                    if Cat.__contains__("Un") and not Cat.__contains__("OT"):
                        rate12unarmed += Hours_tmp
                    elif Cat.__contains__("Armed") and not Cat.__contains__("OT"):
                        rate12armed += Hours_tmp
                    elif Cat.__contains__("Admin"):
                        rate12admin += Hours_tmp
                    elif Cat.__contains__("OT"):
                        rate12OT += Hours_tmp
                    elif Cat.__contains__("Train"):
                        rate12train += Hours_tmp
                    elif Cat.__contains__("Sick"):
                        rate12sick += Hours_tmp
                    elif Cat.__contains__("COVID"):
                        rate12covid += Hours_tmp
                    else:
                        print("Unknown category for hours!")
            #print("test0",rate1, rate1hrs, rate1unarmed, rate1armed, rate1OT)
            #print("test0",rate2, rate2hrs, rate2unarmed, rate2armed, rate2OT)

            if rate2hrs == 0:
                rate2 = 0
            if rate3hrs == 0:
                rate3 = 0
            if rate4hrs == 0:
                rate4 = 0
            if rate5hrs == 0:
                rate5 = 0
            if rate6hrs == 0:
                rate6 = 0
            if rate7hrs == 0:
                rate7 = 0
            if rate8hrs == 0:
                rate8 = 0
            if rate9hrs == 0:
                rate9 = 0
            if rate10hrs == 0:
                rate10 = 0
            if rate11hrs == 0:
                rate11 = 0
            if rate12hrs == 0:
                rate12 = 0
    # End messiness


    # border line parameters
    line = Side(border_style="thin", color="000000")
    if Name_nxt is None:
        hrsTotal = float(round(hrsTotal,2))
        Gross = str(round(Gross, 2))
        Ncell += 1
        c1 = newsheet1.cell(row = Ncell, column = 1)
        c1.value = Name
        l = 1
        for j in range(Ncell, Ncell+k):
            #print(j, Ncell, k,Ncell+k, jrow, rate1,rate1unarmed,rate1armed,rate1admin,rate1OT,"\t",rate2,rate2unarmed,rate2armed,rate2admin,rate2OT)
            c1 = newsheet1.cell(row = j-k+1, column = 3)
            c1.value = "Unarmed:"
            c1 = newsheet1.cell(row = j-k+1, column = 5)
            c1.value = "Armed:"
            c1 = newsheet1.cell(row = j-k+1, column = 7)
            c1.value = "Admin:"
            c1 = newsheet1.cell(row = j-k+1, column = 9)
            c1.value = "OT:"
            c1 = newsheet1.cell(row = j-k+1, column = 11)
            c1.value = "Train:"
            c1 = newsheet1.cell(row = j-k+1, column = 13)
            c1.value = "Sick:"
            c1 = newsheet1.cell(row = j-k+1, column = 15)
            c1.value = "COVID:"
            if l == 1:
                c1 = newsheet1.cell(row = j-k+1, column = 1)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 3)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 5)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 7)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 9)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 11)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 13)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 15)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate1
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate1unarmed
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate1armed
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate1admin
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate1OT
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate1train
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate1sick
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate1covid
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 17)
                c1.value = "Tot hrs:"
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 18)
                c1.value = hrsTotal
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 19)
                c1.value = "Gross:"
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 20)
                c1.value = Gross
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 21)
                c1.border = Border(top = line)
                if Reim != 0:
                    c1 = newsheet1.cell(row = j-k+1, column = 21)
                    c1.value = Reim
                c1 = newsheet1.cell(row = j-k+1, column = 22)
                c1.border = Border(top = line)
                if Note is not None:
                    c1 = newsheet1.cell(row = j-k+1, column = 22)
                    c1.value = Note
            if l == 2 and rate2 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate2
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate2unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate2armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate2admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate2OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate2train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate2sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate2covid
                if Reim != 0:
                    c1 = newsheet1.cell(row = j-k+1, column = 21)
                    c1.value = Reim
                if Note is not None:
                    c1 = newsheet1.cell(row = j-k+1, column = 22)
                    c1.value = Note
            if l == 3 and rate3 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate3
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate3unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate3armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate3admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate3OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate3train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate3sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate3covid
                if Reim != 0:
                    c1 = newsheet1.cell(row = j-k+1, column = 21)
                    c1.value = Reim
                if Note is not None:
                    c1 = newsheet1.cell(row = j-k+1, column = 22)
                    c1.value = Note
            if l == 4 and rate4 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate4
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate4unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate4armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate4admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate4OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate4train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate4sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate4covid
            if l == 5 and rate5 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate5
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate5unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate5armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate5admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate5OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate5train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate5sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate5covid
            if l == 6 and rate6 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate6
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate6unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate6armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate6admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate6OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate6train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate6sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate6covid
            if l == 7 and rate7 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate7
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate7unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate7armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate7admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate7OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate7train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate7sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate7covid
            if l == 8 and rate8 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate8
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate8unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate8armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate8admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate8OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate8train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate8sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate8covid
            if l == 9 and rate9 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate9
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate9unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate9armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate9admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate9OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate9train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate9sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate9covid
            l += 1

        print("Employee:",Name,"  Hours unarmed: ",hrsUnarmed," armed:",hrsArmed," admin:",hrsAdmin," OT:",hrsOT," training:",hrsTrain," sick pay:",hrsSick," COVID:",hrsCOVID," --- Total hours:",hrsTotal," Total pay:",Gross," Reim:",Reim," Note:",Note)
        #print("\trate1:",rate1,"hrs1:",rate1hrs,"rate2:",rate2,"hrs2:",rate2hrs,"rate3:",rate3,"hrs3:",rate3hrs,"rate4:",rate4,"hrs4:",rate4hrs,"rate5:",rate5,"hrs5:",rate5hrs,"rate6:",rate6,"hrs6:",rate6hrs,"rate7:",rate7,"hrs7:",rate7hrs,"rate8:",rate8,"hrs8:",rate8hrs,"rate9:",rate9,"hrs9:",rate9hrs,"hrs10:",rate10,"hrs10:",rate10hrs,"rate11:",rate11hrs,"rate12:",rate12,"hrs12:",rate12hrs)
        break
    if Name_nxt not in Name:
        hrsTotal = float(round(hrsTotal,2))
        Gross = str(round(Gross, 2))
        Ncell += k
        c1 = newsheet1.cell(row = Ncell-k+1, column = 1)
        c1.value = Name
        l = 1
        for j in range(Ncell, Ncell+k):
            c1 = newsheet1.cell(row = j-k+1, column = 3)
            c1.value = "Unarmed:"
            c1 = newsheet1.cell(row = j-k+1, column = 5)
            c1.value = "Armed:"
            c1 = newsheet1.cell(row = j-k+1, column = 7)
            c1.value = "Admin:"
            c1 = newsheet1.cell(row = j-k+1, column = 9)
            c1.value = "OT:"
            c1 = newsheet1.cell(row = j-k+1, column = 11)
            c1.value = "Train:"
            c1 = newsheet1.cell(row = j-k+1, column = 13)
            c1.value = "Sick:"
            c1 = newsheet1.cell(row = j-k+1, column = 15)
            c1.value = "COVID:"
            if l == 1:
                c1 = newsheet1.cell(row = j-k+1, column = 1)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 3)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 5)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 7)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 9)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 11)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 13)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 15)
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate1
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate1unarmed
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate1armed
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate1admin
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate1OT
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate1train
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate1sick
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate1covid
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 17)
                c1.value = "Tot hrs:"
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 18)
                c1.value = hrsTotal
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 19)
                c1.value = "Gross:"
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 20)
                c1.value = Gross
                c1.border = Border(top = line)
                c1 = newsheet1.cell(row = j-k+1, column = 21)
                c1.border = Border(top = line)
                if Reim != 0:
                    c1 = newsheet1.cell(row = j-k+1, column = 21)
                    c1.value = Reim
                c1 = newsheet1.cell(row = j-k+1, column = 22)
                c1.border = Border(top = line)
                if Note is not None:
                    c1 = newsheet1.cell(row = j-k+1, column = 22)
                    c1.value = Note
            if l == 2 and rate2 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate2
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate2unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate2armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate2admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate2OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate2train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate2sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate2covid
                if Reim != 0:
                    c1 = newsheet1.cell(row = j-k+1, column = 21)
                    c1.value = Reim
                if Note is not None:
                    c1 = newsheet1.cell(row = j-k+1, column = 22)
                    c1.value = Note
            if l == 3 and rate3 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate3
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate3unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate3armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate3admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate3OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate3train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate3sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate3covid
                if Reim != 0:
                    c1 = newsheet1.cell(row = j-k+1, column = 21)
                    c1.value = Reim
                if Note is not None:
                    c1 = newsheet1.cell(row = j-k+1, column = 22)
                    c1.value = Note
            if l == 4 and rate4 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate4
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate4unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate4armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate4admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate4OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate4train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate4sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate4covid
            if l == 5 and rate5 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate5
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate5unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate5armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate5admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate5OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate5train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate5sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate5covid
            if l == 6 and rate6 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate6
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate6unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate6armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate6admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate6OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate6train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate6sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate6covid
            if l == 7 and rate7 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate7
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate7unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate7armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate7admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate7OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate7train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate7sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate7covid
            if l == 8 and rate8 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate8
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate8unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate8armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate8admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate8OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate8train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate8sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate8covid
            if l == 9 and rate9 != 0:
                c1 = newsheet1.cell(row = j-k+1, column = 2)
                c1.value = rate9
                c1 = newsheet1.cell(row = j-k+1, column = 4)
                c1.value = rate9unarmed
                c1 = newsheet1.cell(row = j-k+1, column = 6)
                c1.value = rate9armed
                c1 = newsheet1.cell(row = j-k+1, column = 8)
                c1.value = rate9admin
                c1 = newsheet1.cell(row = j-k+1, column = 10)
                c1.value = rate9OT
                c1 = newsheet1.cell(row = j-k+1, column = 12)
                c1.value = rate9train
                c1 = newsheet1.cell(row = j-k+1, column = 14)
                c1.value = rate9sick
                c1 = newsheet1.cell(row = j-k+1, column = 16)
                c1.value = rate9covid
            l += 1

        print("Employee:",Name,"  Hours unarmed: ",hrsUnarmed," armed:",hrsArmed," admin:",hrsAdmin," OT:",hrsOT," training:",hrsTrain," sick pay:",hrsSick," COVID:",hrsCOVID," --- Total hours:",hrsTotal," Total pay:",Gross," Reim:",Reim," Note:",Note)
        #print("\trate1:",rate1,"hrs1:",rate1hrs,"rate2:",rate2,"hrs2:",rate2hrs,"rate3:",rate3,"hrs3:",rate3hrs,"rate4:",rate4,"hrs4:",rate4hrs,"rate5:",rate5,"hrs5:",rate5hrs,"rate6:",rate6,"hrs6:",rate6hrs,"rate7:",rate7,"hrs7:",rate7hrs,"rate8:",rate8,"hrs8:",rate8hrs,"rate9:",rate9,"hrs9:",rate9hrs,"hrs10:",rate10,"hrs10:",rate10hrs,"rate11:",rate11hrs,"rate12:",rate12,"hrs12:",rate12hrs)

        # Clear values for next employee
        Name = Name_nxt # next employee
        Gross = 0
        hrsUnarmed = 0
        hrsArmed = 0
        hrsAdmin = 0
        hrsOT = 0
        hrsTrain = 0
        hrsSick = 0
        hrsCOVID = 0
        hrsTotal = 0
        nrowsEmp = 0
        rate1 = 0
        rate2 = 0
        rate3 = 0
        rate4 = 0
        rate5 = 0
        rate6 = 0
        rate7 = 0
        rate8 = 0
        rate9 = 0
        rate10 = 0
        rate11 = 0
        rate12 = 0
        rate1hrs = 0
        rate2hrs = 0
        rate3hrs = 0
        rate4hrs = 0
        rate5hrs = 0
        rate6hrs = 0
        rate7hrs = 0
        rate8hrs = 0
        rate9hrs = 0
        rate10hrs = 0
        rate11hrs = 0
        rate12hrs = 0
        rate1unarmed = 0
        rate2unarmed = 0
        rate3unarmed = 0
        rate4unarmed = 0
        rate5unarmed = 0
        rate6unarmed = 0
        rate7unarmed = 0
        rate8unarmed = 0
        rate9unarmed = 0
        rate10unarmed = 0
        rate11unarmed = 0
        rate12unarmed = 0
        rate1armed = 0
        rate2armed = 0
        rate3armed = 0
        rate4armed = 0
        rate5armed = 0
        rate6armed = 0
        rate7armed = 0
        rate8armed = 0
        rate9armed = 0
        rate10armed = 0
        rate11armed = 0
        rate12armed = 0
        rate1admin = 0
        rate2admin = 0
        rate3admin = 0
        rate4admin = 0
        rate5admin = 0
        rate6admin = 0
        rate7admin = 0
        rate8admin = 0
        rate9admin = 0
        rate10admin = 0
        rate11admin = 0
        rate12admin = 0
        rate1OT = 0
        rate2OT = 0
        rate3OT = 0
        rate4OT = 0
        rate5OT = 0
        rate6OT = 0
        rate7OT = 0
        rate8OT = 0
        rate9OT = 0
        rate10OT = 0
        rate11OT = 0
        rate12OT = 0
        rate1train = 0
        rate2train = 0
        rate3train = 0
        rate4train = 0
        rate5train = 0
        rate6train = 0
        rate7train = 0
        rate8train = 0
        rate9train = 0
        rate10train = 0
        rate11train = 0
        rate12train = 0
        rate1sick = 0
        rate2sick = 0
        rate3sick = 0
        rate4sick = 0
        rate5sick = 0
        rate6sick = 0
        rate7sick = 0
        rate8sick = 0
        rate9sick = 0
        rate10sick = 0
        rate11sick = 0
        rate12sick = 0
        rate1covid = 0
        rate2covid = 0
        rate3covid = 0
        rate4covid = 0
        rate5covid = 0
        rate6covid = 0
        rate7covid = 0
        rate8covid = 0
        rate9covid = 0
        rate10covid = 0
        rate11covid = 0
        rate12covid = 0
        # end main loop

# Excel spreadhseet options
newsheet1.column_dimensions["A"].width = 25
newsheet1.column_dimensions["B"].width = 6
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[1]
    cell.alignment = Alignment(horizontal='left')
newsheet1.column_dimensions["C"].width = 10
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[2]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["D"].width = 6
newsheet1.column_dimensions["E"].width = 8
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[4]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["F"].width = 7
newsheet1.column_dimensions["G"].width = 8
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[6]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["H"].width = 7
newsheet1.column_dimensions["I"].width = 6
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[8]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["J"].width = 7
newsheet1.column_dimensions["K"].width = 8
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[10]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["L"].width = 5
newsheet1.column_dimensions["M"].width = 6
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[12]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["N"].width = 5
newsheet1.column_dimensions["O"].width = 8
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[14]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["P"].width = 5
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[15]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["Q"].width = 8
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[16]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["R"].width = 6
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[17]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["S"].width = 7
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[18]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["T"].width = 9
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[19]
    cell.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["U"].width = 5
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[20]
    cell.alignment = Alignment(horizontal='right')
for row in newsheet1[2:newsheet1.max_row]:
    cell = row[21]
    cell.alignment = Alignment(horizontal='left')

for i in range(1, 12): # clean empty rows
    for row in range(2,newsheet1.max_row):
        c1 = newsheet1.cell(row, column = 2)
        c1 = c1.value
        if c1 == 0:
            newsheet1.delete_rows(row,1)
        if c1 is None:
            newsheet1.delete_rows(row,1)

GrandHrs = str(round(GrandHrs, 2))
GrandTot = str(round(GrandTot, 2))
GrandReim = str(round(GrandReim, 2))
GrandGross = str(round(GrandGross, 2))

rowmax = newsheet1.max_row
line1 = Side(border_style="medium", color="000000")
c0 = newsheet1.cell(row = rowmax+2, column = 2)
c0.value = "Total hours"
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+2, column = 3)
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+2, column = 4)
c0.border = Border(bottom = line1)
c0.value = "Total pay"
c0 = newsheet1.cell(row = rowmax+2, column = 5)
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+2, column = 6)
c0.border = Border(bottom = line1)
c0.value = "Reimb."
c0 = newsheet1.cell(row = rowmax+2, column = 7)
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+2, column = 8)
c0.border = Border(bottom = line1)
c0.value = "Gross pay"
c0 = newsheet1.cell(row = rowmax+2, column = 9)
c0.border = Border(bottom = line1)
cGrandHrs = newsheet1.cell(row = rowmax+3, column = 2)
cGrandHrs.value = GrandHrs
cGrandTot = newsheet1.cell(row = rowmax+3, column = 4)
cGrandTot.value = GrandTot
cGrandReim = newsheet1.cell(row = rowmax+3, column = 6)
cGrandReim.value = GrandReim
cGrandGross = newsheet1.cell(row = rowmax+3, column = 8)
cGrandGross.value = GrandGross
c0 = newsheet1.cell(row = rowmax+5, column = 2)
c0.value = "Unarmed hours"
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 3)
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 4)
c0.value = "Armed"
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 5)
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 6)
c0.value = "Admin"
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 7)
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 8)
c0.value = "OT"
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 9)
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 10)
c0.value = "Training"
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 11)
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 12)
c0.value = "Sick pay"
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 13)
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 14)
c0.value = "COVID"
c0.border = Border(bottom = line1)
c0 = newsheet1.cell(row = rowmax+5, column = 15)
c0.border = Border(bottom = line1)
cGrandUnarmed = newsheet1.cell(row = rowmax+6, column = 2)
cGrandUnarmed.value = GrandUnarmed
cGrandUnarmed.alignment = Alignment(horizontal = 'left')
cGrandArmed = newsheet1.cell(row = rowmax+6, column = 4)
cGrandArmed.value = GrandArmed
cGrandArmed.alignment = Alignment(horizontal = 'left')
cGrandAdmin = newsheet1.cell(row = rowmax+6, column = 6)
cGrandAdmin.value = GrandAdmin
cGrandAdmin.alignment = Alignment(horizontal = 'left')
cGrandOT = newsheet1.cell(row = rowmax+6, column = 8)
cGrandOT.value = GrandOT
cGrandOT.alignment = Alignment(horizontal = 'left')
cGrandTrain = newsheet1.cell(row = rowmax+6, column = 10)
cGrandTrain.value = GrandTrain
cGrandTrain.alignment = Alignment(horizontal = 'left')
cGrandSick = newsheet1.cell(row = rowmax+6, column = 12)
cGrandSick.value = GrandSick
cGrandSick.alignment = Alignment(horizontal = 'left')
cGrandCovid = newsheet1.cell(row = rowmax+6, column = 14)
cGrandCovid.value = GrandCovid
cGrandCovid.alignment = Alignment(horizontal = 'left')

newbook1.save("/mnt/c/Users/Jacob/macros/Payroll/payroll_output.xlsx")

print()
print("Total hours:",GrandHrs,"\t Total payroll:",GrandTot,"\t Reimbursements:",GrandReim,"\t Total Gross payroll:",GrandGross)
print("Total unarmed:",GrandUnarmed,"\t armed:",GrandArmed,"\t admin:",GrandAdmin,"\t OT:",GrandOT,"\t training:",GrandTrain,"\t sicktime:",GrandSick,"\t COVID:",GrandCovid)
print()
print()
print("file output written to /mnt/c/Users/Jacob/macros/Payroll/payroll_output.xlsx")
print()
print("Done")
    #print(i,Name,Tot,Tot_nxt)
#    if Name_nxt == Name:
#        print(i,Name,Name_nxt)
    #print(Name,"\t",Cat,"\t",Hours,"\t",Rate,"\t",Tot,"\t",Reim,"\t",Gross)


### Next iterration of this code will use arrays to make everything simpler, I promise.
