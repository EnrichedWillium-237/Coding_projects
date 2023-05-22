# Code for calculating OT values
# Reads in file "output_sorted.xlsx"
# Calculates overtime hours

# Source files
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Alignment, Border, Side, Font
from datetime import datetime, date, timedelta

flagDebug1 = False
flagDebug2 = False

# Input file
workbook = load_workbook('output_sorted.xlsx')

sheet = workbook.active
Nrow = sheet.max_row
daterange = sheet['C']
date_list = [daterange[x].value for x in range(2, len(daterange) - 10)]
import datetime
week1start = min(date_list)
week1end = week1start + datetime.timedelta(days = 6)
week2start = week1end + datetime.timedelta(days = 1)
week2end = max(date_list)

# Calculation for crosschecks
GrandHrs = 0
GrandNames = 0
colENum = 1
colPos  = 2
colDate = 3
colHrs  = 6
colName = 7
colRate = 8
for i in range(2, Nrow):
    valHrs = sheet.cell(row = i, column = colHrs)
    GrandHrs += valHrs.value
    valName = sheet.cell(row = i, column = colName).value
    valNameNxt = sheet.cell(row = i + 1, column = colName).value
    if valNameNxt not in valName: GrandNames += 1
print("\n")
print("--- Week 1:", week1start.strftime("%Y-%m-%d"), "to", week1end.strftime("%Y-%m-%d"), "---")
print("--- Week 2:", week2start.strftime("%Y-%m-%d"), "to", week2end.strftime("%Y-%m-%d"), "---")
print("Overall stats:")
print("  Total number of employees:  ", GrandNames)
print("  Total number of shifts:  ", Nrow - 1)
print("  Total hours for all names and positions:  ", f'{GrandHrs:.9}')
print("\n")

# Setup output file and spreadsheet
newbook1 = openpyxl.Workbook()
newsheet1 = newbook1.active
printCnt = 8 # Needed for printing out hours
printCntInit = 8

rowCnt = 1
warnShift = "\nEMPLOYEE HAS WORKED TOO MANY SHIFTS IN ONE WEEK!!!  GIVE THEM SOME TIME OFF!!!\n"
flagMultShiftTot = False
# Main event loop
for i in range(2, Nrow + 1):
    valName = sheet.cell(row = i, column = colName).value
    valNameNxt = sheet.cell(row = i + 1, column = colName).value
    if valNameNxt is not None and valNameNxt.__contains__(valName):
        rowCnt += 1
    else:
        rowmin = i - rowCnt + 1
        rowmax = i
        emptyList = [None, None, 0, 0, 0, 0, 0, None, 0]
        listWeek1 = [emptyList, emptyList, emptyList, emptyList, emptyList,
                     emptyList, emptyList, emptyList, emptyList]
        listWeek2 = [emptyList, emptyList, emptyList, emptyList, emptyList,
                     emptyList, emptyList, emptyList, emptyList]

        # Begin OT calculation
        # Find middle cell between week 1 and week 2
        rowmid = 0
        rowmidcnt = 0
        week1Hrs = 0
        week2Hrs = 0
        for j in range(rowmin, rowmax + 1):
            valDate = sheet.cell(row = j, column = colDate).value
            valHrs = sheet.cell(row = j, column = colHrs).value
            if valHrs > 0 and valHrs < 0.01: valHrs = 0
            if valDate <= week1end:
                week1Hrs += valHrs
                rowmidcnt += 1
            else:
                week2Hrs += valHrs
            rowmid = rowmin + rowmidcnt

        # OT +12 for week one and week 2 without 40+
        if week1Hrs <= 40:
            regHrs = 0
            dayHrsCnt = 0
            dayHrsGap = 0
            flagMultShift1 = False
            valName = sheet.cell(row = rowmid-1, column = colName).value
            list1 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list2 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list3 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list4 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list5 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list6 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list7 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list8 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list9 = [None, None, 0, 0, 0, 0, 0, None, 0]
            for j in range(rowmid - 1, rowmin - 1, -1): # minus one offset in loop because we're counting backwards
                valPos = sheet.cell(row = j, column = colPos).value
                valDate = sheet.cell(row = j, column = colDate).value
                valDateNxt = sheet.cell(row = j + 1, column = colDate).value
                valRate = sheet.cell(row = j, column = colRate).value
                valENum = sheet.cell(row = j, column = colENum).value
                if j > 2 and valDateNxt is not None and valDateNxt.day is valDate.day:
                    flagMultShift1 = True
                    flagMultShiftTot = True
                    if flagDebug1: print("---Multiple shifts in same day this week. Evaluate OT+12 by hand!---")
                valHrs = sheet.cell(row = j, column = colHrs).value
                if valHrs > 0 and valHrs < 0.01: valHrs = 0
                if valHrs > 12:
                    z = valHrs - 12
                    regHrs = valHrs - z
                else:
                    z = 0
                    regHrs = valHrs
                if regHrs > 0 and regHrs < 0.01: regHrs = 0
                if z > 0 and z < 0.01: z = 0
                if j == rowmid - 1:   list1 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmid - 2: list2 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmid - 3: list3 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmid - 4: list4 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmid - 5: list5 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmid - 6: list6 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmid - 7: list7 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmid - 8: list8 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmid - 9: list9 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                if j <= rowmid - 10: print(warnShift, "  ", valName)
                if flagDebug1: print(valName, "  ", valPos, "  Rate:", valRate, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", z)
            listWeek1 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            if flagDebug1: print(valName, "  Week 1 --- total: ", week1Hrs, " shift+40 total: ", 0, "\n")
        if week2Hrs <= 40:
            regHrs = 0
            flagMultShift2 = False
            valName = sheet.cell(row = rowmax, column = colName).value
            list1 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list2 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list3 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list4 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list5 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list6 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list7 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list8 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list9 = [None, None, 0, 0, 0, 0, 0, None, 0]
            for j in range(rowmax, rowmid - 1, -1):
                valPos = sheet.cell(row = j, column = colPos).value
                valDate = sheet.cell(row = j, column = colDate).value
                valDateNxt = sheet.cell(row = j + 1, column = colDate).value
                valRate = sheet.cell(row = j, column = colRate).value
                valENum = sheet.cell(row = j, column = colENum).value
                if j > 2 and valDateNxt is not None and valDateNxt.day is valDate.day:
                    flagMultShift2 = True
                    flagMultShiftTot = True
                    if flagDebug1: print("---Multiple shifts in same day this week. Evaluate OT+12 by hand!---")
                valHrs = sheet.cell(row = j, column = colHrs).value
                if valHrs > 0 and valHrs < 0.01: valHrs = 0
                if valHrs > 12:
                    z = valHrs - 12
                    regHrs = valHrs - z
                else:
                    z = 0
                    regHrs = valHrs
                if regHrs > 0 and regHrs < 0.01: regHrs = 0
                if z > 0 and z < 0.01: z = 0
                if j == rowmax:       list1 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmax - 1: list2 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmax - 2: list3 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmax - 3: list4 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmax - 4: list5 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmax - 5: list6 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmax - 6: list7 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmax - 7: list8 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                elif j == rowmax - 8: list9 = [valName, valPos, valHrs, regHrs, z, 0, valRate, valENum, 0]
                if j <= rowmid - 9: print(warnShift, "  ", valName)
                if flagDebug1: print(valName, "  ", valPos, "  Rate:", valRate, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", z)
            listWeek2 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            if flagDebug1: print(valName, "  Week 2 --- total: ", week2Hrs, " shift+40 total: ", 0, "\n")

        # OT +40 for week one
        if week1Hrs > 40:
            regHrs = 0
            OT40week1 = 0
            OT12 = 0
            OT40week1 = week1Hrs - 40
            OT = OT40week1
            OTn = OT
            OTcnt = 0
            flag1 = False
            flagMultShift1 = False
            valName = sheet.cell(row = rowmid-1, column = colName)
            valName = valName.value
            list1 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list2 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list3 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list4 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list5 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list6 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list7 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list8 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list9 = [None, None, 0, 0, 0, 0, 0, None, 0]
            for j in range(rowmid - 1, rowmin - 1, -1): # minus one offset in for loop because we're counting backwards
                valPos = sheet.cell(row = j, column = colPos).value
                valDate = sheet.cell(row = j, column = colDate).value
                valDateNxt = sheet.cell(row = j + 1, column = colDate).value
                valRate = sheet.cell(row = j, column = colRate).value
                valENum = sheet.cell(row = j, column = colENum).value
                if j > 2 and valDateNxt is not None and valDateNxt.day is valDate.day:
                    flagMultShift1 = True
                    flagMultShiftTot = True
                    if flagDebug1: print("---Multiple shifts in same day this week. Evaluate OT+12 by hand!---")
                valHrs = sheet.cell(row = j, column = colHrs).value
                if valHrs > 0 and valHrs < 0.01: valHrs = 0
                x = valHrs
                if flag1 is True:
                    OTn = 0
                    y = 0
                if flag1 is False:
                    OTn = OTn - x
                    if OTn > 0:
                        y = x
                        OTcnt += y
                    if OTn <= 0:
                        y = OT - OTcnt
                        flag1 = True
                if valHrs > 12: # OT +12 calculation
                    z = valHrs - 12
                    if y > z: z = 0
                    else: z = z - y
                else: z = 0
                regHrs = valHrs - y - z
                if y + z == 0: regHrs = valHrs
                if regHrs > 0 and regHrs < 0.01: regHrs = 0
                if z > 0 and z < 0.01: z = 0
                if y > 0 and y < 0.01: y = 0
                if j == rowmid - 1:   list1 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmid - 2: list2 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmid - 3: list3 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmid - 4: list4 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmid - 5: list5 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmid - 6: list6 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmid - 7: list7 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmid - 8: list8 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmid - 9: list9 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                if j <= rowmid - 10: print(warnShift, "  ", valName)
                if flagDebug1: print(valName, "  ", valPos, "  Rate:", valRate, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", y)

            listWeek1 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            if flagDebug1: print(valName,"  Week 1 --- total: ", week1Hrs," shift+40 total: ",OT40week1,"\n")

        # OT +40 for week two
        if week2Hrs > 40:
            regHrs = 0
            OT40week2 = 0
            OT40week2 = week2Hrs - 40
            OT = OT40week2
            OTn = OT
            OTcnt = 0
            flag1 = False
            flagMultShift2 = False
            valName = sheet.cell(row = rowmax, column = colName)
            valName = valName.value
            list1 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list2 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list3 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list4 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list5 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list6 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list7 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list8 = [None, None, 0, 0, 0, 0, 0, None, 0]
            list9 = [None, None, 0, 0, 0, 0, 0, None, 0]
            for j in range(rowmax, rowmid - 1, -1):
                valPos = sheet.cell(row = j, column = colPos).value
                valDate = sheet.cell(row = j, column = colDate).value
                valDateNxt = sheet.cell(row = j + 1, column = colDate).value
                valRate = sheet.cell(row = j, column = colRate).value
                valENum = sheet.cell(row = j, column = colENum).value
                if j > 2 and valDateNxt is not None and valDateNxt.day is valDate.day:
                    flagMultShift2 = True
                    flagMultShiftTot = True
                    if flagDebug1: print("---Multiple shifts in same day this week. Evaluate OT+12 by hand!---")
                valHrs = sheet.cell(row = j, column = colHrs).value
                if valHrs > 0 and valHrs < 0.01: valHrs = 0
                x = valHrs
                if flag1 is True:
                    OTn = 0
                    y = 0
                if flag1 is False:
                    OTn = OTn - x
                    if OTn > 0:
                        y = x
                        OTcnt += y
                    if OTn <= 0:
                        y = OT - OTcnt
                        flag1 = True
                if valHrs > 12: # OT +12 calculation
                    z = valHrs - 12
                    if y > z: z = 0
                    else: z = z - y
                else: z = 0
                regHrs = valHrs - y - z
                if y + z == 0: regHrs = valHrs
                if regHrs > 0 and regHrs < 0.01: regHrs = 0
                if z > 0 and z < 0.01: z = 0
                if y > 0 and y < 0.01: y = 0
                if j == rowmax:       list1 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmax - 1: list2 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmax - 2: list3 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmax - 3: list4 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmax - 4: list5 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmax - 5: list6 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmax - 6: list7 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmax - 7: list8 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                elif j == rowmax - 8: list9 = [valName, valPos, valHrs, regHrs, z, y, valRate, valENum, 0]
                if j <= rowmax - 9: print(warnShift, "  ", valName)
                if flagDebug1: print(valName, "  ", valPos, "  Rate:", valRate, "  Total:", valHrs, "  Standard: ", regHrs, "  OT+12:", y)

            listWeek2 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            if flagDebug1: print(valName,"  Week 2 --- total: ", week2Hrs," shift+40 total: ",OT40week2,"\n")

        # Sort by position
        for j in range(0, 2):
            if j == 0:
                list1 = (listWeek1[0])
                list2 = (listWeek1[1])
                list3 = (listWeek1[2])
                list4 = (listWeek1[3])
                list5 = (listWeek1[4])
                list6 = (listWeek1[5])
                list7 = (listWeek1[6])
                list8 = (listWeek1[7])
                list9 = (listWeek1[8])
            else:
                list1 = (listWeek2[0])
                list2 = (listWeek2[1])
                list3 = (listWeek2[2])
                list4 = (listWeek2[3])
                list5 = (listWeek2[4])
                list6 = (listWeek2[5])
                list7 = (listWeek2[6])
                list8 = (listWeek2[7])
                list9 = (listWeek2[8])
            pos1 = list1[1]
            pos2 = list2[1]
            pos3 = list3[1]
            pos4 = list4[1]
            pos5 = list5[1]
            pos6 = list6[1]
            pos7 = list7[1]
            pos8 = list8[1]
            pos9 = list9[1]
            hrsPos1 = list1[2]
            RegHrs1 = list1[3]
            totOT12_1 = list1[4]
            totOT40_1 = list1[5]
            nshift = 0
            # Position 1
            if pos1 is not None:
                if pos2 is not None and pos2 in pos1:
                    hrsPos1 += list2[2]
                    RegHrs1 += list2[3]
                    totOT12_1 += list2[4]
                    totOT40_1 += list2[5]
                    nshift += 1
                if pos3 is not None and pos3 in pos1:
                    hrsPos1 += list3[2]
                    RegHrs1 += list3[3]
                    totOT12_1 += list3[4]
                    totOT40_1 += list3[5]
                    nshift += 1
                if pos4 is not None and pos4 in pos1:
                    hrsPos1 += list4[2]
                    RegHrs1 += list4[3]
                    totOT12_1 += list4[4]
                    totOT40_1 += list4[5]
                    nshift += 1
                if pos5 is not None and pos5 in pos1:
                    hrsPos1 += list5[2]
                    RegHrs1 += list5[3]
                    totOT12_1 += list5[4]
                    totOT40_1 += list5[5]
                    nshift += 1
                if pos6 is not None and pos6 in pos1:
                    hrsPos1 += list6[2]
                    RegHrs1 += list6[3]
                    totOT12_1 += list6[4]
                    totOT40_1 += list6[5]
                    nshift += 1
                if pos7 is not None and pos7 in pos1:
                    hrsPos1 += list7[2]
                    RegHrs1 += list7[3]
                    totOT12_1 += list7[4]
                    totOT40_1 += list7[5]
                    nshift += 1
                if pos8 is not None and pos8 in pos1:
                    hrsPos1 += list8[2]
                    RegHrs1 += list8[3]
                    totOT12_1 += list8[4]
                    totOT40_1 += list8[5]
                    nshift += 1
                if pos9 is not None and pos9 in pos1:
                    hrsPos1 += list9[2]
                    RegHrs1 += list9[3]
                    totOT12_1 += list9[4]
                    totOT40_1 += list9[5]
                    nshift += 1
            list1 = [valName, pos1, hrsPos1, RegHrs1, totOT12_1, totOT40_1, list1[6], list1[7], nshift]
            # Position 2
            hrsPos2 = 0
            RegHrs2 = 0
            totOT12_2 = 0
            totOT40_2 = 0
            nshift = 0
            if pos2 is not None:
                if (    pos1 is not None and
                        pos2 not in pos1):
                    hrsPos2 = list2[2]
                    RegHrs2 = list2[3]
                    totOT12_2 = list2[4]
                    totOT40_2 = list2[5]
                    nshift += 1
                    if pos3 is not None and pos3 in pos2:
                        hrsPos2 += list3[2]
                        RegHrs2 += list3[3]
                        totOT12_2 += list3[4]
                        totOT40_2 += list3[5]
                        nshift += 1
                    if pos4 is not None and pos4 in pos2:
                        hrsPos2 += list4[2]
                        RegHrs2 += list4[3]
                        totOT12_2 += list4[4]
                        totOT40_2 += list4[5]
                        nshift += 1
                    if pos5 is not None and pos5 in pos2:
                        hrsPos2 += list5[2]
                        RegHrs2 += list5[3]
                        totOT12_2 += list5[4]
                        totOT40_2 += list5[5]
                        nshift += 1
                    if pos6 is not None and pos6 in pos2:
                        hrsPos2 += list6[2]
                        RegHrs2 += list6[3]
                        totOT12_2 += list6[4]
                        totOT40_2 += list6[5]
                    if pos7 is not None and pos7 in pos2:
                        hrsPos2 += list7[2]
                        RegHrs2 += list7[3]
                        totOT12_2 += list7[4]
                        totOT40_2 += list7[5]
                        nshift += 1
                    if pos8 is not None and pos8 in pos2:
                        hrsPos2 += list8[2]
                        RegHrs2 += list8[3]
                        totOT12_2 += list8[4]
                        totOT40_2 += list8[5]
                        nshift += 1
                    if pos9 is not None and pos9 in pos2:
                        hrsPos2 += list9[2]
                        RegHrs2 += list9[3]
                        totOT12_2 += list9[4]
                        totOT40_2 += list9[5]
                        nshift += 1
                list2 = [valName, pos2, hrsPos2, RegHrs2, totOT12_2, totOT40_2, list2[6], list2[7], nshift]
            # Position 3
            hrsPos3 = 0
            RegHrs3 = 0
            totOT12_3 = 0
            totOT40_3 = 0
            nshift = 0
            if pos3 is not None:
                if (    pos2 is not None and pos1 is not None and
                        pos3 not in pos1 and pos3 not in pos2):
                    hrsPos3 += list3[2]
                    RegHrs3 += list3[3]
                    totOT12_3 += list3[4]
                    totOT40_3 += list3[5]
                    nshift += 1
                    if pos4 is not None and pos4 in pos3:
                        hrsPos3 += list4[2]
                        RegHrs3 += list4[3]
                        totOT12_3 += list4[4]
                        totOT40_3 += list4[5]
                        nshift += 1
                    if pos5 is not None and pos5 in pos3:
                        hrsPos3 += list5[2]
                        RegHrs3 += list5[3]
                        totOT12_3 += list5[4]
                        totOT40_3 += list5[5]
                        nshift += 1
                    if pos6 is not None and pos6 in pos3:
                        hrsPos3 += list6[2]
                        RegHrs3 += list6[3]
                        totOT12_3 += list6[4]
                        totOT40_3 += list6[5]
                        nshift += 1
                    if pos7 is not None and pos7 in pos3:
                        hrsPos3 += list7[2]
                        RegHrs3 += list7[3]
                        totOT12_3 += list7[4]
                        totOT40_3 += list7[5]
                        nshift += 1
                    if pos8 is not None and pos8 in pos3:
                        hrsPos3 += list8[2]
                        RegHrs3 += list8[3]
                        totOT12_3 += list8[4]
                        totOT40_3 += list8[5]
                        nshift += 1
                    if pos9 is not None and pos9 in pos3:
                        hrsPos3 += list9[2]
                        RegHrs3 += list9[3]
                        totOT12_3 += list9[4]
                        totOT40_3 += list9[5]
                        nshift += 1
                list3 = [valName, pos3, hrsPos3, RegHrs3, totOT12_3, totOT40_3, list3[6], list3[7], nshift]
            # Position 4
            hrsPos4 = 0
            RegHrs4 = 0
            totOT12_4 = 0
            totOT40_4 = 0
            nshift = 0
            if pos4 is not None:
                if (    pos3 is not None and pos2 is not None and pos1 is not None and
                        pos4 not in pos1 and pos4 not in pos2 and pos4 not in pos3):
                    hrsPos4 += list4[2]
                    RegHrs4 += list4[3]
                    totOT12_4 += list4[4]
                    totOT40_4 += list4[5]
                    nshift += 1
                    if pos5 is not None and pos5 in pos4:
                        hrsPos4 += list5[2]
                        RegHrs4 += list5[3]
                        totOT12_4 += list5[4]
                        totOT40_4 += list5[5]
                        nshift += 1
                    if pos6 is not None and pos6 in pos4:
                        hrsPos4 += list6[2]
                        RegHrs4 += list6[3]
                        totOT12_4 += list6[4]
                        totOT40_4 += list6[5]
                        nshift += 1
                    if pos7 is not None and pos7 in pos4:
                        hrsPos4 += list7[2]
                        RegHrs4 += list7[3]
                        totOT12_4 += list7[4]
                        totOT40_4 += list7[5]
                        nshift += 1
                    if pos8 is not None and pos8 in pos4:
                        hrsPos4 += list8[2]
                        RegHrs4 += list8[3]
                        totOT12_4 += list8[4]
                        totOT40_4 += list8[5]
                        nshift += 1
                    if pos9 is not None and pos9 in pos4:
                        hrsPos4 += list9[2]
                        RegHrs4 += list9[3]
                        totOT12_4 += list9[4]
                        totOT40_4 += list9[5]
                        nshift += 1
                list4 = [valName, pos4, hrsPos4, RegHrs4, totOT12_4, totOT40_4, list4[6], list4[7], nshift]
            # Position 5
            hrsPos5 = 0
            RegHrs5 = 0
            totOT12_5 = 0
            totOT40_5 = 0
            nshift = 0
            if pos5 is not None:
                if (    pos4 is not None and pos3 is not None and pos2 is not None and
                        pos1 is not None and
                        pos5 not in pos1 and pos5 not in pos2 and pos5 not in pos3 and
                        pos5 not in pos4):
                    hrsPos5 += list5[2]
                    RegHrs5 += list5[3]
                    totOT12_5 += list5[4]
                    totOT40_5 += list5[5]
                    nshift += 1
                    if pos6 is not None and pos6 in pos5:
                        hrsPos5 += list6[2]
                        RegHrs5 += list6[3]
                        totOT12_5 += list6[4]
                        totOT40_5 += list6[5]
                        nshift += 1
                    if pos7 is not None and pos7 in pos5:
                        hrsPos5 += list7[2]
                        RegHrs5 += list7[3]
                        totOT12_5 += list7[4]
                        totOT40_5 += list7[5]
                        nshift += 1
                    if pos8 is not None and pos8 in pos5:
                        hrsPos5 += list8[2]
                        RegHrs5 += list8[3]
                        totOT12_5 += list8[4]
                        totOT40_5 += list8[5]
                        nshift += 1
                    if pos9 is not None and pos9 in pos5:
                        hrsPos5 += list9[2]
                        RegHrs5 += list9[3]
                        totOT12_5 += list9[4]
                        totOT40_5 += list9[5]
                        nshift += 1
                list5 = [valName, pos5, hrsPos5, RegHrs5, totOT12_5, totOT40_5, list5[6], list5[7], nshift]
            # Position 6
            hrsPos6 = 0
            RegHrs6 = 0
            totOT12_6 = 0
            totOT40_6 = 0
            nshift = 0
            if pos6 is not None:
                if (    pos5 is not None and pos4 is not None and pos3 is not None and
                        pos2 is not None and pos1 is not None and
                        pos6 not in pos1 and pos6 not in pos2 and pos6 not in pos3 and
                        pos6 not in pos4 and pos6 not in pos5):
                    hrsPos6 += list6[2]
                    RegHrs6 += list6[3]
                    totOT12_6 += list6[4]
                    totOT40_6 += list6[5]
                    nshift += 1
                    if pos7 is not None and pos7 in pos6:
                        hrsPos6 += list7[2]
                        RegHrs6 += list7[3]
                        totOT12_6 += list7[4]
                        totOT40_6 += list7[5]
                        nshift += 1
                    if pos8 is not None and pos8 in pos6:
                        hrsPos6 += list8[2]
                        RegHrs6 += list8[3]
                        totOT12_6 += list8[4]
                        totOT40_6 += list8[5]
                        nshift += 1
                    if pos9 is not None and pos9 in pos6:
                        hrsPos6 += list9[2]
                        RegHrs6 += list9[3]
                        totOT12_6 += list9[4]
                        totOT40_6 += list9[5]
                        nshift += 1
                list6 = [valName, pos6, hrsPos6, RegHrs6, totOT12_6, totOT40_6, list6[6], list6[7], nshift]
            # Position 7
            hrsPos7 = 0
            RegHrs7 = 0
            totOT12_7 = 0
            totOT40_7 = 0
            nshift = 1
            if pos7 is not None:
                if (    pos6 is not None and pos5 is not None and pos4 is not None and
                        pos3 is not None and pos2 is not None and pos1 is not None and
                        pos7 not in pos1 and pos7 not in pos2 and pos7 not in pos3 and
                        pos7 not in pos4 and pos7 not in pos5 and pos7 not in pos6):
                    hrsPos7 += list7[2]
                    RegHrs7 += list7[3]
                    totOT12_7 += list7[4]
                    totOT40_7 += list7[5]
                    nshift += 1
                    if pos8 is not None and pos8 in pos7:
                        hrsPos7 += list8[2]
                        RegHrs7 += list8[3]
                        totOT12_7 += list8[4]
                        totOT40_7 += list8[5]
                        nshift += 1
                    if pos9 is not None and pos9 in pos7:
                        hrsPos7 += list9[2]
                        RegHrs7 += list9[3]
                        totOT12_7 += list9[4]
                        totOT40_7 += list9[5]
                        nshift += 1
                list7 = [valName, pos7, hrsPos7, RegHrs7, totOT12_7, totOT40_7, list7[6], list7[7], nshift]
            # Position 8
            hrsPos8 = 0
            RegHrs8 = 0
            totOT12_8 = 0
            totOT40_8 = 0
            nshift = 0
            if pos8 is not None:
                if (    pos7 is not None and pos6 is not None and pos5 is not None and
                        pos4 is not None and pos3 is not None and pos2 is not None and
                        pos1 is not None and
                        pos8 not in pos1 and pos8 not in pos2 and pos8 not in pos3 and
                        pos8 not in pos4 and pos8 not in pos5 and pos8 not in pos6 and
                        pos8 not in pos7):
                    hrsPos8 += list8[2]
                    RegHrs8 += list8[3]
                    totOT12_8 += list8[4]
                    totOT40_8 += list8[5]
                    nshift += 1
                    if pos9 is not None and pos9 in pos8:
                        hrsPos8 += list9[2]
                        RegHrs8 += list9[3]
                        totOT12_8 += list9[4]
                        totOT40_8 += list9[5]
                        nshift += 1
                list8 = [valName, pos8, hrsPos8, RegHrs8, totOT12_8, totOT40_8, list8[6], list8[7], nshift]
            # Position 9
            hrsPos9 = 0
            RegHrs9 = 0
            totOT12_9 = 0
            totOT40_9 = 0
            nshift = 1
            if pos9 is not None:
                if (    pos8 is not None and pos7 is not None and pos6 is not None and
                        pos5 is not None and pos4 is not None and pos3 is not None and
                        pos2 is not None and pos1 is not None and
                        pos9 not in pos1 and pos9 not in pos2 and pos9 not in pos3 and
                        pos9 not in pos4 and pos9 not in pos5 and pos9 not in pos6 and
                        pos9 not in pos7 and pos9 not in pos8):
                    hrsPos9 += list9[2]
                    RegHrs9 += list9[3]
                    totOT12_9 += list9[4]
                    totOT40_9 += list9[5]
                    nshift += 1
                list9 = [valName, pos9, hrsPos9, RegHrs9, totOT12_9, totOT40_9, list9[6], list9[7], nshift]

            if (    list2[1] is not None and (list2[1] is list1[1])):
                list2 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list3[1] is not None and (list3[1] is list1[1] or list3[1] is list2[1])):
                list3 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list4[1] is not None and (list4[1] is list1[1] or list4[1] is list2[1] or list4[1] is list3[1])):
                list4 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list5[1] is not None and (list5[1] is list1[1] or list5[1] is list2[1] or list5[1] is list3[1] or
                    list5[1] is list4[1])):
                list5 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list6[1] is not None and (list6[1] is list1[1] or list6[1] is list2[1] or list6[1] is list3[1] or
                    list6[1] is list4[1] or list6[1] is list5[1])):
                list6 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list7[1] is not None and (list7[1] is list1[1] or list7[1] is list2[1] or list7[1] is list3[1] or
                    list7[1] is list4[1] or list7[1] is list5[1] or list7[1] is list6[1])):
                list7 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list8[1] is not None and (list8[1] is list1[1] or list8[1] is list2[1] or list8[1] is list3[1] or
                    list8[1] is list4[1] or list8[1] is list5[1] or list8[1] is list6[1] or list8[1] is list7[1])):
                list8 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if (    list9[1] is not None and (list9[1] is list1[1] or list9[1] is list2[1] or list9[1] is list3[1] or
                    list9[1] is list4[1] or list9[1] is list5[1] or list9[1] is list6[1] or list9[1] is list7[1] or
                    list9[1] is list8[1])):
                list9 = [None, None, 0, 0, 0, 0, 0, None, 0]

            if list8[1] is None and list9[1] is not None:
                list8 = list9
                list9 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list7[1] is None and list8[1] is not None:
                list7 = list8
                list8 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list6[1] is None and list7[1] is not None:
                list6 = list7
                list7 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list5[1] is None and list6[1] is not None:
                list5 = list6
                list6 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list4[1] is None and list5[1] is not None:
                list4 = list5
                list5 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list3[1] is None and list4[1] is not None:
                list3 = list4
                list4 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if list2[1] is None and list3[1] is not None:
                list2 = list3
                list3 = [None, None, 0, 0, 0, 0, 0, None, 0]
            if j == 0: listWeek1 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]
            else:      listWeek2 = [list1, list2, list3, list4, list5, list6, list7, list8, list9]

        if flagDebug2:
            print("\n", valName)
            print("--Week 1 totals--")
            print("    Position:\t\t  total hrs:   reg hrs:   OT12:   OT40:   Rate:")
            for j in range(0, 9):
                if listWeek1[j][0] is not None:
                    if (listWeek1[j][2] != 0. or listWeek1[j][5] != 0.):
                        print(listWeek1[j][1], " ", listWeek1[j][2], " ", listWeek1[j][3], " ", listWeek1[j][4], " ", listWeek1[j][5], " ", listWeek1[j][6])
            print("--Week 2 totals--")
            print("    Position:\t\t  total hrs:   reg hrs:   OT12:   OT40:   Rate:")
            for j in range(0, 9):
                if listWeek2[j][0] is not None:
                    if (listWeek2[j][2] != 0. or listWeek2[j][5] != 0.):
                        print(listWeek2[j][1], " ", listWeek2[j][2], " ", listWeek2[j][3], " ", listWeek2[j][4], " ", listWeek2[j][5], " ", listWeek2[j][6])


        # Print values to .xlsx file
        line = Side(border_style="thin", color="000000")
        c0 = newsheet1.cell(row = printCnt, column = 1)
        c0.value = listWeek1[0][0]
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt, column = 2)
        c0.border = Border(bottom = line)
        c0.value = "Rate"
        c0 = newsheet1.cell(row = printCnt, column = 3)
        c0.value = "Regular"
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt, column = 4)
        c0.value = "OT+12"
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt, column = 5)
        c0.value = "OT+40"
        c0.border = Border(bottom = line)
        c0 = newsheet1.cell(row = printCnt, column = 6)
        c0.value = "OT Total"
        c0.border = Border(bottom = line)
        if listWeek1[0][7] is not None or listWeek1[0][7] == 0:
            c0 = newsheet1.cell(row = printCnt, column = 8)
            c0.value = listWeek1[0][7]
        if flagMultShift1 is True or flagMultShift2 is True:
            c0 = newsheet1.cell(row = printCnt, column = 10)
            c0.value = "Check OT+12"
        c0 = newsheet1.cell(row = printCnt + 1, column = 1)
        c0.value = "---Week 1---"
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0 = newsheet1.cell(row = printCnt + 1, column = 2)
        c0.border = Border(right = line)
        c0 = newsheet1.cell(row = printCnt + 2, column = 1)
        if listWeek1[0][1] is None: c0.value = "No shifts"
        else: c0.value = listWeek1[0][1]
        c0.border = Border(right = line)
        c0 = newsheet1.cell(row = printCnt + 2, column = 2)
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0.value = listWeek1[0][6]
        c0 = newsheet1.cell(row = printCnt + 2, column = 3)
        c0.value = listWeek1[0][3]
        c0 = newsheet1.cell(row = printCnt + 2, column = 4)
        c0.value = listWeek1[0][4]
        c0 = newsheet1.cell(row = printCnt + 2, column = 5)
        c0.value = listWeek1[0][5]
        c0 = newsheet1.cell(row = printCnt + 2, column = 6)
        c0.value = listWeek1[0][4] + listWeek1[0][5]
        c0 = newsheet1.cell(row = printCnt + 2, column = 9)
        if flagMultShift1 is True:
            c0 = newsheet1.cell(row = printCnt + 2, column = 10)
            c0.value = "Multiple shifts in same day for week 1. Check for OT+12 by hand."
        if listWeek1[1][0] is None: printCnt += 4
        if listWeek1[1][0] is not None:
            c0 = newsheet1.cell(row = printCnt + 3, column = 1)
            c0.value = listWeek1[1][1]
            c0.border = Border(right = line)
            c0 = newsheet1.cell(row = printCnt + 3, column = 2)
            c0.border = Border(right = line)
            c0.font = Font(italic = 'single')
            c0.value = listWeek1[1][6]
            c0 = newsheet1.cell(row = printCnt + 3, column = 3)
            c0.value = listWeek1[1][3]
            c0 = newsheet1.cell(row = printCnt + 3, column = 4)
            c0.value = listWeek1[1][4]
            c0 = newsheet1.cell(row = printCnt + 3, column = 5)
            c0.value = listWeek1[1][5]
            c0 = newsheet1.cell(row = printCnt + 3, column = 6)
            c0.value = listWeek1[1][4] + listWeek1[1][5]
            c0 = newsheet1.cell(row = printCnt + 3, column = 9)
            printCnt += 5
            for j in range(2, 9):
                if listWeek1[j][0] is not None:
                    if (listWeek1[j][2] != 0. or listWeek1[j][5] != 0.):
                        c0 = newsheet1.cell(row = printCnt - 1, column = 1)
                        c0.value = listWeek1[j][1]
                        c0.border = Border(right = line)
                        c0 = newsheet1.cell(row = printCnt - 1, column = 2)
                        c0.border = Border(right = line)
                        c0.font = Font(italic = 'single')
                        c0.value = listWeek1[j][6]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 3)
                        c0.value = listWeek1[j][3]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 4)
                        c0.value = listWeek1[j][4]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 5)
                        c0.value = listWeek1[j][5]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 6)
                        c0.value = listWeek1[j][4] + listWeek1[j][5]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 9)
                        printCnt += 1

        c0 = newsheet1.cell(row = printCnt - 1, column = 1)
        c0.value = "---Week 2---"
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0 = newsheet1.cell(row = printCnt - 1, column = 2)
        c0.border = Border(right = line)
        c0 = newsheet1.cell(row = printCnt, column = 1)
        if listWeek2[0][1] is None: c0.value = "No shifts"
        else: c0.value = listWeek2[0][1]
        c0.border = Border(right = line)
        c0 = newsheet1.cell(row = printCnt, column = 2)
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0.value = listWeek2[0][6]
        c0 = newsheet1.cell(row = printCnt, column = 3)
        c0.value = listWeek2[0][3]
        c0 = newsheet1.cell(row = printCnt, column = 4)
        c0.value = listWeek2[0][4]
        c0 = newsheet1.cell(row = printCnt, column = 5)
        c0.value = listWeek2[0][5]
        c0 = newsheet1.cell(row = printCnt, column = 6)
        c0.value = listWeek2[0][4] + listWeek2[0][5]
        c0 = newsheet1.cell(row = printCnt, column = 9)
        if flagMultShift2 is True:
            c0 = newsheet1.cell(row = printCnt, column = 10)
            c0.value = "Multiple shifts in same day for week 2. Check for OT+12 by hand."
        if listWeek2[1][0] is None: printCnt += 2
        if listWeek2[1][0] is not None:
            c0 = newsheet1.cell(row = printCnt + 1, column = 1)
            c0.value = listWeek2[1][1]
            c0.border = Border(right = line)
            c0 = newsheet1.cell(row = printCnt + 1, column = 2)
            c0.border = Border(right = line)
            c0.font = Font(italic = 'single')
            c0.value = listWeek2[1][6]
            c0 = newsheet1.cell(row = printCnt + 1, column = 3)
            c0.value = listWeek2[1][3]
            c0 = newsheet1.cell(row = printCnt + 1, column = 4)
            c0.value = listWeek2[1][4]
            c0 = newsheet1.cell(row = printCnt + 1, column = 5)
            c0.value = listWeek2[1][5]
            c0 = newsheet1.cell(row = printCnt + 1, column = 6)
            c0.value = listWeek2[1][4] + listWeek2[1][5]
            c0 = newsheet1.cell(row = printCnt + 1, column = 9)
            printCnt += 3
            for j in range(2, 9):
                if listWeek2[j][0] is not None:
                    if (listWeek2[j][2] != 0. or listWeek2[j][5] != 0.):
                        c0 = newsheet1.cell(row = printCnt - 1, column = 1)
                        c0.value = listWeek2[j][1]
                        c0.border = Border(right = line)
                        c0 = newsheet1.cell(row = printCnt - 1, column = 2)
                        c0.border = Border(right = line)
                        c0.font = Font(italic = 'single')
                        c0.value = listWeek2[j][6]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 3)
                        c0.value = listWeek2[j][3]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 4)
                        c0.value = listWeek2[j][4]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 5)
                        c0.value = listWeek2[j][5]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 6)
                        c0.value = listWeek2[j][4] + listWeek2[j][5]
                        c0 = newsheet1.cell(row = printCnt - 1, column = 9)
                        printCnt += 1

        # Add up both weeks
        xreg = 0
        xOT = 0
        c0 = newsheet1.cell(row = printCnt - 1, column = 1)
        c0.value = "---Total (reg, OT, shifts)---"
        c0.border = Border(right = line)
        c0.font = Font(italic = 'single')
        c0 = newsheet1.cell(row = printCnt - 1, column = 2)
        c0.border = Border(right = line)
        for j in range(0, 9):
            if listWeek1[j][1] is not None:
                if (listWeek1[j][2] != 0. or listWeek1[j][5] != 0.):
                    xreg = listWeek1[j][3]
                    xOT  = listWeek1[j][4] + listWeek1[j][5]
                    xshift = listWeek1[j][8]
                    if listWeek2[0][1] is not None and listWeek2[0][1] in listWeek1[j][1]:
                        xreg += listWeek2[0][3]
                        xOT  += listWeek2[0][4] + listWeek2[0][5]
                        xshift += listWeek2[0][8]
                    if listWeek2[1][1] is not None and listWeek2[1][1] in listWeek1[j][1]:
                        xreg += listWeek2[1][3]
                        xOT  += listWeek2[1][4] + listWeek2[1][5]
                        xshift += listWeek2[1][8]
                    if listWeek2[2][1] is not None and listWeek2[2][1] in listWeek1[j][1]:
                        xreg += listWeek2[2][3]
                        xOT  += listWeek2[2][4] + listWeek2[2][5]
                        xshift += listWeek2[2][8]
                    if listWeek2[3][1] is not None and listWeek2[3][1] in listWeek1[j][1]:
                        xreg += listWeek2[3][3]
                        xOT  += listWeek2[3][4] + listWeek2[3][5]
                        xshift += listWeek2[3][8]
                    if listWeek2[4][1] is not None and listWeek2[4][1] in listWeek1[j][1]:
                        xreg += listWeek2[4][3]
                        xOT  += listWeek2[4][4] + listWeek2[4][5]
                        xshift += listWeek2[4][8]
                    if listWeek2[5][1] is not None and listWeek2[5][1] in listWeek1[j][1]:
                        xreg += listWeek2[5][3]
                        xOT  += listWeek2[5][4] + listWeek2[5][5]
                        xshift += listWeek2[5][8]
                    if listWeek2[6][1] is not None and listWeek2[6][1] in listWeek1[j][1]:
                        xreg += listWeek2[6][3]
                        xOT  += listWeek2[6][4] + listWeek2[6][5]
                        xshift += listWeek2[6][8]
                    if listWeek2[7][1] is not None and listWeek2[7][1] in listWeek1[j][1]:
                        xreg += listWeek2[7][3]
                        xOT  += listWeek2[7][4] + listWeek2[7][5]
                        xshift += listWeek2[7][8]
                    if listWeek2[8][1] is not None and listWeek2[8][1] in listWeek1[j][1]:
                        xreg += listWeek2[8][3]
                        xOT  += listWeek2[8][4] + listWeek2[8][5]
                        xshift += listWeek2[8][8]
                    if xshift == 0: xshift = 1
                    c0 = newsheet1.cell(row = printCnt, column = 1)
                    c0.value = listWeek1[j][1]
                    c0.border = Border(right = line)
                    c0 = newsheet1.cell(row = printCnt, column = 2)
                    c0.border = Border(right = line)
                    c0.font = Font(italic = 'single')
                    c0.value = listWeek1[j][6]
                    c0 = newsheet1.cell(row = printCnt, column = 3)
                    c0.font = Font(bold = 'single')
                    c0.value = xreg
                    c0 = newsheet1.cell(row = printCnt, column = 4)
                    c0.font = Font(bold = 'single')
                    c0.value = xOT
                    c0 = newsheet1.cell(row = printCnt, column = 6)
                    c0.font = Font(italic = 'single')
                    c0.value = xshift
                    printCnt += 1
        xreg = 0
        xOT = 0
        for j in range(0, 9):
            if listWeek2[j][1] is not None:
                if (listWeek2[j][2] != 0. or listWeek2[j][5] != 0.):
                    if   (listWeek1[0][1] is not None and listWeek2[j][1].__contains__(listWeek1[0][1]) ): continue
                    elif (listWeek1[1][1] is not None and listWeek2[j][1].__contains__(listWeek1[1][1]) ): continue
                    elif (listWeek1[2][1] is not None and listWeek2[j][1].__contains__(listWeek1[2][1]) ): continue
                    elif (listWeek1[3][1] is not None and listWeek2[j][1].__contains__(listWeek1[3][1]) ): continue
                    elif (listWeek1[4][1] is not None and listWeek2[j][1].__contains__(listWeek1[4][1]) ): continue
                    elif (listWeek1[5][1] is not None and listWeek2[j][1].__contains__(listWeek1[5][1]) ): continue
                    elif (listWeek1[6][1] is not None and listWeek2[j][1].__contains__(listWeek1[6][1]) ): continue
                    elif (listWeek1[7][1] is not None and listWeek2[j][1].__contains__(listWeek1[7][1]) ): continue
                    elif (listWeek1[8][1] is not None and listWeek2[j][1].__contains__(listWeek1[8][1]) ): continue
                    else:
                        xreg = listWeek2[j][3]
                        xOT  = listWeek2[j][4] + listWeek2[j][5]
                        xshift = listWeek2[j][8]
                        if xshift == 0: xshift = 1
                        c0 = newsheet1.cell(row = printCnt, column = 1)
                        c0.value = listWeek2[j][1]
                        c0.border = Border(right = line)
                        c0 = newsheet1.cell(row = printCnt, column = 2)
                        c0.border = Border(right = line)
                        c0.font = Font(italic = 'single')
                        c0.value = listWeek2[j][6]
                        c0 = newsheet1.cell(row = printCnt, column = 3)
                        c0.font = Font(bold = 'single')
                        c0.value = xreg
                        c0 = newsheet1.cell(row = printCnt, column = 4)
                        c0.font = Font(bold = 'single')
                        c0.value = xOT
                        c0 = newsheet1.cell(row = printCnt, column = 6)
                        c0.font = Font(italic = 'single')
                        c0.value = xshift
                        printCnt += 1
        printCnt += 2
        rowCnt = 1 # Reset count parameter for next employee

# Check for empty rows
for i in range(1, Nrow):
    c0 = newsheet1.cell(row = i, column = 1).value
    c1 = newsheet1.cell(row = i, column = 1).value
    c2 = newsheet1.cell(row = i, column = 3).value
    c3 = newsheet1.cell(row = i, column = 4).value
    c4 = newsheet1.cell(row = i, column = 5).value
    c5 = newsheet1.cell(row = i, column = 6).value
    if c0 is c1 and c2 == 0 and c3 == 0 and c4 == 0 and c5 == 0:
        newsheet1.delete_rows(i, 1)


# More headers and design options for output
c0 = newsheet1.cell(row = 1, column = 1)
c0.value = "Week 1 dates:  "
c0 = newsheet1.cell(row = 1, column = 2)
c0.value = week1start.strftime("%Y-%m-%d") + "  to  " + week1end.strftime("%Y-%m-%d")
c0 = newsheet1.cell(row = 2, column = 1)
c0.value = "Week 2 dates:  "
c0 = newsheet1.cell(row = 2, column = 2)
c0.value = week2start.strftime("%Y-%m-%d") + "  to  " + week2end.strftime("%Y-%m-%d")
c0 = newsheet1.cell(row = 3, column = 1)
c0.value = "Total number of employees:  "
c0 = newsheet1.cell(row = 3, column = 2)
c0.value = GrandNames
c0 = newsheet1.cell(row = 4, column = 1)
c0.value = "Total number of shifts:  "
c0 = newsheet1.cell(row = 4, column = 2)
c0.value = Nrow - 1
c0 = newsheet1.cell(row = 5, column = 1)
c0.value = "Total number of hours:  "
c0 = newsheet1.cell(row = 5, column = 2)
c0.value = GrandHrs
c0 = newsheet1.cell(row = 7, column = 8)
c0.value = "Employee Number"
c0 = newsheet1.cell(row = 7, column = 10)
c0.value = "Notes"

c0 = newsheet1.cell(row = 1, column = 1)
c0.alignment = Alignment(horizontal='right')
c0 = newsheet1.cell(row = 2, column = 1)
c0.alignment = Alignment(horizontal='right')
c0 = newsheet1.cell(row = 3, column = 1)
c0.alignment = Alignment(horizontal='right')
c0 = newsheet1.cell(row = 4, column = 1)
c0.alignment = Alignment(horizontal='right')
c0 = newsheet1.cell(row = 5, column = 1)
c0.alignment = Alignment(horizontal='right')
newsheet1.column_dimensions["A"].width = 48
newsheet1.column_dimensions["B"].width = 8
newsheet1.column_dimensions["C"].width = 8
newsheet1.column_dimensions["D"].width = 8

output_name = "output_detail.xlsx"
newbook1.save(output_name)
print("\n\n")
if flagMultShiftTot is True: print("====================================================")
if flagMultShiftTot is True: print("---WARNING! Hand-check days with multiple shifts!---")
if flagMultShiftTot is True: print("====================================================\n")
print("\nOT calculation done.\n")
print("file output written to", output_name,"\n")
